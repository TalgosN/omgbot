import requests
import json
from datetime import datetime, timedelta
import pytz
import locale
import pandas as pd
import math
from collections import defaultdict
from copy import copy
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook
from telebot import *
import os
import sqlite3
from constants import *
from sender import safe_send
from permissions import ROLE_OWNER, require_role

locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')



########## FT

headers_ft = {'Accept': 'application/json',
  'Content-Type': 'application/json',
           'Authorization':f"Bearer {FT_API_KEY}"}
           


params_ft = {
  "type": "nal" 
}

########## AQSI
headers = {'Accept': 'application/json',
  'Content-Type': 'application/json',
           'x-client-key':f"Application {AQSI_API_KEY}"}
           
           

############################## Some consts

pay_method = {0: "Нал",1:"Карта", 2:"QR"}

content_type = { 1: "Приход",
                 2:"Возврат прихода",
                 3:"Расход",
                 4:"Возврат расхода"}


op_type = {'Приход':'income', "Возврат прихода":'outcome'}

ft_cat = {
    "Кофе": "Снеки",
    "Снеки": "Снеки",
    "Квест": "Квесты",
    "Автосим": "Автосим",
    "Абонемент": "Абонемент",
    "Классика" : "Классика",
    "Майнкрафт": "Майнкрафт",
    "Мероприятие": "Мероприятие",
    "Лаунж": "Лаунж",
    "Сертификат в конверте": "Сертификаты",
    "Сертификат в коробке": "Сертификаты"}

def is_difference_one_day(date_str1, date_str2):
    # Преобразуем строки в объекты datetime
    date_format = "%Y-%m-%dT%H:%M:%S"
    date1 = datetime.strptime(date_str1, date_format)
    date2 = datetime.strptime(date_str2, date_format)
    
    # Вычисляем разницу между датами
    difference = abs(date1 - date2)
    
    # Проверяем, равна ли разница 1 суткам
    return difference == timedelta(days=1)
    
bdays_rate =500  
################################
def define_goods():
    response_goods_groups = requests.request("GET", f'https://api.aqsi.ru/pub/v2/GoodsCategory/list', headers=headers, timeout=15)

    response_dict_goods_groups = response_goods_groups.json()

    response_goods = requests.request("GET", f'https://api.aqsi.ru/pub/v2/Goods/list', headers=headers, timeout=15)

    response_goods = response_goods.json()
    
    return response_dict_goods_groups, response_goods
    
################################ Bot Enterpoint  
def finance(message,bot):
    if not require_role(message, bot, ROLE_OWNER):
        return
    bot.send_message(message.chat.id, f'Этот раздел посвящен финансам')
    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add(*funclist_fin)
    bot.send_message(message.chat.id, 'Что вы хотите сделать? 👀', reply_markup=markup)
    bot.register_next_step_handler(message, func_fin,bot)

def returnback(message,bot):
        from menu import hello
        hello (message.chat.id,bot)

import traceback 

def func_fin(message, bot):
    if not require_role(message, bot, ROLE_OWNER):
        return
    if message.text in ['📑 Отчет по приходам', '💸 Внести приходы по наличке', '💰 Инкассация', '👨🏻‍💻 ЗП за период', '📊 Сводный отчет']:
        operation = message.text
        markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        
        # Для инкассации логичнее дни, для остального - периоды
        if operation == '💰 Инкассация':
            markup.add('Вчера', 'Сегодня', '⬅️ Вернуться')
        else:
            markup.add('Текущий месяц', 'Прошлый месяц')
            markup.add('Текущая неделя', 'Вчера')
            markup.add('⬅️ Вернуться')
            
        bot.send_message(message.chat.id, 'Выбери быстрый период или введи дату начала вручную (01.01.2025)', reply_markup=markup)
        bot.register_next_step_handler(message, handle_start, bot, operation)
        
    elif message.text == '👀 Сверка финансов':
        try:
            text = check_cash(datetime.now(pytz.timezone('Europe/Moscow')).strftime('%d.%m.%Y'))
            if len(text) >= 4096:
                TEXTS['ui']['login_logout'] = text[:4096]
                TEXTS['ui']['readiness'] = text[4096:]
                bot.send_message(message.chat.id, TEXTS['ui']['login_logout'])
                bot.send_message(message.chat.id, TEXTS['ui']['readiness'])
            else:
                bot.send_message(message.chat.id, text)
            finance(message, bot)
        except Exception as er:
            import traceback
            error_details = traceback.format_exc()
            bot.send_message(message.chat.id, f"Ошибка: {er}\n\nДетали:\n{error_details}")
            finance(message, bot)
            
    elif message.text == '⬅️ Вернуться':
        returnback(message, bot)
    else:
        finance(message, bot)


def handle_start(message, bot, operation):
    if not require_role(message, bot, ROLE_OWNER):
        return
    if message.text == '⬅️ Вернуться':
        finance(message, bot)
        return

    quick_ranges = ['Текущий месяц', 'Прошлый месяц', 'Текущая неделя', 'Вчера', 'Сегодня']
    
    if message.text in quick_ranges:
        tz = pytz.timezone('Europe/Moscow')
        today = datetime.now(tz).replace(hour=0, minute=0, second=0, microsecond=0)
        
        if message.text == 'Сегодня':
            d_start = today
            d_end = today + timedelta(days=1)
        elif message.text == 'Вчера':
            d_start = today - timedelta(days=1)
            d_end = today
        elif message.text == 'Текущая неделя':
            d_start = today - timedelta(days=today.weekday())
            d_end = today + timedelta(days=1)
        elif message.text == 'Текущий месяц':
            d_start = today.replace(day=1)
            d_end = today + timedelta(days=1)
        elif message.text == 'Прошлый месяц':
            d_end = today.replace(day=1)
            d_start = (d_end - timedelta(days=1)).replace(day=1)
        
        date_start = d_start.strftime("%Y-%m-%dT%H:%M:%S")
        date_end = d_end.strftime("%Y-%m-%dT%H:%M:%S")
        
        # Скрываем клавиатуру и запускаем сразу сбор данных
        bot.send_message(message.chat.id, f"Выбран период: с {d_start.strftime('%d.%m.%Y')} по {(d_end - timedelta(days=1)).strftime('%d.%m.%Y')}. Собираю данные...", reply_markup=telebot.types.ReplyKeyboardRemove())
        execute_fin_operation(operation, date_start, date_end, message, bot)
    else:
        # Старая логика ручного ввода
        try:
            dt = datetime.strptime(message.text, '%d.%m.%Y')
            date_start = dt.strftime("%Y-%m-%dT%H:%M:%S")
            
            markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
            markup.add('⬅️ Вернуться')
            bot.send_message(message.chat.id, 'Введи дату конца отcчета в формате 01.02.2025', reply_markup=markup)
            bot.register_next_step_handler(message, handle_end, date_start, bot, operation)
        except Exception:
            markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
            markup.add('⬅️ Вернуться')
            bot.send_message(message.chat.id, 'Что-то пошло не так!\nВведи дату начала отcчета в формате 01.01.2025', reply_markup=markup)
            bot.register_next_step_handler(message, handle_start, bot, operation)

def handle_end(message, date_start, bot, operation):
    if not require_role(message, bot, ROLE_OWNER):
        return
    if message.text == '⬅️ Вернуться':
        finance(message, bot)
    else:
        try:
            dt = datetime.strptime(message.text, '%d.%m.%Y')
            dt = dt + timedelta(days=1)
            date_end = dt.strftime("%Y-%m-%dT%H:%M:%S")
            
            bot.send_message(message.chat.id, "Собираю данные...", reply_markup=telebot.types.ReplyKeyboardRemove())
            execute_fin_operation(operation, date_start, date_end, message, bot)
            
        except Exception as er:
            markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
            markup.add('⬅️ Вернуться')
            bot.send_message(message.chat.id, f'Что-то пошло не так!\n{er}\nВведи дату конца отcчета:', reply_markup=markup)
            bot.register_next_step_handler(message, handle_end, date_start, bot, operation)

def execute_fin_operation(operation, date_start, date_end, message, bot):
    """Вынесенный роутер операций для того, чтобы не дублировать код в handle_start и handle_end"""
    if not require_role(message, bot, ROLE_OWNER):
        return
    try:
        if operation == '📑 Отчет по приходам':
            create_otchet(date_start, date_end, message, bot)
            
        elif operation == '💰 Инкассация':
            if is_difference_one_day(date_start, date_end):
                inkass(date_start, date_end, message, bot)
            else:
                markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
                markup.add('⬅️ Вернуться')
                bot.send_message(message.chat.id, 'Для проверки списаний разница между датами должна быть ровно сутки.\nВведи дату начала:', reply_markup=markup)
                bot.register_next_step_handler(message, handle_start, bot, operation)
                
        elif operation == '💸 Внести приходы по наличке':
            nal_to_dt(date_start, date_end, message, bot)
            
        elif operation == '👨🏻‍💻 ЗП за период':
            pay_report(date_start, date_end, message, bot)
            
        elif operation == '📊 Сводный отчет':
            generate_summary_report(date_start, date_end, message, bot)
            
    except Exception as er:
        bot.send_message(message.chat.id, f"Ошибка при формировании отчета: {er}")
        finance(message, bot)
        
################################ Excel Transactions
def create_otchet (start_dt,end_dt,message,bot):

    data, raw_data, errors = create_data(start_dt,end_dt,message,bot)

    file = open("./Reports/Errors_Report.txt", "w")
    file.write(errors)
    file.close()         
    
    file = open("./Reports/Raw_Data.txt", "w")
    file.write(raw_data)
    file.close()

    df = pd.DataFrame(data)

    # Save the DataFrame to an Excel file
 
    dt = datetime.strptime(start_dt,"%Y-%m-%dT%H:%M:%S")
    date_start_str = dt.strftime('%d.%m.%Y')

    dt = datetime.strptime(end_dt,"%Y-%m-%dT%H:%M:%S")
    date_end_str = dt.strftime('%d.%m.%Y')

    template_path = './Reports/Шаблон.xlsx'  # путь к вашему шаблону
    output_path = f'./Reports/Отчет_{date_start_str}-{date_end_str}.xlsx'  # имя для сохранения



    # Загружаем книгу и выбираем лист
    wb = load_workbook(template_path)
    ws = wb['Данные']  # или wb['SheetName'], если знаете имя листа

    # Записываем заголовки из DataFrame в первую строку
    for c_idx, column_name in enumerate(df.columns):
        ws.cell(row=1, column=c_idx + 1, value=column_name)  # Записываем заголовки в первую строку

    # Записываем данные в нужные ячейки, начиная со второй строки
    for r_idx, row in df.iterrows():
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)  # +2 для пропуска заголовков


  
    # Удаляем последние пустые строки
    max_row = ws.max_row
    while ws[max_row][0].value is None and max_row > 1:# Проверяем только строки после заголовка
        ws.delete_rows(max_row)
        max_row -= 1
      
    # Сохраняем новый файл
    wb.save(output_path)
    # Сохраняем новый файл
  
  
    
    #bot.send_message(message.chat.id, 'Готово!')

    doc = open(output_path, 'rb')
    bot.send_document(message.chat.id, doc)

    if os.stat("./Reports/Errors_Report.txt").st_size == 0:
        bot.send_message(message.chat.id, 'Все прошло без ошибок!')
    else:
        bot.send_message(message.chat.id, 'В ходе формирования отcчета были ошибки!')
        er_doc = open("./Reports/Errors_Report.txt", 'rb')
        bot.send_document(message.chat.id, er_doc)
   
    finance(message,bot)

################################ Data from AQSI

def create_data (start_dt,end_dt,message,bot):
    # Кассы
    response_dev = requests.request("GET", f'https://api.aqsi.ru/pub/v3/Devices', headers=headers)

    response_dict_dev = response_dev.json()


    # Магазины
    response_shop = requests.request("GET", f'https://api.aqsi.ru/pub/v2/Shops/list', headers=headers)

    response_dict_shops = response_shop.json()

    errors = ''      
    data = []
    raw_data=''
  
    response_dict_goods_groups, response_goods = define_goods()
  
    params = {"filtered.beginDate": start_dt,
              "filtered.endDate": end_dt}


    response = requests.request("GET", f'https://api.aqsi.ru/pub/v2/Receipts', headers=headers, params = params)

    response_dict = response.json()

    pages = response_dict['pages']
    message_bot = None

    for j in range (pages):

        params = {"filtered.beginDate": start_dt,
                 "filtered.endDate": end_dt,
                 'page': j}
  
        response = requests.request("GET", f'https://api.aqsi.ru/pub/v2/Receipts', headers=headers, params = params)
        response_dict = response.json() 

        raw_data=f'{raw_data} {json.dumps(response_dict, indent=4, sort_keys=True, ensure_ascii=False)}\n'

        if j==0:
            message_bot = bot.send_message (message.chat.id, f'{math.floor(j/pages*100)} %')
        else:
            bot.edit_message_text(chat_id=message.chat.id, message_id=message_bot.message_id, text=f'{math.floor(j/pages*100)} %')
    
        for i in response_dict['rows']:
            try:
                dt = datetime.fromisoformat(i['processedAt'])
                date_pay = dt.strftime("%d.%m.%Y")
                time_pay = dt.strftime("%H:%M:%S")

                payments = i['content']['checkClose']['payments']
                positions = i['content']['positions']
    
                payment_index = 0
                remaining_payment = payments[payment_index]['amount']
    
                for t in range(len(positions)):
                    item_price = positions[t]['price']
                    item_quantity = positions[t]['quantity']
                    item_total = item_price * item_quantity
        
                    while item_total > 0 and payment_index < len(payments):
                        if remaining_payment > 0:
                            # Определяем, сколько можем оплатить
                            payment_amount = min(remaining_payment, item_total)
                
                            row_data = {}
                            row_data['Тип'] = content_type[i['content']['type']]
                            row_data['Дата'] = date_pay
                            row_data['Время'] = time_pay
                        
                            pay_type = i['content']['checkClose']['payments'][payment_index]['type']
                            
                            if pay_type ==1:
                                pay_type = 2 if i['content']['checkClose']['payments'][payment_index]["acquiringData"]["apn"]=="" else 1
          
                            row_data ['Метод оплаты'] = pay_method[pay_type]
                            

                            dev_id_pay = i ['deviceSN']
                        
                            for v in response_dict_dev['rows']:
                                if v['serialNumber']==dev_id_pay:
                                    shop_id=v['shop']['id']
                                    break
      
                            for b in response_dict_shops:
                                if b['id']==shop_id:
                                    row_data['Место'] = b['name']
                                    break

                            row_data['Сумма платежа'] = payment_amount
                
                            # Заполняем данные о позиции
                            row_data[f'Позиция'] = positions[t]['text']

                            goodId = positions[t]['externalId']
                
                            for n in response_goods['rows']:
                                if n['id'] == goodId:
                                    good_group_id = n['group_id']
                                    break
                
                            for b in response_dict_goods_groups:
                                if b['id'] == good_group_id:
                                    good_group = b['name']
                                    break

                            row_data[f'Категория позиции'] = good_group.strip()
                
                            data.append(row_data)
                
                            # Обновляем суммы
                            remaining_payment -= payment_amount
                            item_total -= payment_amount
            
                        # Если текущий платеж исчерпан, переходим к следующему
                        if remaining_payment <= 0:
                            payment_index += 1
                            if payment_index < len(payments):
                                remaining_payment = payments[payment_index]['amount']
    
                # Запись ошибок
            except Exception as e:
                errors=f'{errors}{e}{i.get("id", "Unknown")}\n'
                continue
                
    if message_bot:
        bot.delete_message(message.chat.id, message_bot.message_id)
    
    return data, raw_data, errors

################################ Инкассы, изъятия

def inkass (start_dt,end_dt,message,bot):
    
    params = {
      "filtered.beginDate": start_dt,
      
      "filtered.endDate": end_dt}
      
    # Кассы
    response_dev = requests.request("GET", f'https://api.aqsi.ru/pub/v3/Devices', headers=headers)

    response_dict_dev = response_dev.json()


    # Магазины
    response_shop = requests.request("GET", f'https://api.aqsi.ru/pub/v2/Shops/list', headers=headers)

    response_dict_shops = response_shop.json()

    # Определения числа страниц по всему запросу (Смены)
    response = requests.request("GET", f'https://api.aqsi.ru/pub/v2/Shifts', headers=headers, params = params)

    response_dict = response.json()
    pages = response_dict['pages']


    data_check, raw_data, errors = create_data(start_dt,end_dt,message,bot)     
    data_cash = []
    raw_data_cash=''

    # Обработка всех страниц и строк

    for j in range (pages):

      params = {
      "filtered.beginDate": start_dt,
      
      "filtered.endDate": end_dt,

      'page': j}
      
      response = requests.request("GET", f'https://api.aqsi.ru/pub/v2/Shifts', headers=headers, params = params)
      response_dict = response.json() 

      raw_data_cash=f'{raw_data_cash} {json.dumps(response_dict, indent=4, sort_keys=True, ensure_ascii=False)}\n'
      
      for i in response_dict['rows']:
          
          dict_cash = {}
          dict_cash['cash_start']= i['cashAtStart']
          dict_cash['cash_end'] = i['cashAtEnd']
          
          dev_id = i['deviceSN']
          
          for t in response_dict_dev['rows']:
              if t['serialNumber']==dev_id:
                  shop_id=t['shop']['id']
                  break
          
          for b in response_dict_shops:
              if b['id']==shop_id:
                  dict_cash['shop'] = b['name']
                  break
          data_cash.append(dict_cash)


    
    inkass =[]
    for i in data_cash:
        row_inkass = {}
        row_inkass['Место']=i['shop']
        
        #if i['cash_end']==i['cash_start']:
            #continue
          
        for j in data_check:
          
            if i['shop']==j['Место'] and j['Метод оплаты']=='Нал':
              
                if j['Тип']=='Приход':
                    
                    i["cash_start"]= float(i['cash_start'])+float(j['Сумма платежа']) 
                    
                    
                    
                elif j['Тип']=='Возврат прихода':
                  
                    i["cash_start"]= float(i['cash_start'])-float(j['Сумма платежа'])
                    
        
        row_inkass['Был инкасс?']=float(i['cash_end'])!=float(i['cash_start'])
        row_inkass['Сумма инкасса']=float(i['cash_end'])-float(i['cash_start'])
        inkass.append(row_inkass)
        
        
        

    text_inkass = 'По моим данным на указанную дату были следующие изменения в балансе:\n\n'
    
    for i in inkass:
        text_inkass=f'{text_inkass}{i["Место"]}: {i["Сумма инкасса"]}\n\n'
    
    bot.send_message(message.chat.id, text_inkass)
    
    file = open("./Reports/Errors_Inkass.txt", "w")
    file.write(errors)
    file.close()  
    
    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Да','Нет')
    
    if os.stat("./Reports/Errors_Inkass.txt").st_size == 0:
        bot.send_message(message.chat.id, 'Все прошло без ошибок! Желаете внести в ФТ?',reply_markup=markup)
    else:
        bot.send_message(message.chat.id, 'В ходе формирования отcчета были ошибки! Желаете внести в ФТ?',reply_markup=markup)
        er_doc = open("./Reports/Errors_Inkass.txt", 'rb')
        bot.send_document(message.chat.id, er_doc)
    
    bot.register_next_step_handler(message,confirm_inkass,inkass,bot,start_dt)
    
def confirm_inkass(message,inkass,bot,start_dt):
    if not require_role(message, bot, ROLE_OWNER):
        return
    if message.text=='Нет':
        finance(message,bot)
    elif message.text=='Да':
        insert_inkass(message,inkass, bot,start_dt)
    else:
        markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        markup.add('Да','Нет')
        bot.send_message(message.chat.id, 'Желаете внести в ФТ?',reply_markup=markup)
        bot.register_next_step_handler(message,confirm_inkass,inkass,bot,start_dt)


def insert_inkass(message,inkass,bot,start_dt):
    if not require_role(message, bot, ROLE_OWNER):
        return
    
    date_ins = datetime.strptime(start_dt,"%Y-%m-%dT%H:%M:%S")
    
    response_mb = requests.request("GET", f'https://api.fintablo.ru/v1/moneybag', headers=headers_ft, params = params_ft)

    moneybags = response_mb.json()



    #############################################################


    #############################################################
    response_direct = requests.request("GET", f'https://api.fintablo.ru/v1/direction', headers=headers_ft)

    directs = response_direct.json()

    #############################################################
    
    flag = 1
    errors = ''
    for k in inkass:
        if k['Был инкасс?']:

          amount =abs(k['Сумма инкасса'])
          date_ins2 = date_ins.strftime('%d.%m.%Y')
          
          desc = 'Инкассация/списание' if k['Сумма инкасса']<0 else 'Пополнение/выравнивание'
          #cat = 'Инкассация'
          club = f'Нал {k["Место"]}'
          
          group = 'outcome' if k['Сумма инкасса']<0 else 'income'
          
          for j in moneybags['items']:
            
              if j['name']==club:
                  club_id =j['id']
                  break
          
            
          for p in directs['items']:
              if p['name']==k["Место"]:
                  id_direct = p['id']
                  break
                  
          payload = json.dumps({
                                "value": amount,
                                "moneybagId": club_id,
                                "group": group,
                                "description": desc,
                                #"categoryId": id_cat,
                                "directionId":id_direct,
                                "date": date_ins2
                                })

          response2 = requests.request("POST", f'https://api.fintablo.ru/v1/transaction', headers=headers_ft, data = payload)

          response_dict2 = response2.json()
          if response_dict2['status']==200:
            flag = flag
          else:
            flag=0
            dic = {
                                "value": amount,
                                "moneybagId": club_id,
                                "group": group,
                                "description": desc,
                                #"categoryId": id_cat,
                                "directionId":id_direct,
                                "date": date_ins2
                                }
            
            errors = f'{errors}{response_dict2["status"]}{dic}\n\n'
    
    if flag ==1:
        bot.send_message(message.chat.id, 'Все прошло успешно!')
    else:
        bot.send_message(message.chat.id, 'Некоторые операции не были внесены! Представлю их ниже')
        bot.send_message(message.chat.id, errors)
    
    finance(message,bot)        
################################ Проверка финансов
def check_cash (date_check):

    

    # Кассы
    response_dev = requests.request("GET", f'https://api.aqsi.ru/pub/v3/Devices', headers=headers)

    response_dict_dev = response_dev.json()


    # Магазины
    response_shop = requests.request("GET", f'https://api.aqsi.ru/pub/v2/Shops/list', headers=headers)

    response_dict_shops = response_shop.json()
    
    #ФТ
    response_moneybags = requests.request("GET", f'https://api.fintablo.ru/v1/moneybag', headers=headers_ft, params = params_ft)
    moneybags = response_moneybags.json()

    
    start_dt = datetime.strptime(date_check,'%d.%m.%Y')
    
    end_dt=start_dt
    start_dt = start_dt-timedelta(days=1)
    
    #БД
    db_date = start_dt.strftime("%Y-%m-%d")
    
    end_dt=end_dt.strftime("%Y-%m-%dT%H:%M:%S")
    start_dt = start_dt.strftime("%Y-%m-%dT%H:%M:%S")
    
    
    
    conn=sqlite3.connect('db/omgbot.sql')
    cur = conn.cursor()
    cur.execute("SELECT club, amount FROM nal WHERE drep=?", (db_date,))
    cash_by_admin = cur.fetchall()
    cur.close()
    conn.close()
    
    #Акси
    params = {
      "filtered.beginDate": start_dt,
      
      "filtered.endDate": end_dt}
      
    response = requests.request("GET", f'https://api.aqsi.ru/pub/v2/Shifts', headers=headers, params = params)

    response_dict = response.json()
    
    data_cash= []
    print (response_dict_shops)  
    for i in response_dict['rows']:
        dict_cash = {}
        
        dict_cash['Акси'] = i['cashAtEnd']
      
        dev_id = i['deviceSN']
      
        for t in response_dict_dev['rows']:
            if t['serialNumber']==dev_id:
                shop_id=t['shop']['id']
                break
      
        for b in response_dict_shops:
            if b['id']==shop_id:
                dict_cash['Клуб'] = b['name']
                break
                
        for j in cash_by_admin:
            if j[0]==dict_cash['Клуб']:
                dict_cash['БД']=j[1]
                break
                
        for t in moneybags['items']:
          if t['name'].replace('Нал ','')==dict_cash['Клуб']:
              dict_cash['ФТ']=t['balance']
              break
            
        data_cash.append(dict_cash)
    
      
    text=f'Сверка баланса на {date_check}\n\n'
    for i in data_cash:

        text=f'{text}{i["Клуб"]}\n'
        text=f'{text}ФТ: {i["ФТ"]}\n'
        text=f'{text}БД: {i["БД"]}\n'
        text=f'{text}Акси: {i["Акси"]}\n\n'
        
        if i["Акси"] is None:
            text=f'{text}Не удалось получить данные из Акси!!\n\n'
        
        elif i["ФТ"] is None:
            text=f'{text}Не удалось получить данные из ФТ!!\n\n'
        
        elif i["БД"] is None:
            text=f'{text}Не удалось получить данные из БД!!\n\n'
        
        else:
            if float(i["ФТ"])>float(i["Акси"]):

                text=f'{text}Возможно, ты уже перенес операции за сегодня по ФТ!\n\n'

            elif float(i["ФТ"]) <float(i["Акси"]):

                text=f'{text}Вероятно, не перенесены все транзакции в ФТ\n\n'

            elif float(i["БД"])!=float(i["ФТ"]) and float(i["БД"]!=i["Акси"]):

                text=f'{text}Администратор неверно подсчитал кассу!\n\n'

            elif float(i["БД"])==float(i["ФТ"]) and float(i["БД"])==float(i["Акси"]):

                text=f'{text}Ура, все сходится!\n\n'
          
    return f'{text}'
    
    
def nal_to_dt (start_dt,end_dt,message,bot):
    response_mb = requests.request("GET", f'https://api.fintablo.ru/v1/moneybag', headers=headers_ft, params = params_ft)

    moneybags = response_mb.json()



    #############################################################
    response_ft_cat = requests.request("GET", f'https://api.fintablo.ru/v1/category', headers=headers_ft)

    cats_ft = response_ft_cat.json()

    #############################################################
    response_direct = requests.request("GET", f'https://api.fintablo.ru/v1/direction', headers=headers_ft)

    directs = response_direct.json()

    #############################################################
    data, raw_data, errors = create_data(start_dt,end_dt,message,bot)
    flag = 1
    errors = ''
    for k in data:
        if k['Метод оплаты']=='Нал':

          amount =k['Сумма платежа']
          date = k['Дата']
          time = k['Время']
          group = op_type[k['Тип']]
          
          if group=='income':
              if k['Категория позиции']=='VR':
                  cat = ft_cat[k['Позиция']]
              else:
                  cat = k['Категория позиции']
          elif group=='outcome':
              cat = 'Возврат'

              
          
          
          club = f'Нал {k["Место"]}'
          
          
          for j in moneybags['items']:
            
              if j['name']==club:
                  club_id =j['id']
                  break
          
          for i in cats_ft['items']:
              if i['name']==cat:
                  id_cat = i['id']
                  break
            
          for p in directs['items']:
              if p['name']==k["Место"]:
                  id_direct = p['id']
                  break
                  
          payload = json.dumps({
                                "value": amount,
                                "moneybagId": club_id,
                                "group": group,
                                "description": cat,
                                "categoryId": id_cat,
                                "directionId":id_direct,
                                "date": date
                                })

          response2 = requests.request("POST", f'https://api.fintablo.ru/v1/transaction', headers=headers_ft, data = payload)

          response_dict2 = response2.json()
          if response_dict2['status']==200:
            flag = flag
          else:
            flag=0
            dic = {
                                "value": amount,
                                "moneybagId": club_id,
                                "group": group,
                                "description": cat,
                                "categoryId": id_cat,
                                "directionId":id_direct,
                                "date": date
                                }
            
            errors = f'{errors}{response_dict2["status"]} {dic}\n\n'
    
    if flag ==1:
        bot.send_message(message.chat.id, 'Все прошло успешно!')
    else:
        bot.send_message(message.chat.id, 'Некоторые операции не были внесены! Представлю их ниже')
        bot.send_message(message.chat.id, errors)
    
    finance(message,bot)

################################ Pay Report



PAYROLL_DB_PATH = 'db/omgbot.sql'
PAYROLL_TEMPLATE_PATH = './Reports/Шаблон_ЗП.xlsx'
PAYROLL_EMPLOYEE_COLUMNS = 30
PAYROLL_SPECIAL_ROWS = ('Двойные без клуба', 'Автосим', 'Активации')


def _canonical_login(login):
    return str(login or '').strip().lower()


def _decimal(value):
    return Decimal(str(value or 0))


def _money(value):
    return float(value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))


def _parse_payroll_period(date_start, date_end):
    start_dt = datetime.strptime(date_start, '%Y-%m-%d %H:%M:%S')
    end_dt = datetime.strptime(date_end, '%Y-%m-%d %H:%M:%S')
    if end_dt <= start_dt:
        raise ValueError('Дата конца отчёта должна быть позже даты начала')
    return start_dt, end_dt


def _fetch_missing_payroll_shifts(conn, start_dt, end_dt):
    """Дополняет только полностью отсутствующие в БД даты через OMG Shift."""
    start_iso = start_dt.date().isoformat()
    end_iso = end_dt.date().isoformat()
    existing_dates = {
        row[0] for row in conn.execute(
            """SELECT DISTINCT date(dt_shift)
               FROM shifts
               WHERE date(dt_shift) >= date(?) AND date(dt_shift) < date(?)""",
            (start_iso, end_iso),
        )
    }
    location_names = {
        club['source_name']: club['name'] for club in get_schedule_locations()
    }
    current_date = start_dt.date()
    end_date = end_dt.date()

    while current_date < end_date:
        date_iso = current_date.isoformat()
        current_date += timedelta(days=1)
        if date_iso in existing_dates:
            continue

        response = requests.get(
            f"{SHIFTON_API_URL}/api/bot/schedule?date={date_iso}",
            headers={"Authorization": f"Bearer {SHIFTON_API_TOKEN}"},
            timeout=10,
        )
        response.raise_for_status()
        payload = response.json()
        if not isinstance(payload, dict) or not payload.get('ok'):
            error = payload.get('error', 'invalid_response') if isinstance(payload, dict) else 'invalid_response'
            raise RuntimeError(f'OMG Shift не вернул расписание за {date_iso}: {error}')

        rows = []
        seen = set()
        for location in payload.get('locations', []):
            source_club = location.get('title', 'Неизвестно')
            club = location_names.get(source_club, source_club)
            for shift in location.get('shifts', []):
                employee = str(shift.get('employee') or '').strip()
                if not employee or employee.upper() == 'СВОБОДНАЯ СМЕНА':
                    continue

                start_time = datetime.strptime(shift['start'], '%H:%M')
                end_time = datetime.strptime(shift['end'], '%H:%M')
                if end_time <= start_time:
                    end_time += timedelta(days=1)
                duration = round((end_time - start_time).total_seconds() / 3600, 1)

                name_parts = employee.split()
                second_name = name_parts[0]
                first_name = name_parts[1] if len(name_parts) > 1 else ''
                login = str(shift.get('telegram') or '').strip() or None
                if login and not login.startswith('@'):
                    login = f'@{login}'

                row_key = (
                    second_name, first_name, club, shift['start'], shift['end'],
                    _canonical_login(login),
                )
                if row_key in seen:
                    continue
                seen.add(row_key)
                rows.append((second_name, first_name, date_iso, club, duration, login))

        if rows:
            conn.executemany(
                """INSERT INTO shifts
                   (shift_second_name, shift_first_name, dt_shift, club, dur, shift_login, source)
                   VALUES (?, ?, ?, ?, ?, ?, 'omg_shift')""",
                rows,
            )
            conn.commit()


def _load_payroll_rates(conn, start_dt, end_dt):
    table_exists = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name='payroll_rates'"
    ).fetchone()
    if not table_exists:
        raise RuntimeError('Таблица ставок payroll_rates не создана. Перезапусти бота и импортируй ставки')

    rates = defaultdict(list)
    for login, club, hourly_rate, valid_from, valid_to in conn.execute(
        """SELECT login, club, hourly_rate, valid_from, valid_to
           FROM payroll_rates
           WHERE date(valid_from) < date(?)
             AND (valid_to IS NULL OR date(valid_to) >= date(?))
           ORDER BY date(valid_from) DESC""",
        (end_dt.date().isoformat(), start_dt.date().isoformat()),
    ):
        rates[(_canonical_login(login), club)].append(
            (datetime.strptime(valid_from, '%Y-%m-%d').date(),
             datetime.strptime(valid_to, '%Y-%m-%d').date() if valid_to else None,
             _decimal(hourly_rate))
        )
    return rates


def _find_payroll_rate(rates, login, club, shift_date):
    for rate_club in (club, '*'):
        for valid_from, valid_to, hourly_rate in rates.get((_canonical_login(login), rate_club), []):
            if valid_from <= shift_date and (valid_to is None or shift_date <= valid_to):
                return hourly_rate
    return None


def get_data_pay_report(date_start, date_end):
    start_dt, end_dt = _parse_payroll_period(date_start, date_end)
    start_iso = start_dt.date().isoformat()
    end_iso = end_dt.date().isoformat()
    conn = sqlite3.connect(PAYROLL_DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        _fetch_missing_payroll_shifts(conn, start_dt, end_dt)
        rates = _load_payroll_rates(conn, start_dt, end_dt)

        users = {}
        user_columns = {row[1] for row in conn.execute('PRAGMA table_info(users)')}
        if {'second_name', 'first_name'} <= user_columns:
            user_rows = conn.execute(
                "SELECT login, second_name, first_name FROM users WHERE login IS NOT NULL"
            )
            for row in user_rows:
                login = _canonical_login(row['login'])
                full_name = ' '.join(filter(None, (row['second_name'], row['first_name']))).strip()
                users[login] = full_name or row['login']
        elif 'name' in user_columns:
            for row in conn.execute("SELECT login, name FROM users WHERE login IS NOT NULL"):
                users[_canonical_login(row['login'])] = row['name'] or row['login']

        data = {}

        def employee_data(login, fallback_name=''):
            login_key = _canonical_login(login)
            if not login_key:
                raise RuntimeError(f'У смены сотрудника {fallback_name or "без имени"} не указан Telegram-логин')
            if login_key not in data:
                data[login_key] = {
                    'Логин': login,
                    'Имя': fallback_name or users.get(login_key) or login,
                    'Клубы': defaultdict(Decimal),
                    'Двойные без клуба': Decimal('0'),
                    'Автосим': Decimal('0'),
                    'Активации': Decimal('0'),
                }
            return data[login_key]

        shifts = conn.execute(
            """SELECT shift_second_name, shift_first_name, date(dt_shift) AS shift_date,
                      club, dur, shift_login
               FROM shifts
               WHERE date(dt_shift) >= date(?) AND date(dt_shift) < date(?)""",
            (start_iso, end_iso),
        ).fetchall()

        shift_clubs = defaultdict(set)
        missing_rates = set()
        for shift in shifts:
            fallback_name = ' '.join(filter(None, (
                shift['shift_second_name'], shift['shift_first_name']
            ))).strip()
            employee = employee_data(shift['shift_login'], fallback_name)
            shift_date = datetime.strptime(shift['shift_date'], '%Y-%m-%d').date()
            login_key = _canonical_login(shift['shift_login'])
            rate = _find_payroll_rate(rates, login_key, shift['club'], shift_date)
            if rate is None:
                missing_rates.add((employee['Имя'], employee['Логин'], shift['club'], shift['shift_date']))
                continue
            employee['Клубы'][shift['club']] += _decimal(shift['dur']) * rate
            shift_clubs[(login_key, shift['shift_date'])].add(shift['club'])

        doubles = conn.execute(
            """SELECT who, date(d_rep) AS report_date, amount
               FROM double
               WHERE date(d_rep) >= date(?) AND date(d_rep) < date(?)""",
            (start_iso, end_iso),
        ).fetchall()
        for item in doubles:
            employee = employee_data(item['who'])
            login_key = _canonical_login(item['who'])
            item_date = datetime.strptime(item['report_date'], '%Y-%m-%d').date()
            clubs = shift_clubs.get((login_key, item['report_date']), set())
            club = next(iter(clubs)) if len(clubs) == 1 else None
            rate = _find_payroll_rate(rates, login_key, club or '*', item_date)
            if rate is None:
                missing_rates.add((employee['Имя'], employee['Логин'], club or 'без клуба', item['report_date']))
                continue
            amount = _decimal(item['amount']) * rate
            if club:
                employee['Клубы'][club] += amount
            else:
                employee['Двойные без клуба'] += amount

        birthdays = conn.execute(
            """SELECT who, club, COUNT(DISTINCT ID) AS amount
               FROM birthday
               WHERE date(dt_rep) >= date(?) AND date(dt_rep) < date(?)
                 AND status = 'Одобрено'
               GROUP BY who, club""",
            (start_iso, end_iso),
        ).fetchall()
        for item in birthdays:
            employee_data(item['who'])['Клубы'][item['club']] += _decimal(item['amount']) * _decimal(bdays_rate)

        for table, field in (('autosim', 'Автосим'), ('activation', 'Активации')):
            rows = conn.execute(
                f"""SELECT who, SUM(amount) AS amount FROM {table}
                    WHERE date(d_rep) >= date(?) AND date(d_rep) < date(?)
                    GROUP BY who""",
                (start_iso, end_iso),
            ).fetchall()
            for item in rows:
                employee_data(item['who'])[field] += _decimal(item['amount'])

        if missing_rates:
            details = ', '.join(
                f'{name} ({login}), {club}, {report_date}'
                for name, login, club, report_date in sorted(missing_rates)[:10]
            )
            if len(missing_rates) > 10:
                details += f' и ещё {len(missing_rates) - 10}'
            raise RuntimeError(f'Не найдены ставки: {details}')

        for employee in data.values():
            employee['Клубы'] = {
                club: _money(amount) for club, amount in employee['Клубы'].items()
            }
            for field in PAYROLL_SPECIAL_ROWS:
                employee[field] = _money(employee[field])
        return data
    finally:
        conn.close()


def _find_excel_row(ws, label):
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == label:
            return row
    raise RuntimeError(f'В шаблоне зарплаты не найдена строка «{label}»')


def _copy_excel_row_style(ws, source_row, target_row):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def _prepare_payroll_rows(ws, row_count):
    anchor_row = _find_excel_row(ws, 'Возмещение / Доп')
    available_rows = anchor_row - 2
    if row_count > available_rows:
        extra_rows = row_count - available_rows
        ws.insert_rows(anchor_row, amount=extra_rows)
        for row in range(anchor_row, anchor_row + extra_rows):
            _copy_excel_row_style(ws, 2, row)
        anchor_row += extra_rows

    for row in range(2, anchor_row):
        for column in range(1, 32):
            ws.cell(row=row, column=column).value = None
        ws.cell(row=row, column=32, value=f'=SUM(B{row}:AE{row})')
        ws.cell(row=row, column=34, value=f'=SUM(B{row}:AE{row})-AG{row}')
    return anchor_row


def _write_payroll_formulas(ws):
    payment_row = _find_excel_row(ws, 'К оплате:')
    comment_row = _find_excel_row(ws, 'Комментарий')
    for label in ('Возмещение / Доп', 'Штраф', 'Бонус KPI', 'ОФ'):
        row = _find_excel_row(ws, label)
        ws.cell(row, 32, value=f'=SUM(B{row}:AE{row})')
        ws.cell(row, 34, value=f'=SUM(B{row}:AE{row})')
    for column in range(2, 32):
        letter = ws.cell(row=1, column=column).column_letter
        ws.cell(payment_row, column, value=f'=SUM({letter}1:{letter}{payment_row - 1})')
    ws.cell(payment_row, 32, value=f'=SUM(B{payment_row}:AE{payment_row})')
    ws.cell(payment_row, 34, value=f'=AF{payment_row}-AG{comment_row}')
    ws.cell(comment_row, 32, value=f'=SUM(AF2:AF{payment_row - 1})')
    ws.cell(comment_row, 33, value=f'=SUM(AG2:AG{payment_row - 1})')
    ws.cell(comment_row, 34, value=f'=SUM(AH2:AH{payment_row - 1})')


def pay_report(date_start, date_end, message, bot):
    start_dt = datetime.strptime(date_start, '%Y-%m-%dT%H:%M:%S')
    end_dt = datetime.strptime(date_end, '%Y-%m-%dT%H:%M:%S')
    data = get_data_pay_report(
        start_dt.strftime('%Y-%m-%d %H:%M:%S'),
        end_dt.strftime('%Y-%m-%d %H:%M:%S'),
    )
    if not data:
        raise RuntimeError('За выбранный период нет смен и начислений')

    employees = sorted(data.values(), key=lambda item: (item['Имя'].lower(), item['Логин'].lower()))
    if len(employees) > PAYROLL_EMPLOYEE_COLUMNS:
        raise RuntimeError(f'Шаблон зарплаты поддерживает не более {PAYROLL_EMPLOYEE_COLUMNS} сотрудников')

    duplicate_names = defaultdict(int)
    for employee in employees:
        duplicate_names[employee['Имя']] += 1
    column_names = [
        employee['Имя'] if duplicate_names[employee['Имя']] == 1
        else f"{employee['Имя']} ({employee['Логин']})"
        for employee in employees
    ]

    used_clubs = {club for employee in employees for club in employee['Клубы']}
    configured_clubs = [club for club in get_clubs() if club in used_clubs]
    clubs = configured_clubs + sorted(used_clubs - set(configured_clubs), key=str.lower)
    report_rows = clubs + list(PAYROLL_SPECIAL_ROWS)

    wb = load_workbook(PAYROLL_TEMPLATE_PATH)
    ws = wb['ЗП']
    _prepare_payroll_rows(ws, len(report_rows))

    for column in range(2, 2 + PAYROLL_EMPLOYEE_COLUMNS):
        ws.cell(row=1, column=column).value = None
    for column, name in enumerate(column_names, start=2):
        ws.cell(row=1, column=column, value=name)

    for row, row_name in enumerate(report_rows, start=2):
        ws.cell(row=row, column=1, value=row_name)
        for column, employee in enumerate(employees, start=2):
            value = employee['Клубы'].get(row_name, 0) if row_name in clubs else employee[row_name]
            ws.cell(row=row, column=column, value=value)
            ws.cell(row=row, column=column).number_format = '#,##0.00'

    _write_payroll_formulas(ws)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    for column in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value is not None), default=0)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    start_label = start_dt.strftime('%d.%m.%y')
    end_label = (end_dt - timedelta(days=1)).strftime('%d.%m.%y')
    output_path = f'./Reports/Отчет_ЗП_{start_label}-{end_label}.xlsx'
    wb.save(output_path)

    with open(output_path, 'rb') as document:
        bot.send_document(message.chat.id, document)
    finance(message, bot)


################################ Сводка
def format_money(amount):
    """Красивое форматирование денег: 15000.5 -> 15 000.50"""
    return f"{amount:,.2f}".replace(',', ' ')

def generate_summary_report(start_dt, end_dt, message, bot):
    # Получаем данные через твою готовую функцию
    data, _, _ = create_data(start_dt, end_dt, message, bot)
    
    total_income = 0
    by_method = {}
    by_club = {}
    by_category = {}
    
    for row in data:
        amount = row['Сумма платежа']
        # Учитываем только чистый приход (Приход минус Возвраты)
        if row['Тип'] == 'Возврат прихода':
            amount = -amount
        elif row['Тип'] != 'Приход':
            continue
            
        total_income += amount
        
        # Группировка по методам оплаты
        method = row['Метод оплаты']
        by_method[method] = by_method.get(method, 0) + amount
        
        # Группировка по клубам
        club = row.get('Место', 'Неизвестно')
        by_club[club] = by_club.get(club, 0) + amount
        
        # Группировка по категориям
        cat = row.get('Категория позиции', 'Без категории')
        by_category[cat] = by_category.get(cat, 0) + amount

    # Формируем красивый текст
    dt_s = datetime.strptime(start_dt, "%Y-%m-%dT%H:%M:%S").strftime("%d.%m.%Y")
    dt_e = (datetime.strptime(end_dt, "%Y-%m-%dT%H:%M:%S") - timedelta(days=1)).strftime("%d.%m.%Y")
    
    text = f"📊 <b>Сводный отчет ({dt_s} - {dt_e})</b>\n\n"
    text += f"💰 <b>ОБЩАЯ СУММА:</b> {format_money(total_income)} р.\n\n"
    
    text += "💳 <b>МЕТОДЫ ОПЛАТЫ:</b>\n"
    for m, val in sorted(by_method.items(), key=lambda x: x[1], reverse=True):
        text += f" • {m}: {format_money(val)} р.\n"
        
    text += "\n🏠 <b>ПО КЛУБАМ:</b>\n"
    for c, val in sorted(by_club.items(), key=lambda x: x[1], reverse=True):
        text += f" • {c}: {format_money(val)} р.\n"
        
    text += "\n🛒 <b>ПО КАТЕГОРИЯМ:</b>\n"
    for cat, val in sorted(by_category.items(), key=lambda x: x[1], reverse=True):
        text += f" • {cat}: {format_money(val)} р.\n"

    bot.send_message(message.chat.id, text, parse_mode='HTML')
    finance(message, bot)

def get_aqsi_data_silent(start_dt, end_dt):
    """Тихая функция получения данных для фоновых задач (без bot.send_message)"""
    response_dict_goods_groups, response_goods = define_goods()
    
    # Оптимизация поиска: делаем плоские словари для мгновенного доступа O(1)
    goods_to_group = {n['id']: n.get('group_id') for n in response_goods.get('rows', []) if 'id' in n}
    group_to_name = {b['id']: b['name'].strip() for b in response_dict_goods_groups if 'id' in b}
    
    response_dev = requests.request("GET", f'https://api.aqsi.ru/pub/v3/Devices', headers=headers, timeout=15).json()
    response_shop = requests.request("GET", f'https://api.aqsi.ru/pub/v2/Shops/list', headers=headers, timeout=15).json()
    
    params = {"filtered.beginDate": start_dt, "filtered.endDate": end_dt}
    response = requests.request("GET", f'https://api.aqsi.ru/pub/v2/Receipts', headers=headers, params=params, timeout=15).json()
    
    if 'pages' not in response: return []
    
    data = []
    for j in range(response['pages']):
        params['page'] = j
        page_resp = requests.request("GET", f'https://api.aqsi.ru/pub/v2/Receipts', headers=headers, params=params, timeout=15).json()
        
        for i in page_resp.get('rows', []):
            try:
                payments = i['content']['checkClose']['payments']
                positions = i['content']['positions']
                payment_index = 0
                remaining_payment = payments[payment_index]['amount']
                
                for t in range(len(positions)):
                    item_total = positions[t]['price'] * positions[t]['quantity']
                    while item_total > 0 and payment_index < len(payments):
                        if remaining_payment > 0:
                            payment_amount = min(remaining_payment, item_total)
                            
                            # Определяем метод оплаты
                            pay_type = payments[payment_index]['type']
                            if pay_type == 1:
                                pay_type = 2 if payments[payment_index].get("acquiringData", {}).get("apn") == "" else 1
                            method_name = pay_method.get(pay_type, "Неизвестно")
                            
                            # Определяем категорию позиции
                            g_id = positions[t]['externalId']
                            group_id = goods_to_group.get(g_id)
                            category_name = group_to_name.get(group_id, "Без категории")
                            
                            row_data = {
                                'Тип': content_type[i['content']['type']],
                                'Сумма платежа': payment_amount,
                                'Метод оплаты': method_name,
                                'Категория позиции': category_name
                            }
                            
                            dev_id_pay = i['deviceSN']
                            shop_id = next((v['shop']['id'] for v in response_dev.get('rows', []) if v['serialNumber'] == dev_id_pay), None)
                            row_data['Место'] = next((b['name'] for b in response_shop if b['id'] == shop_id), 'Неизвестно')
                            
                            data.append(row_data)
                            remaining_payment -= payment_amount
                            item_total -= payment_amount
                        if remaining_payment <= 0:
                            payment_index += 1
                            if payment_index < len(payments):
                                remaining_payment = payments[payment_index]['amount']
            except Exception:
                continue
    return data

def calc_diff(current, previous):
    """Вычисляет процентную разницу и возвращает красивую строку со стрелочкой"""
    if previous == 0:
        return " (📈 +100%)" if current > 0 else ""
    diff = ((current - previous) / previous) * 100
    if diff > 0:
        return f" (📈 +{diff:.1f}%)"
    elif diff < 0:
        return f" (📉 {diff:.1f}%)"
    return " (➖ 0%)"

def auto_weekly_report(bot, target_chat_id=None):
    """Отправка сводного отчета в канал по понедельникам"""
    tz = pytz.timezone('Europe/Moscow')
    today = datetime.now(tz).replace(hour=0, minute=0, second=0, microsecond=0)
    
    last_week_start = today - timedelta(days=today.weekday() + 7)
    last_week_end = last_week_start + timedelta(days=7)
    
    prev_week_start = last_week_start - timedelta(days=7)
    prev_week_end = last_week_start

    data_last = get_aqsi_data_silent(last_week_start.strftime("%Y-%m-%dT%H:%M:%S"), last_week_end.strftime("%Y-%m-%dT%H:%M:%S"))
    data_prev = get_aqsi_data_silent(prev_week_start.strftime("%Y-%m-%dT%H:%M:%S"), prev_week_end.strftime("%Y-%m-%dT%H:%M:%S"))
    
    def get_totals(data_set):
        total = 0
        clubs = {}
        methods = {}
        categories = {}
        for r in data_set:
            amt = r['Сумма платежа']
            if r['Тип'] == 'Возврат прихода': amt = -amt
            elif r['Тип'] != 'Приход': continue
            
            total += amt
            
            c = r.get('Место', 'Неизвестно')
            clubs[c] = clubs.get(c, 0) + amt
            
            m = r.get('Метод оплаты', 'Неизвестно')
            methods[m] = methods.get(m, 0) + amt
            
            cat = r.get('Категория позиции', 'Без категории')
            categories[cat] = categories.get(cat, 0) + amt
        return total, clubs, methods, categories

    total_last, clubs_last, methods_last, categories_last = get_totals(data_last)
    total_prev, clubs_prev, methods_prev, categories_prev = get_totals(data_prev)

    text = f"#финансы\n\n📊 <b>Еженедельный фин. отчет ({last_week_start.strftime('%d.%m')} - {(last_week_end-timedelta(days=1)).strftime('%d.%m')})</b>\n\n"
    text += f"💰 <b>ВЫРУЧКА:</b> {format_money(total_last)} р.{calc_diff(total_last, total_prev)}\n\n"
    
    text += "💳 <b>МЕТОДЫ ОПЛАТЫ:</b>\n"
    for m, val in sorted(methods_last.items(), key=lambda x: x[1], reverse=True):
        val_prev = methods_prev.get(m, 0)
        text += f" • {m}: {format_money(val)} р.{calc_diff(val, val_prev)}\n"
        
    text += "\n🏠 <b>ПО КЛУБАМ:</b>\n"
    for c, val in sorted(clubs_last.items(), key=lambda x: x[1], reverse=True):
        val_prev = clubs_prev.get(c, 0)
        text += f" • {c}: {format_money(val)} р.{calc_diff(val, val_prev)}\n"
        
    text += "\n🛒 <b>ПО КАТЕГОРИЯМ:</b>\n"
    for cat, val in sorted(categories_last.items(), key=lambda x: x[1], reverse=True):
        val_prev = categories_prev.get(cat, 0)
        text += f" • {cat}: {format_money(val)} р.{calc_diff(val, val_prev)}\n"
        
    text += "\n<i>* Динамика указана в сравнении с позапрошлой неделей.</i>"
    
    try:
        from constants import CHATS
        target = target_chat_id if target_chat_id is not None else CHATS['reports']
        safe_send(bot, target, text, photo=None, parse_mode='HTML')
    except Exception as e:
        print(f"Ошибка отправки еженедельного отчета: {e}")
