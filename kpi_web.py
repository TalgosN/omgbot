import calendar
import hashlib
import hmac
import html
import io
import json
import os
import random
import re
import sqlite3
import threading
import time
from datetime import UTC, date, datetime, timedelta
from functools import wraps
from urllib.parse import parse_qsl
from zoneinfo import ZoneInfo

import telebot
from flask import Flask, g, jsonify, request, send_file, send_from_directory
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from bukza import (
    BUKZA_CLUB_CODES,
    active_orders_for_day,
    booking_freshness,
    initialize_bukza_schema,
    upcoming_unpaid_orders,
)
from constants import CHATS, TELEGRAM_API_KEY, TEXTS, extra_tags, get_clubs, tags_main
from group_membership import is_main_group_member
from kpi_calculator import (
    active_kpi_employee_logins,
    add_penalty,
    calculate_daily_kpi_series,
    calculate_monthly_kpi,
    cancel_penalty,
    disable_custom_goal,
    get_hashtag_entries,
    get_hashtag_summaries,
    get_month_status,
    get_metric_entries,
    get_kpi_freshness,
    get_kpi_settings,
    initialize_kpi_calculation_schema,
    initialize_shift_time_schema,
    list_penalties,
    save_kpi_settings,
    save_custom_goal,
    set_month_status,
    set_monthly_stream,
)
from permissions import (
    ACTIVE_ROLES,
    ROLE_MANAGER,
    ROLE_NAMES,
    ROLE_OWNER,
    ROLE_TECHNICIAN,
    get_user,
)
from openclose import (
    initialize_club_status_dashboard_schema,
    refresh_club_status_dashboard,
)
from repair_catalog import (
    ZONE_COUNTS,
    add_repair_event,
    catalog_payload,
    create_repair_case,
    equipment_list_payload,
    equipment_payload,
    initialize_repair_schema,
    migration_review_payload,
    repair_payload,
    replace_equipment_unit,
)
from shift_config_store import (
    ACTIONS as SHIFT_ACTIONS,
    MAX_PHOTO_QUESTIONS,
    get_editor_config,
    list_versions,
    rollback_version,
    save_editor_config,
)
from task_notifications import (
    BOT_TASK_TYPE,
    GENERAL_TASK_TYPE,
    REPAIR_TASK_TYPE,
    created_task_notification,
    progress_task_notification,
)
from task_analytics import (
    build_task_analytics,
    build_task_report,
    format_task_report_html,
    initialize_task_analytics_schema,
    record_task_event,
    task_activity_payload,
    task_actor_snapshot,
)


DB_PATH = 'db/omgbot.sql'
STATIC_DIR = os.path.join(os.path.dirname(__file__), 'kpi_static')
AUTH_MAX_AGE_SECONDS = int(os.getenv('KPI_WEBAPP_AUTH_MAX_AGE', '86400'))
OMG_SHIFT_URL = os.getenv(
    'OMG_SHIFT_URL',
    'http://31.129.109.167/?page=settings',
).strip()
CAMERA_TEST_RECIPIENT_CHAT_ID = str(
    os.getenv('CAMERA_TEST_RECIPIENT_CHAT_ID') or CHATS.get('me') or ''
).strip()
CAMERA_TEST_COOLDOWN_SECONDS = 8
ANALYTICS_CACHE_SECONDS = 60
KPI_CACHE_SECONDS = 60
_analytics_cache = {}
_analytics_cache_lock = threading.Lock()
_kpi_cache = {}
_kpi_cache_lock = threading.Lock()
_membership_bot = None
_membership_bot_lock = threading.Lock()
_camera_test_sent_at = {}
_camera_test_lock = threading.Lock()
_shift_report_test_sent_at = {}
_shift_report_test_lock = threading.Lock()

app = Flask(__name__, static_folder=None)
app.config['MAX_CONTENT_LENGTH'] = 22 * 1024 * 1024
initialize_bukza_schema(DB_PATH)
initialize_shift_time_schema(DB_PATH)


def _clear_analytics_cache():
    with _analytics_cache_lock:
        _analytics_cache.clear()
    with _kpi_cache_lock:
        _kpi_cache.clear()


def update_table_open():
    from sheets import update_table_open as sync_update_table_open

    return sync_update_table_open()


def _validate_month(value):
    raw = str(value or '').strip()
    try:
        return datetime.strptime(raw, '%Y-%m').date().replace(day=1).isoformat()
    except ValueError as error:
        raise ValueError('Месяц должен быть в формате YYYY-MM') from error


def _validate_day(value, month):
    raw = str(value or '').strip()
    try:
        selected = datetime.strptime(raw, '%Y-%m-%d').date()
    except ValueError as error:
        raise ValueError('Дата должна быть в формате YYYY-MM-DD') from error
    if selected.strftime('%Y-%m') != month[:7]:
        raise ValueError('Выбранная дата должна относиться к выбранному месяцу')
    return selected.isoformat()


def _default_day(month):
    month_start = datetime.strptime(month, '%Y-%m-%d').date()
    today = date.today()
    if (today.year, today.month) == (month_start.year, month_start.month):
        return today.isoformat()
    return month_start.replace(
        day=calendar.monthrange(month_start.year, month_start.month)[1],
    ).isoformat()


KPI_EXPORT_MONTH_NAMES = (
    'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
    'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь',
)
KPI_EXPORT_BIRTHDAY_RATE = 500


def _kpi_export_workbook(rows, month):
    month_start = datetime.strptime(month, '%Y-%m-%d').date()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (
        f'KPI {KPI_EXPORT_MONTH_NAMES[month_start.month - 1]} '
        f'{month_start.year}'
    )
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = 'A3'
    sheet.merge_cells('C1:E1')

    sheet['A1'] = 'Сотрудник'
    sheet['B1'] = 'Смены'
    sheet['C1'] = 'Итого'
    sheet['A2'] = 'Ник'
    sheet['B2'] = 'По 6 ч'
    sheet['C2'] = '%'
    sheet['D2'] = 'Рейтинг'
    sheet['E2'] = 'ДРшки'

    yellow = PatternFill('solid', fgColor='FFE27A')
    green = PatternFill('solid', fgColor='B7D7AE')
    pale_green = PatternFill('solid', fgColor='DCEAD8')
    first_place = PatternFill('solid', fgColor='8EC37D')
    zone_fills = {
        '🟢': PatternFill('solid', fgColor='C7E2B8'),
        '🟡': PatternFill('solid', fgColor='FFF0BD'),
        '🔴': PatternFill('solid', fgColor='F3C4C6'),
        '⚪': PatternFill('solid', fgColor='E8E8E8'),
    }
    thin = Side(style='thin', color='202020')
    medium = Side(style='medium', color='202020')
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_border = Border(left=medium, right=medium, top=medium, bottom=thin)
    header_font = Font(name='Comfortaa', size=11, bold=True, color='111111')
    data_font = Font(name='Comfortaa', size=11, color='202020')

    for cell in sheet[1]:
        cell.fill = yellow if cell.column == 1 else green
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border
    for cell in sheet[2]:
        cell.fill = yellow if cell.column == 1 else pale_green
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = data_border

    participants = sorted(
        (row for row in rows if float(row.get('shifts') or 0) > 0),
        key=lambda row: (
            row.get('rank') is None,
            row.get('rank') or 0,
            str(row.get('nickname') or '').casefold(),
        ),
    )
    for row_number, employee in enumerate(participants, start=3):
        values = (
            employee.get('nickname') or employee.get('login') or '—',
            float(employee.get('shifts') or 0),
            float(employee.get('total_pct') or 0),
            employee.get('rank'),
            int(round(
                float(employee.get('birthdays') or 0)
                * KPI_EXPORT_BIRTHDAY_RATE
            )),
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.font = data_font
            cell.border = data_border
            cell.alignment = Alignment(
                horizontal='left' if column == 1 else 'right',
                vertical='center',
            )
            if employee.get('rank') == 1:
                cell.fill = first_place
            elif column == 1:
                cell.fill = zone_fills.get(employee.get('zone'), yellow)
        sheet.cell(row=row_number, column=2).number_format = '0'
        sheet.cell(row=row_number, column=3).number_format = '0%'
        sheet.cell(row=row_number, column=4).number_format = '0'
        sheet.cell(row=row_number, column=5).number_format = '# ##0'
        sheet.row_dimensions[row_number].height = 24

    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 22
    for column, width in {'A': 25, 'B': 11, 'C': 11, 'D': 12, 'E': 14}.items():
        sheet.column_dimensions[column].width = width
    last_row = max(2, len(participants) + 2)
    sheet.auto_filter.ref = f'A2:E{last_row}'
    sheet.print_area = f'A1:E{last_row}'
    sheet.page_setup.orientation = 'portrait'
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    return workbook


def _task_report_workbook(report):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = 'Сводка'
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet.merge_cells('A1:D1')
    summary_sheet['A1'] = 'Доска проблем OMG VR'
    summary_sheet['A2'] = 'Период'
    summary_sheet['B2'] = report['period']['label']
    summary_sheet['A3'] = 'Сформировано'
    summary_sheet['B3'] = datetime.fromisoformat(
        report['generated_at']
    ).strftime('%d.%m.%Y %H:%M')
    summary_rows = (
        ('Закрыто за период', report['summary']['completed']),
        ('Создано за период', report['summary']['created']),
        (
            'Средний первый ответ',
            _task_duration_label(
                report['summary']['average_first_response_seconds'],
                report['summary']['first_response_precision'],
            ),
        ),
        (
            'Среднее полное решение',
            _task_duration_label(
                report['summary']['average_resolution_seconds'],
                report['summary']['resolution_precision'],
            ),
        ),
        ('Сейчас в работе', report['summary']['work']),
        ('На проверке', report['summary']['review']),
        ('Всего незакрытых', report['summary']['open']),
    )
    for row_number, values in enumerate(summary_rows, start=5):
        summary_sheet.cell(row=row_number, column=1, value=values[0])
        summary_sheet.cell(row=row_number, column=2, value=values[1])
    closed_start = 14
    closed_headers = (
        'Закрытые по клубам', 'Количество', 'Первый ответ', 'Полное решение',
    )
    for column, value in enumerate(closed_headers, start=1):
        summary_sheet.cell(row=closed_start, column=column, value=value)
    for row_number, club in enumerate(
        report['closed_clubs'], start=closed_start + 1,
    ):
        summary_sheet.cell(row=row_number, column=1, value=club['label'])
        summary_sheet.cell(row=row_number, column=2, value=club['count'])
        summary_sheet.cell(
            row=row_number,
            column=3,
            value=_task_duration_label(
                club['average_first_response_seconds'],
                club['first_response_precision'],
            ),
        )
        summary_sheet.cell(
            row=row_number,
            column=4,
            value=_task_duration_label(
                club['average_resolution_seconds'],
                club['resolution_precision'],
            ),
        )

    open_start = closed_start + len(report['closed_clubs']) + 3
    summary_sheet.cell(row=open_start, column=1, value='Незакрытые по клубам')
    summary_sheet.cell(row=open_start, column=2, value='Количество')
    for row_number, club in enumerate(
        report['open_clubs'], start=open_start + 1,
    ):
        summary_sheet.cell(row=row_number, column=1, value=club['label'])
        summary_sheet.cell(row=row_number, column=2, value=club['open'])

    purple = PatternFill('solid', fgColor='7030A0')
    thin = Side(style='thin', color='D7C6E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    summary_sheet['A1'].fill = purple
    summary_sheet['A1'].font = Font(
        name='Comfortaa', color='FFFFFF', bold=True, size=16,
    )
    summary_sheet['A1'].alignment = Alignment(horizontal='center')
    for row in range(2, 12):
        for column in range(1, 5):
            cell = summary_sheet.cell(row=row, column=column)
            cell.border = border
            cell.font = Font(name='Comfortaa', size=10)
            if column == 1:
                cell.font = Font(name='Comfortaa', size=10, bold=True)
    for header_row in (closed_start, open_start):
        for cell in summary_sheet[header_row]:
            cell.fill = purple
            cell.font = Font(name='Comfortaa', color='FFFFFF', bold=True)
    for row in summary_sheet.iter_rows(min_row=closed_start + 1):
        for cell in row:
            cell.border = border
            if not cell.font.bold:
                cell.font = Font(name='Comfortaa', size=10)
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 22
    summary_sheet.column_dimensions['C'].width = 22
    summary_sheet.column_dimensions['D'].width = 22

    tasks_sheet = workbook.create_sheet('Закрытые')
    tasks_sheet.sheet_view.showGridLines = False
    tasks_sheet.freeze_panes = 'A2'
    headers = (
        'ID', 'Создано', 'Закрыто', 'Клуб', 'Тип', 'Название', 'Описание',
        'Первый ответ', 'Полное решение', 'Возвраты', 'Итоговое решение',
    )
    tasks_sheet.append(headers)
    for cell in tasks_sheet[1]:
        cell.fill = purple
        cell.font = Font(name='Comfortaa', color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    for task in report['rows']:
        tasks_sheet.append((
            task['id'],
            _display_date(task['date']),
            _display_date(task['closed_at']),
            _excel_text(task['club']),
            _excel_text(task['type']),
            _excel_text(task['title']),
            _excel_text(task['description']),
            _task_duration_label(
                task['first_response_seconds'], task['first_response_precision'],
            ),
            _task_duration_label(
                task['resolution_seconds'], task['resolution_precision'],
            ),
            task['return_count'],
            _excel_text(task['final_solution'] or 'Итог не записан'),
        ))
    for row in tasks_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.font = Font(name='Comfortaa', size=10)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    widths = (8, 13, 13, 18, 22, 35, 45, 20, 20, 10, 50)
    for column, width in enumerate(widths, start=1):
        tasks_sheet.column_dimensions[
            tasks_sheet.cell(row=1, column=column).column_letter
        ].width = width
    tasks_sheet.auto_filter.ref = f'A1:K{max(tasks_sheet.max_row, 1)}'
    tasks_sheet.row_dimensions[1].height = 28
    return workbook


def _task_duration_label(seconds, precision):
    if seconds is None:
        return 'Нет данных'
    if precision == 'day' and float(seconds) == 0:
        return '≈ в тот же день'
    prefix = '≈ ' if precision == 'day' else ''
    minutes = max(round(float(seconds) / 60), 0)
    if minutes < 1:
        return 'Меньше минуты'
    if minutes < 60:
        return f'{prefix}{minutes} мин.'
    hours, remaining_minutes = divmod(minutes, 60)
    if hours < 24:
        return f'{prefix}{hours} ч {remaining_minutes} мин.'
    days, remaining_hours = divmod(hours, 24)
    return f'{prefix}{days} дн. {remaining_hours} ч'


def _display_date(value):
    parsed = str(value or '').split('T', 1)[0]
    try:
        return datetime.strptime(parsed, '%Y-%m-%d').strftime('%d.%m.%Y')
    except ValueError:
        return ''


def _excel_text(value):
    text = str(value or '')[:32000]
    return f"'{text}" if text.startswith(('=', '+', '-', '@')) else text


def _shift_month(month, offset):
    month_start = datetime.strptime(month, '%Y-%m-%d').date()
    absolute_month = month_start.year * 12 + month_start.month - 1 + offset
    return date(
        absolute_month // 12,
        absolute_month % 12 + 1,
        1,
    ).isoformat()


def _analytics_employee(row):
    return {
        'login': row['login'],
        'nickname': row['nickname'],
        'kpi': row['total_pct'],
        'weighted_kpi': row['weighted_pct'],
        'rank': row['rank'],
        'shifts': row['shifts'],
        'weighted_shifts': row['weighted_shifts'],
        'reviews': row['reviews'],
        'forms': row['forms'],
        'extensions': row['extensions'],
        'certificates': row['certificates'],
        'subscriptions': row['subscriptions'],
        'initiatives': row['initiatives'],
        'penalties': row['penalties'],
        'stream': row['stream'],
    }


def _analytics_point(label, rows):
    participants = [row for row in rows if row['shifts'] > 0]
    count = len(participants)
    totals = {
        metric: sum(row[metric] for row in participants)
        for metric in (
            'shifts',
            'weighted_shifts',
            'reviews',
            'forms',
            'extensions',
            'certificates',
            'subscriptions',
            'initiatives',
        )
    }
    return {
        'label': label,
        'team': {
            'employees': count,
            'kpi': (
                sum(row['total_pct'] for row in participants) / count
                if count else 0
            ),
            'weighted_kpi': (
                sum(row['weighted_pct'] for row in participants) / count
                if count else 0
            ),
            'zones': {
                zone: sum(row['zone'] == zone for row in participants)
                for zone in ('🟢', '🟡', '🔴')
            },
            **totals,
        },
        'employees': [
            _analytics_employee(row)
            for row in rows
        ],
    }


def _validate_init_data(init_data, bot_token, now=None):
    if not init_data or not bot_token:
        return None
    values = dict(parse_qsl(init_data, keep_blank_values=True))
    received_hash = values.pop('hash', '')
    if not received_hash:
        return None

    data_check_string = '\n'.join(
        f'{key}={values[key]}'
        for key in sorted(values)
    )
    secret_key = hmac.new(
        b'WebAppData',
        bot_token.encode('utf-8'),
        hashlib.sha256,
    ).digest()
    expected_hash = hmac.new(
        secret_key,
        data_check_string.encode('utf-8'),
        hashlib.sha256,
    ).hexdigest()
    if not hmac.compare_digest(received_hash, expected_hash):
        return None

    try:
        auth_date = int(values.get('auth_date', 0))
        current_time = int(now if now is not None else time.time())
        if auth_date <= 0 or abs(current_time - auth_date) > AUTH_MAX_AGE_SECONDS:
            return None
        user = json.loads(values.get('user', '{}'))
        telegram_id = int(user['id'])
    except (KeyError, TypeError, ValueError, json.JSONDecodeError):
        return None
    return {
        'telegram_id': telegram_id,
        'telegram_user': user,
        'auth_date': auth_date,
    }


def _request_user():
    auth = _validate_init_data(
        request.headers.get('X-Telegram-Init-Data', ''),
        TELEGRAM_API_KEY,
    )
    if not auth:
        return None
    user = get_user(telegram_id=auth['telegram_id'])
    if not user or user['status'] not in ACTIVE_ROLES:
        return None
    if not is_main_group_member(
        _main_group_membership_bot(),
        CHATS['main_group'],
        auth['telegram_id'],
    ):
        return None
    return dict(user)


def _main_group_membership_bot():
    global _membership_bot
    if _membership_bot is not None:
        return _membership_bot
    if not TELEGRAM_API_KEY:
        return None
    with _membership_bot_lock:
        if _membership_bot is None:
            _membership_bot = telebot.TeleBot(TELEGRAM_API_KEY)
    return _membership_bot


def require_user(handler):
    @wraps(handler)
    def wrapped(*args, **kwargs):
        user = _request_user()
        if not user:
            return jsonify({'error': 'Откройте приложение через Telegram-бота.'}), 401
        g.kpi_user = user
        return handler(*args, **kwargs)
    return wrapped


def require_manager(handler):
    @wraps(handler)
    @require_user
    def wrapped(*args, **kwargs):
        if int(g.kpi_user['status']) < ROLE_MANAGER:
            return jsonify({'error': 'Действие доступно менеджерам и руководству.'}), 403
        return handler(*args, **kwargs)
    return wrapped


def require_technician(handler):
    @wraps(handler)
    @require_user
    def wrapped(*args, **kwargs):
        if int(g.kpi_user['status']) < ROLE_TECHNICIAN:
            return jsonify({
                'error': 'Аналитика доступна ремонтникам, менеджерам и руководству.',
            }), 403
        return handler(*args, **kwargs)
    return wrapped


def require_owner(handler):
    @wraps(handler)
    @require_user
    def wrapped(*args, **kwargs):
        if int(g.kpi_user['status']) != ROLE_OWNER:
            return jsonify({
                'error': 'Настройки KPI доступны только руководству.',
            }), 403
        return handler(*args, **kwargs)
    return wrapped


def _active_employee_logins():
    return active_kpi_employee_logins(DB_PATH)


def _employee_logins_with_month_shifts(employee_logins, month):
    active_logins = set(employee_logins)
    if not active_logins:
        return []
    conn = sqlite3.connect(DB_PATH)
    try:
        rows = conn.execute(
            '''
            SELECT DISTINCT lower(ns.login)
            FROM shifts sh
            JOIN users ns ON (
                sh.shift_login IS NOT NULL
                AND lower(sh.shift_login)=lower(ns.login)
            ) OR (
                sh.shift_login IS NULL
                AND sh.shift_second_name=ns.second_name
                AND sh.shift_first_name=ns.first_name
            )
            WHERE date(substr(sh.dt_shift, 1, 10)) >= date(?)
              AND date(substr(sh.dt_shift, 1, 10)) < date(?, '+1 month')
            ''',
            (month, month),
        ).fetchall()
        return [
            row[0]
            for row in rows
            if row[0] in active_logins
        ]
    finally:
        conn.close()


def _employee_metadata(employee_logins, month):
    normalized = sorted({
        str(login or '').strip().lower()
        for login in employee_logins
        if str(login or '').strip()
    })
    if not normalized:
        return {}
    placeholders = ','.join('?' for _ in normalized)
    conn = sqlite3.connect(DB_PATH)
    try:
        users = conn.execute(
            f'''
            SELECT lower(login), status
            FROM users
            WHERE lower(login) IN ({placeholders})
            ''',
            normalized,
        ).fetchall()
        clubs = conn.execute(
            f'''
            SELECT DISTINCT lower(ns.login), sh.club
            FROM shifts sh
            JOIN users ns ON (
                sh.shift_login IS NOT NULL
                AND lower(sh.shift_login)=lower(ns.login)
            ) OR (
                sh.shift_login IS NULL
                AND sh.shift_second_name=ns.second_name
                AND sh.shift_first_name=ns.first_name
            )
            WHERE lower(ns.login) IN ({placeholders})
              AND date(substr(sh.dt_shift, 1, 10)) >= date(?)
              AND date(substr(sh.dt_shift, 1, 10)) < date(?, '+1 month')
              AND sh.club IS NOT NULL
              AND trim(sh.club) <> ''
            ORDER BY sh.club
            ''',
            (*normalized, month, month),
        ).fetchall()
    finally:
        conn.close()

    result = {
        login: {
            'role': int(role),
            'role_name': ROLE_NAMES.get(int(role), str(role)),
            'clubs': [],
        }
        for login, role in users
    }
    for login, club in clubs:
        if login in result and club not in result[login]['clubs']:
            result[login]['clubs'].append(club)
    return result


def _actor_login():
    return str(g.kpi_user.get('login') or '').strip().lower()


def _legacy_text_variants(value):
    variants = {value}
    for encoding in ('cp1251', 'utf-8'):
        try:
            variants.add(value.encode(encoding).decode('latin1'))
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass
    return tuple(variants)


def _normalize_legacy_text(value):
    raw = str(value or '')
    for source, target in (('latin1', 'cp1251'), ('latin1', 'utf-8')):
        try:
            decoded = raw.encode(source).decode(target)
        except (UnicodeEncodeError, UnicodeDecodeError):
            continue
        if any('А' <= char <= 'я' or char == 'ё' for char in decoded):
            return decoded
    return raw


def _status_sql(statuses):
    values = []
    for status in statuses:
        for variant in _legacy_text_variants(status):
            if variant not in values:
                values.append(variant)
    return ','.join('?' for _ in values), values


def _plain_task_feedback(value):
    without_tags = re.sub(r'</?(?:b|i)>', '', str(value or ''), flags=re.I)
    return html.unescape(without_tags).strip()


def _moscow_today():
    return datetime.now(ZoneInfo('Europe/Moscow')).date()


def _upcoming_shifts(login, limit=3):
    today = _moscow_today().isoformat()
    conn = sqlite3.connect(DB_PATH)
    try:
        rows = conn.execute(
            '''
            SELECT date(substr(sh.dt_shift, 1, 10)), sh.club,
                   ROUND(SUM(COALESCE(sh.dur, 0)), 1),
                   MIN(sh.shift_start), MAX(sh.shift_end)
            FROM shifts sh
            JOIN users employee ON lower(employee.login)=?
             AND (
                (sh.shift_login IS NOT NULL
                 AND lower(sh.shift_login)=lower(employee.login))
                OR
                (sh.shift_login IS NULL
                 AND sh.shift_second_name=employee.second_name
                 AND sh.shift_first_name=employee.first_name)
             )
            WHERE date(substr(sh.dt_shift, 1, 10)) >= date(?)
            GROUP BY date(substr(sh.dt_shift, 1, 10)), sh.club
            ORDER BY date(substr(sh.dt_shift, 1, 10)), sh.club
            LIMIT ?
            ''',
            (login, today, limit),
        ).fetchall()
    finally:
        conn.close()
    return [
        {
            'date': row[0],
            'club': row[1],
            'duration': float(row[2] or 0),
            'start': row[3],
            'end': row[4],
        }
        for row in rows
    ]


def _shift_clock_minutes(value):
    match = re.match(r'^(\d{1,2}):(\d{2})', str(value or '').strip())
    if not match:
        return None
    hour, minute = map(int, match.groups())
    if hour > 23 or minute > 59:
        return None
    return hour * 60 + minute


def _select_shift_report_test_shift(login, now=None, requested_club=None):
    current = now or datetime.now(ZoneInfo('Europe/Moscow'))
    today = current.date().isoformat()
    shifts = [
        shift for shift in _upcoming_shifts(login, limit=50)
        if shift.get('date') == today
    ]
    if requested_club:
        normalized_request = str(requested_club).strip().casefold()
        shifts = [
            shift for shift in shifts
            if str(shift.get('club') or '').strip().casefold() == normalized_request
            or _shift_report_test_club(shift.get('club'))[0] == requested_club
        ]
    if not shifts:
        return None

    current_minutes = current.hour * 60 + current.minute

    def score(shift):
        start = _shift_clock_minutes(shift.get('start'))
        end = _shift_clock_minutes(shift.get('end'))
        if start is None:
            return 2, 0, str(shift.get('club') or '').casefold()
        if end is None:
            duration = max(0, int(round(float(shift.get('duration') or 0) * 60)))
            end = start + duration
        if end < start:
            end += 24 * 60
        comparable_now = current_minutes
        if end >= 24 * 60 and current_minutes < start:
            comparable_now += 24 * 60
        if start <= comparable_now <= end:
            return 0, 0, str(shift.get('club') or '').casefold()
        return 1, abs(comparable_now - start), str(shift.get('club') or '').casefold()

    return min(shifts, key=score)


def _shift_report_test_club(shift_club):
    normalized = str(shift_club or '').strip().casefold()
    for club_name, info in get_clubs().items():
        aliases = {str(club_name).strip().casefold()}
        aliases.add(str(info.get('shift_name') or '').strip().casefold())
        if normalized in aliases:
            return club_name, info
    return None, None


def _shift_report_test_owner_clubs(action):
    if action not in SHIFT_ACTIONS:
        raise ValueError('Выберите открытие или закрытие')
    action_name = SHIFT_ACTIONS[action]
    return sorted(
        [
            club_name
            for club_name, club in get_clubs().items()
            if club.get('questions', {}).get(action_name)
        ],
        key=str.casefold,
    )


def _initialize_shift_report_schema(db_path=DB_PATH):
    conn = sqlite3.connect(db_path)
    try:
        with conn:
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS shift_webapp_runs (
                       id TEXT PRIMARY KEY,
                       login TEXT NOT NULL,
                       chatid TEXT NOT NULL,
                       club TEXT NOT NULL,
                       action TEXT NOT NULL,
                       shift_date TEXT NOT NULL,
                       scenario_version TEXT NOT NULL,
                       variant_index INTEGER NOT NULL,
                       started_at TEXT NOT NULL,
                       early_close INTEGER NOT NULL DEFAULT 0,
                       start_notified_at TEXT,
                       finished_at TEXT,
                       answers_json TEXT,
                       photo_count INTEGER,
                       report_sent_at TEXT,
                       warning_sent_at TEXT,
                       main_message TEXT,
                       main_sent_at TEXT,
                       activity_id INTEGER,
                       sheet_synced_at TEXT,
                       completed_at TEXT
                   )'''
            )
    finally:
        conn.close()


def _shift_report_is_early_close(action, club, now=None):
    if action != 'close':
        return False
    current = now or datetime.now(ZoneInfo('Europe/Moscow'))
    raw_limit = str(club.get('schedule', {}).get('early_check_time') or '')
    try:
        limit_time = datetime.strptime(raw_limit, '%H:%M:%S').time()
    except ValueError:
        return False
    limit = current.replace(
        hour=limit_time.hour,
        minute=limit_time.minute,
        second=0,
        microsecond=0,
    )
    if current.hour < 5:
        limit -= timedelta(days=1)
    return current < limit


def _shift_report_late_minutes(club, started_at):
    schedule = club.get('schedule', {}).get('open_strict', {})
    target_raw = (
        schedule.get('weekend')
        if started_at.weekday() >= 5
        else schedule.get('weekdays')
    )
    try:
        target_time = datetime.strptime(str(target_raw), '%H:%M:%S').time()
    except ValueError:
        return 0
    target = started_at.replace(
        hour=target_time.hour,
        minute=target_time.minute,
        second=0,
        microsecond=0,
    )
    return max(0, int((started_at - target).total_seconds() / 60))


def _shift_report_test_scenario(
    action, variant_index=None, now=None, requested_club=None,
):
    if action not in SHIFT_ACTIONS:
        raise ValueError('Выберите открытие или закрытие')
    shift = _select_shift_report_test_shift(
        _actor_login(), now=now, requested_club=requested_club,
    )
    manual_club = None
    if (
        not shift
        and requested_club
        and int(g.kpi_user['status']) == ROLE_OWNER
        and not _select_shift_report_test_shift(_actor_login(), now=now)
    ):
        manual_club = _shift_report_test_club(requested_club)
        if not manual_club[1]:
            raise ValueError('Выбранный клуб не найден в настройках OMG Shift')
        current = now or datetime.now(ZoneInfo('Europe/Moscow'))
        shift = {
            'date': current.date().isoformat(),
            'club': manual_club[0],
            'duration': 0,
            'start': None,
            'end': None,
            'manual': True,
        }
    if not shift:
        raise ValueError('На сегодня смена в OMG Shift не найдена')

    club_name, club = manual_club or _shift_report_test_club(shift.get('club'))
    if not club:
        raise ValueError('Для клуба из расписания не найден сценарий смены')
    action_name = SHIFT_ACTIONS[action]
    variants = club.get('questions', {}).get(action_name, [])
    if not variants:
        raise ValueError(f'Для клуба «{club_name}» не настроен этот сценарий')

    if variant_index is None:
        selected_index = random.randrange(len(variants))
    else:
        try:
            selected_index = int(variant_index)
        except (TypeError, ValueError) as error:
            raise ValueError('Номер набора сценария задан неверно') from error
        if selected_index < 0 or selected_index >= len(variants):
            raise ValueError('Сохранённый набор сценария больше недоступен')

    raw_questions = variants[selected_index]
    questions = [
        {
            'id': f'q{index}',
            'position': index,
            'text': str(question.get('text') or '').strip(),
            'type': str(question.get('type') or '').strip(),
            'checklist': str(question.get('checklist') or '').strip(),
        }
        for index, question in enumerate(raw_questions, 1)
    ]
    if not questions or any(
        not question['text'] or question['type'] not in {'text', 'num', 'photo'}
        for question in questions
    ):
        raise ValueError('Сценарий смены настроен некорректно')
    if sum(question['type'] == 'photo' for question in questions) > MAX_PHOTO_QUESTIONS:
        raise ValueError('В сценарии больше десяти фото-вопросов')

    checklist = [question['checklist'] for question in questions if question['checklist']]
    version_payload = {
        'club': club_name,
        'action': action,
        'variant_index': selected_index,
        'checklist': checklist,
        'questions': questions,
    }
    version = hashlib.sha256(json.dumps(
        version_payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(',', ':'),
    ).encode('utf-8')).hexdigest()[:16]
    return {
        'production_mode': True,
        'action': action,
        'action_label': 'Открытие' if action == 'open' else 'Закрытие',
        'club': club_name,
        'shift': shift,
        'variant_index': selected_index,
        'variant_label': chr(ord('A') + selected_index),
        'version': version,
        'checklist': checklist,
        'questions': questions,
        'early_close': _shift_report_is_early_close(action, club, now=now),
        'user_login': _actor_login(),
        'user_name': (
            g.kpi_user.get('nick_name')
            or g.kpi_user.get('first_name')
            or g.kpi_user['login']
        ),
        'draft_ttl_hours': 18,
    }


def _start_shift_report_run(scenario, run_id, early_confirmed=False):
    if not re.fullmatch(r'[A-Za-z0-9._:-]{8,180}', str(run_id or '')):
        raise ValueError('Идентификатор отчёта передан неверно')
    _initialize_shift_report_schema(DB_PATH)
    initialize_club_status_dashboard_schema(DB_PATH)
    actor_login = _actor_login()
    actor_chatid = str(g.kpi_user['chatid'])
    now = datetime.now(ZoneInfo('Europe/Moscow'))
    started_at = now.strftime('%Y-%m-%d %H:%M:%S')
    club_name, club = _shift_report_test_club(scenario['club'])
    if not club:
        raise ValueError('Клуб больше не найден в настройках OMG Shift')
    early_close = _shift_report_is_early_close(
        scenario['action'], club, now=now,
    )
    if early_close and not early_confirmed:
        return {'requires_early_confirmation': True}

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute('BEGIN IMMEDIATE')
        existing = conn.execute(
            'SELECT * FROM shift_webapp_runs WHERE id=?',
            (run_id,),
        ).fetchone()
        if existing:
            if (
                existing['login'] != actor_login
                or existing['club'] != club_name
                or existing['action'] != scenario['action']
                or existing['scenario_version'] != scenario['version']
            ):
                raise ValueError('Черновик открытия или закрытия больше недействителен')
            conn.commit()
            run = dict(existing)
        else:
            status_row = conn.execute(
                'SELECT status FROM clubs WHERE club=?',
                (club_name,),
            ).fetchone()
            if not status_row:
                raise ValueError('Клуб не найден в рабочей базе')
            new_status = 'Открыт' if scenario['action'] == 'open' else 'Закрыт'
            if status_row['status'] == new_status:
                raise ValueError(f'Клуб «{club_name}» уже {new_status.lower()}')
            conn.execute(
                'UPDATE clubs SET status=? WHERE club=?',
                (new_status, club_name),
            )
            conn.execute(
                '''INSERT INTO club_status_updates (club, changed_at)
                   VALUES (?, ?)
                   ON CONFLICT(club) DO UPDATE SET changed_at=excluded.changed_at''',
                (club_name, started_at),
            )
            conn.execute(
                '''INSERT INTO shift_webapp_runs (
                       id, login, chatid, club, action, shift_date,
                       scenario_version, variant_index, started_at, early_close
                   ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (
                    run_id,
                    actor_login,
                    actor_chatid,
                    club_name,
                    scenario['action'],
                    scenario['shift']['date'],
                    scenario['version'],
                    scenario['variant_index'],
                    started_at,
                    int(early_close),
                ),
            )
            conn.commit()
            run = dict(conn.execute(
                'SELECT * FROM shift_webapp_runs WHERE id=?',
                (run_id,),
            ).fetchone())
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    bot = _notification_bot()
    if not bot:
        raise RuntimeError('Telegram-бот временно недоступен')
    refresh_club_status_dashboard(bot, DB_PATH)
    if not run.get('start_notified_at'):
        name = (
            g.kpi_user.get('nick_name')
            or g.kpi_user.get('first_name')
            or actor_login
        )
        action_text = 'зашёл в' if scenario['action'] == 'open' else 'начинает закрывать'
        bot.send_message(
            CHATS['reports'],
            f'⚠️ {name} {action_text} {club_name} в {run["started_at"][11:16]}',
        )
        notified_at = datetime.now(ZoneInfo('Europe/Moscow')).strftime(
            '%Y-%m-%d %H:%M:%S'
        )
        conn = sqlite3.connect(DB_PATH)
        try:
            with conn:
                conn.execute(
                    'UPDATE shift_webapp_runs SET start_notified_at=? WHERE id=?',
                    (notified_at, run_id),
                )
        finally:
            conn.close()
        run['start_notified_at'] = notified_at
    return {
        'started': True,
        'started_at': run['started_at'],
        'early_close': bool(run['early_close']),
    }


def _shift_month_summary(login, month):
    conn = sqlite3.connect(DB_PATH)
    try:
        row = conn.execute(
            '''
            SELECT COUNT(*), ROUND(SUM(grouped.duration), 1)
            FROM (
                SELECT date(substr(sh.dt_shift, 1, 10)), sh.club,
                       SUM(COALESCE(sh.dur, 0)) AS duration
                FROM shifts sh
                JOIN users employee ON lower(employee.login)=?
                 AND (
                    (sh.shift_login IS NOT NULL
                     AND lower(sh.shift_login)=lower(employee.login))
                    OR
                    (sh.shift_login IS NULL
                     AND sh.shift_second_name=employee.second_name
                     AND sh.shift_first_name=employee.first_name)
                 )
                WHERE substr(sh.dt_shift, 1, 7)=?
                GROUP BY date(substr(sh.dt_shift, 1, 10)), sh.club
            ) grouped
            ''',
            (login, month),
        ).fetchone()
    finally:
        conn.close()
    return {
        'shifts': int(row[0] or 0),
        'hours': float(row[1] or 0),
    }


def _today_shift_clubs(login):
    today = _moscow_today().isoformat()
    conn = sqlite3.connect(DB_PATH)
    try:
        rows = conn.execute(
            '''
            SELECT DISTINCT sh.club
            FROM shifts sh
            JOIN users employee ON lower(employee.login)=?
             AND (
                (sh.shift_login IS NOT NULL
                 AND lower(sh.shift_login)=lower(employee.login))
                OR
                (sh.shift_login IS NULL
                 AND sh.shift_second_name=employee.second_name
                 AND sh.shift_first_name=employee.first_name)
             )
            WHERE date(substr(sh.dt_shift, 1, 10))=date(?)
              AND sh.club IS NOT NULL
              AND trim(sh.club) <> ''
            ORDER BY sh.club
            ''',
            (login, today),
        ).fetchall()
    finally:
        conn.close()
    return [row[0] for row in rows]


def _public_booking(order, include_order=False):
    result = {
        'start': order.get('reservation_at'),
        'end': order.get('reservation_end_at'),
        'format': order.get('booking_format') or order.get('resource') or '',
        'participants': float(order.get('participants') or 0),
    }
    if include_order:
        result.update({
            'date': order.get('date').isoformat() if order.get('date') else None,
            'club': order.get('club'),
            'number': order.get('number'),
            'url': order.get('url'),
        })
    return result


def _club_booking_groups(clubs, today):
    club_order = list(dict.fromkeys(clubs))
    grouped = {
        club: {
            'club': club,
            'count': 0,
            'participants': 0,
            'bookings': [],
        }
        for club in club_order
    }
    for order in active_orders_for_day(today, club_order, DB_PATH):
        club = order.get('club')
        if club not in grouped:
            continue
        grouped[club]['bookings'].append(_public_booking(order))
        grouped[club]['count'] += 1
        grouped[club]['participants'] += float(order.get('participants') or 0)
    return [grouped[club] for club in club_order]


def _task_counts():
    placeholders, statuses = _status_sql(('В работе', 'На проверке'))
    conn = sqlite3.connect(DB_PATH)
    try:
        rows = conn.execute(
            f'''SELECT status, COUNT(*) FROM tasks
                WHERE status IN ({placeholders}) GROUP BY status''',
            statuses,
        ).fetchall()
    finally:
        conn.close()
    result = {'work': 0, 'review': 0}
    for status, count in rows:
        normalized = _normalize_legacy_text(status)
        if normalized == 'В работе':
            result['work'] += int(count)
        elif normalized == 'На проверке':
            result['review'] += int(count)
    return result


def _team_snapshot(month, selected_day):
    logins = _active_employee_logins()
    rows = calculate_monthly_kpi(
        month,
        employee_logins=logins,
        period_end=selected_day,
    )
    metadata = _employee_metadata(logins, month)
    for row in rows:
        row.update(metadata.get(row['login'], {'clubs': []}))
    participants = [row for row in rows if float(row.get('shifts', 0) or 0) > 0]
    average = (
        sum(float(row.get('total_pct', 0) or 0) for row in participants)
        / len(participants)
        if participants else 0
    )
    return rows, {
        'participants': len(participants),
        'average_pct': average,
        'red_zone': sum(row.get('zone') == '🔴' for row in participants),
        'active_penalties': sum(
            int(row.get('penalties', 0) or 0) for row in rows
        ),
    }


def _club_dashboard(team_rows, problem_counts):
    today = _moscow_today().isoformat()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        has_updates = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' "
            "AND name='club_status_updates'"
        ).fetchone()
        update_join = (
            'LEFT JOIN club_status_updates updates ON updates.club=clubs.club'
            if has_updates else ''
        )
        update_column = 'updates.changed_at' if has_updates else 'NULL'
        clubs = conn.execute(
            f'''SELECT clubs.club, clubs.status, {update_column}
                FROM clubs {update_join} ORDER BY clubs.club COLLATE NOCASE'''
        ).fetchall()
        shifts = conn.execute(
            '''
            SELECT sh.club,
                   COALESCE(NULLIF(employee.nick_name, ''),
                            NULLIF(employee.first_name, ''), employee.login),
                   employee.login
            FROM shifts sh
            LEFT JOIN users employee ON (
                sh.shift_login IS NOT NULL
                AND lower(sh.shift_login)=lower(employee.login)
            ) OR (
                sh.shift_login IS NULL
                AND sh.shift_second_name=employee.second_name
                AND sh.shift_first_name=employee.first_name
            )
            WHERE date(substr(sh.dt_shift, 1, 10))=date(?)
            GROUP BY sh.club, employee.login, employee.first_name,
                     employee.nick_name
            ORDER BY sh.club, employee.ID
            ''',
            (today,),
        ).fetchall()
    finally:
        conn.close()

    shift_people = {}
    for club, name, login in shifts:
        if name:
            people = shift_people.setdefault(club, [])
            person = {'name': name, 'login': login}
            if person not in people:
                people.append(person)
    physical_clubs = {
        name
        for name, settings in get_clubs().items()
        if settings.get('is_physical') is True
    }
    return [
        {
            'club': row['club'],
            'status': _normalize_legacy_text(row['status']),
            'changed_at': row[2],
            'on_shift': [
                person['name'] for person in shift_people.get(row['club'], [])
            ],
            'on_shift_contacts': shift_people.get(row['club'], []),
            'problems': problem_counts.get(row['club'], {'work': 0, 'review': 0}),
        }
        for row in clubs
        if row['club'] in physical_clubs
    ]


def _problem_counts_by_club():
    placeholders, statuses = _status_sql(('В работе', 'На проверке'))
    conn = sqlite3.connect(DB_PATH)
    try:
        rows = conn.execute(
            f'''SELECT club, status, COUNT(*) FROM tasks
                WHERE status IN ({placeholders})
                GROUP BY club, status''',
            statuses,
        ).fetchall()
    finally:
        conn.close()
    result = {}
    for club, status, count in rows:
        item = result.setdefault(club, {'work': 0, 'review': 0})
        normalized = _normalize_legacy_text(status)
        if normalized == 'В работе':
            item['work'] += int(count)
        elif normalized == 'На проверке':
            item['review'] += int(count)
    return result


def _notification_bot():
    return telebot.TeleBot(TELEGRAM_API_KEY) if TELEGRAM_API_KEY else None


def _camera_test_caption(diagnostics, media_type, media_size):
    user = g.kpi_user
    name = (
        user.get('nick_name') or user.get('first_name') or user.get('login')
    )
    fields = (
        ('Платформа', 'platform'),
        ('Telegram WebApp', 'telegram_version'),
        ('Способ съёмки', 'capture_method'),
        ('Формат', 'mime_type'),
        ('Разрешение', 'resolution'),
        ('Продолжительность', 'duration'),
        ('Камера', 'camera_status'),
        ('Микрофон', 'microphone_status'),
        ('MediaRecorder', 'recorder_status'),
        ('WebView', 'user_agent'),
        ('Ошибка', 'error'),
    )
    lines = [
        '🧪 <b>Тест камеры Mini App</b>',
        '',
        f"👤 {html.escape(str(name))} · {html.escape(str(user.get('login') or '—'))}",
        f"📎 {html.escape(media_type)} · {media_size / (1024 * 1024):.2f} МБ",
    ]
    for label, key in fields:
        value = str(diagnostics.get(key) or '').strip()
        if value:
            lines.append(f"<b>{label}:</b> {html.escape(value[:180])}")
    return '\n'.join(lines)[:1000]


def _camera_test_upload():
    upload = request.files.get('media')
    if not upload or not upload.filename:
        return None
    mimetype = str(upload.mimetype or '').lower().split(';', 1)[0]
    image_types = {
        'image/jpeg', 'image/png', 'image/webp', 'image/heic', 'image/heif',
    }
    video_types = {'video/mp4', 'video/quicktime', 'video/webm'}
    if mimetype in image_types:
        maximum = 6 * 1024 * 1024
        media_type = 'Фото'
    elif mimetype in video_types:
        maximum = 20 * 1024 * 1024
        media_type = 'Видео'
    else:
        raise ValueError('Тест камеры принимает только JPEG, PNG, WebP, HEIC, MP4, MOV или WebM')
    content = upload.read(maximum + 1)
    if not content:
        raise ValueError('Камера вернула пустой файл')
    if len(content) > maximum:
        raise ValueError(
            'Фото должно быть не больше 6 МБ, видео — не больше 20 МБ'
        )
    return {
        'content': content,
        'filename': str(upload.filename or 'camera-test').replace('/', '_')[:100],
        'mimetype': mimetype,
        'media_type': media_type,
    }


def _shift_report_test_data(scenario, payload):
    if not isinstance(payload, dict):
        raise ValueError('Отчёт передан неверно')
    if str(payload.get('version') or '') != scenario['version']:
        raise ValueError('Сценарий смены изменился. Начните отчёт заново')

    raw_answers = payload.get('answers')
    if not isinstance(raw_answers, dict):
        raise ValueError('Ответы на вопросы переданы неверно')
    answers = {}
    photo_questions = []
    for question in scenario['questions']:
        question_id = question['id']
        if question['type'] == 'photo':
            photo_questions.append(question)
            continue
        value = str(raw_answers.get(question_id) or '').strip()
        if not value:
            raise ValueError(f"Ответьте на вопрос «{question['text']}»")
        if len(value) > 1000:
            raise ValueError('Ответ должен быть не длиннее 1000 символов')
        if question['type'] == 'num' and not value.isdecimal():
            raise ValueError(f"Для вопроса «{question['text']}» нужно указать целое число")
        answers[question_id] = value

    raw_photo_ids = payload.get('photo_ids')
    if not isinstance(raw_photo_ids, list):
        raise ValueError('Список фотографий передан неверно')
    expected_photo_ids = [question['id'] for question in photo_questions]
    photo_ids = [str(value) for value in raw_photo_ids]
    if photo_ids != expected_photo_ids:
        raise ValueError('Нужно приложить по одной фотографии к каждому фото-вопросу')

    uploads = request.files.getlist('photos')
    if len(uploads) != len(photo_questions):
        raise ValueError('Количество фотографий не совпадает со сценарием')
    allowed_types = {'image/jpeg', 'image/png', 'image/webp'}
    photos = []
    total_size = 0
    for upload, question in zip(uploads, photo_questions):
        mimetype = str(upload.mimetype or '').lower().split(';', 1)[0]
        if mimetype not in allowed_types:
            raise ValueError('Фотографии должны быть в формате JPEG, PNG или WebP')
        content = upload.read(3 * 1024 * 1024 + 1)
        if not content:
            raise ValueError('Получена пустая фотография')
        if len(content) > 3 * 1024 * 1024:
            raise ValueError('Каждая фотография должна быть не больше 3 МБ')
        total_size += len(content)
        if total_size > 20 * 1024 * 1024:
            raise ValueError('Общий размер фотографий должен быть не больше 20 МБ')
        extension = {'image/jpeg': 'jpg', 'image/png': 'png', 'image/webp': 'webp'}[mimetype]
        photos.append({
            'content': content,
            'filename': f"{question['id']}.{extension}",
            'question': question,
        })
    return answers, photos


def _shift_report_test_messages(scenario, answers):
    user = g.kpi_user
    name = user.get('nick_name') or user.get('first_name') or user.get('login')
    login = str(user.get('login') or '—')
    shift = scenario['shift']
    time_label = (
        'клуб выбран вручную'
        if shift.get('manual')
        else '–'.join(
            value for value in (shift.get('start'), shift.get('end')) if value
        ) or 'время не указано'
    )
    heading_emoji = '🌅' if scenario['action_label'] == 'Открытие' else '🌙'
    heading_lines = [
        f"{heading_emoji} <b>{html.escape(scenario['action_label'])} смены</b>",
        '',
        f"📍 <b>Клуб:</b> {html.escape(scenario['club'])}",
        f"👤 <b>Сотрудник:</b> {html.escape(str(name))} · {html.escape(login)}",
        f"📅 <b>Дата:</b> {html.escape(str(shift.get('date') or '—'))}",
    ]
    if scenario.get('started_at') and scenario.get('finished_at'):
        heading_lines.extend([
            f"⏰ <b>Начато:</b> {html.escape(scenario['started_at'][11:16])}",
            f"✅ <b>Завершено:</b> {html.escape(scenario['finished_at'][11:16])}",
        ])
    else:
        heading_lines.extend([
            f"⏰ <b>Смена:</b> {html.escape(time_label)}",
            f"✅ <b>Отправлено:</b> {datetime.now(ZoneInfo('Europe/Moscow')).strftime('%d.%m.%Y %H:%M')}",
        ])
    heading = '\n'.join(heading_lines)
    blocks = [heading]
    for question in scenario['questions']:
        if question['type'] == 'photo':
            continue
        question_parts = [
            question['text'][index:index + 400]
            for index in range(0, len(question['text']), 400)
        ]
        answer = answers[question['id']]
        answer_parts = [
            answer[index:index + 500]
            for index in range(0, len(answer), 500)
        ]
        blocks.append(
            f"<b>{question['position']}. {html.escape(question_parts[0])}</b>"
        )
        blocks.extend(
            f"<b>{html.escape(part)}</b>"
            for part in question_parts[1:]
        )
        blocks.extend(
            f"{'— ' if index == 0 else ''}{html.escape(part)}"
            for index, part in enumerate(answer_parts)
        )

    messages = []
    current = ''
    for block in blocks:
        candidate = f'{current}\n\n{block}' if current else block
        if len(candidate) > 900 and current:
            messages.append(current)
            current = f'📝 <b>Продолжение отчёта</b>\n\n{block}'
        else:
            current = candidate
    if current:
        messages.append(current)
    return messages


def _send_shift_report_test_photos(
    bot, photos, report_captions=None, chat_id=None,
):
    target_chat = chat_id or CAMERA_TEST_RECIPIENT_CHAT_ID
    total = len(photos)

    def caption(index, photo):
        if report_captions is not None:
            caption_index = 0 if total == 1 else index - 2
            if 0 <= caption_index < len(report_captions):
                return report_captions[caption_index]
            return None
        photo_label = (
            f"📸 <b>{index}/{total}</b> · "
            f"{html.escape(photo['question']['text'][:900])}"
        )
        return photo_label

    if total == 1:
        photo = photos[0]
        media_file = io.BytesIO(photo['content'])
        media_file.name = photo['filename']
        bot.send_photo(
            target_chat,
            media_file,
            caption=caption(1, photo),
            parse_mode='HTML',
        )
        return

    media = []
    for index, photo in enumerate(photos, 1):
        media_file = io.BytesIO(photo['content'])
        media_file.name = photo['filename']
        item_caption = caption(index, photo)
        media.append(telebot.types.InputMediaPhoto(
            media_file,
            caption=item_caption,
            parse_mode='HTML' if item_caption else None,
        ))
    bot.send_media_group(target_chat, media=media)


def _send_shift_report_test(scenario, answers, photos, chat_id=None):
    bot = _notification_bot()
    if not bot:
        raise RuntimeError('Telegram-бот временно недоступен')
    messages = _shift_report_test_messages(scenario, answers)
    caption_slots = 1 if len(photos) == 1 else max(0, len(photos) - 1)
    if photos and len(messages) <= caption_slots:
        try:
            _send_shift_report_test_photos(
                bot, photos, messages, chat_id=chat_id,
            )
            return bot
        except Exception as error:
            print(f'Единая отправка отчёта смены не удалась: {error}')

    for message in messages:
        bot.send_message(
            chat_id or CAMERA_TEST_RECIPIENT_CHAT_ID,
            message,
            parse_mode='HTML',
        )
    if photos:
        _send_shift_report_test_photos(bot, photos, chat_id=chat_id)
    return bot


def _complete_shift_report_run(scenario, payload, answers, photos):
    run_id = str(payload.get('run_id') or '')
    if not re.fullmatch(r'[A-Za-z0-9._:-]{8,180}', run_id):
        raise ValueError('Сначала начните открытие или закрытие смены')
    _initialize_shift_report_schema(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        run_row = conn.execute(
            'SELECT * FROM shift_webapp_runs WHERE id=?',
            (run_id,),
        ).fetchone()
    finally:
        conn.close()
    if not run_row:
        raise ValueError('Начало открытия или закрытия не найдено')
    run = dict(run_row)
    if (
        run['login'] != _actor_login()
        or run['club'] != scenario['club']
        or run['action'] != scenario['action']
        or run['scenario_version'] != scenario['version']
        or int(run['variant_index']) != int(scenario['variant_index'])
    ):
        raise ValueError('Отчёт не соответствует начатой смене')
    if run.get('completed_at'):
        return {'completed': True, 'already_completed': True}

    def mark(**values):
        if not values:
            return
        assignments = ', '.join(f'{column}=?' for column in values)
        update_conn = sqlite3.connect(DB_PATH)
        try:
            with update_conn:
                update_conn.execute(
                    f'UPDATE shift_webapp_runs SET {assignments} WHERE id=?',
                    (*values.values(), run_id),
                )
        finally:
            update_conn.close()
        run.update(values)

    if not run.get('finished_at'):
        mark(
            finished_at=datetime.now(ZoneInfo('Europe/Moscow')).strftime(
                '%Y-%m-%d %H:%M:%S'
            ),
            answers_json=json.dumps(answers, ensure_ascii=False),
            photo_count=len(photos),
        )

    started_at = datetime.strptime(
        run['started_at'], '%Y-%m-%d %H:%M:%S'
    ).replace(tzinfo=ZoneInfo('Europe/Moscow'))
    club_name, club = _shift_report_test_club(run['club'])
    if not club:
        raise ValueError('Клуб больше не найден в настройках OMG Shift')
    late_minutes = (
        _shift_report_late_minutes(club, started_at)
        if scenario['action'] == 'open'
        else 0
    )
    report_scenario = dict(scenario)
    report_scenario['started_at'] = run['started_at']
    report_scenario['finished_at'] = run['finished_at']

    if not run.get('report_sent_at'):
        _send_shift_report_test(
            report_scenario,
            answers,
            photos,
            chat_id=CHATS['reports'],
        )
        mark(report_sent_at=datetime.now(ZoneInfo('Europe/Moscow')).strftime(
            '%Y-%m-%d %H:%M:%S'
        ))

    bot = _notification_bot()
    if not bot:
        raise RuntimeError('Telegram-бот временно недоступен')
    if not run.get('warning_sent_at'):
        warning = None
        if run['early_close'] and scenario['action'] == 'close':
            warning = f'⚠️ Внимание! Раннее закрытие!\n{tags_main}'
        elif scenario['action'] == 'open' and late_minutes > 5:
            warning = (
                f'😡 Внимание! ОПОЗДАНИЕ на {late_minutes} мин!\n{tags_main}'
            )
        if warning:
            bot.send_message(CHATS['reports'], warning)
            mark(warning_sent_at=datetime.now(ZoneInfo('Europe/Moscow')).strftime(
                '%Y-%m-%d %H:%M:%S'
            ))

    name = (
        g.kpi_user.get('nick_name')
        or g.kpi_user.get('first_name')
        or _actor_login()
    )
    if not run.get('main_message'):
        if scenario['action'] == 'open' and late_minutes > 5:
            phrase_type = 'penalty_phrases'
            penalty = f'🚨 ШТРАФ (опоздание {late_minutes} мин)! 🚨\n'
        elif scenario['action'] == 'open':
            phrase_type = 'good_morning'
            penalty = ''
        else:
            phrase_type = 'good_night'
            penalty = ''
        phrases = TEXTS.get(
            phrase_type,
            ['Смена открыта/закрыта.', 'Хорошего отдыха!'],
        )
        action_text = SHIFT_ACTIONS[scenario['action']].lower().replace('ть', 'л')
        main_message = (
            f'{name} {action_text} в {club_name} '
            f'в {run["finished_at"][11:16]}! {penalty}{random.choice(phrases)}'
        )
        mark(main_message=main_message)
    if not run.get('main_sent_at'):
        bot.send_message(CHATS['main_group'], run['main_message'])
        mark(main_sent_at=datetime.now(ZoneInfo('Europe/Moscow')).strftime(
            '%Y-%m-%d %H:%M:%S'
        ))

    if not run.get('activity_id'):
        db_conn = sqlite3.connect(DB_PATH)
        try:
            with db_conn:
                cursor = db_conn.execute(
                    '''INSERT INTO activity (dtrep, login, club, action)
                       VALUES (?, ?, ?, ?)''',
                    (
                        run['finished_at'],
                        _actor_login(),
                        club_name,
                        SHIFT_ACTIONS[scenario['action']],
                    ),
                )
                first_answer = next(iter(answers.values()), '')
                if str(first_answer).isdigit():
                    try:
                        db_conn.execute(
                            'INSERT INTO nal (drep, club, amount) VALUES (?, ?, ?)',
                            (run['finished_at'], club_name, first_answer),
                        )
                    except Exception as error:
                        print(f'Ошибка записи нала из Mini App: {error}')
                activity_id = cursor.lastrowid
        finally:
            db_conn.close()
        mark(activity_id=activity_id)

    if not run.get('sheet_synced_at'):
        try:
            update_table_open()
        except Exception as error:
            print(f'Ошибка выгрузки открытия или закрытия из Mini App: {error}')
        else:
            mark(sheet_synced_at=datetime.now(ZoneInfo('Europe/Moscow')).strftime(
                '%Y-%m-%d %H:%M:%S'
            ))

    completed_at = datetime.now(ZoneInfo('Europe/Moscow')).strftime(
        '%Y-%m-%d %H:%M:%S'
    )
    mark(completed_at=completed_at)
    return {'completed': True, 'already_completed': False}


def _problem_mentions(task_type, club):
    club_tag = str(get_clubs().get(club, {}).get('tag') or '').strip()
    if task_type == REPAIR_TASK_TYPE:
        repair_tag = extra_tags.get(task_type, '')
        return ' '.join(
            value for value in (repair_tag if club_tag != repair_tag else '', club_tag)
            if value
        )
    if task_type == BOT_TASK_TYPE:
        return extra_tags.get(task_type, '')
    return club_tag


def _send_problem_notification(
    event, task, message='', photo=None, video=None, actor=None,
):
    bot = _notification_bot()
    if not bot:
        return None
    task_type = _normalize_legacy_text(task['type'])
    mentions = _problem_mentions(task_type, task['club'])
    video_reference = None
    try:
        if event == 'created':
            full, short, _confirmation = created_task_notification(
                task_type,
                task['club'],
                task['title'],
                task.get('description'),
                actor=actor,
            )
            report_text = f"#задачи\n\n{full}\n\n@OMGVR_Admin_Bot"
            repair_chat = (
                CHATS.get('repair_extra')
                if task_type == REPAIR_TASK_TYPE else None
            )
            telegram_media = None
            if video and CHATS.get('reports'):
                video_file = io.BytesIO(video['content'])
                video_file.name = video['filename']
                if video['mimetype'] == 'video/mp4':
                    sent = bot.send_video(
                        CHATS['reports'], video_file, caption=report_text,
                        parse_mode='HTML', supports_streaming=True,
                    )
                    telegram_media = sent.video
                else:
                    sent = bot.send_document(
                        CHATS['reports'], video_file, caption=report_text,
                        parse_mode='HTML',
                    )
                    telegram_media = sent.document
                video_reference = {
                    'file_id': telegram_media.file_id,
                    'file_unique_id': telegram_media.file_unique_id,
                    'mimetype': video['mimetype'],
                    'file_size': len(video['content']),
                }
            elif photo and CHATS.get('reports'):
                photo_file = io.BytesIO(photo)
                photo_file.name = 'problem.jpg'
                bot.send_photo(
                    CHATS['reports'], photo_file, caption=report_text,
                    parse_mode='HTML',
                )
            elif CHATS.get('reports'):
                bot.send_message(CHATS['reports'], report_text, parse_mode='HTML')
            if CHATS.get('main_group'):
                bot.send_message(
                    CHATS['main_group'], f"{mentions}\n\n{short}",
                    parse_mode='HTML',
                )
            if repair_chat:
                if video and telegram_media:
                    if video['mimetype'] == 'video/mp4':
                        bot.send_video(
                            repair_chat, telegram_media.file_id, caption=full,
                            parse_mode='HTML', supports_streaming=True,
                        )
                    else:
                        bot.send_document(
                            repair_chat, telegram_media.file_id, caption=full,
                            parse_mode='HTML',
                        )
                elif photo:
                    repair_photo = io.BytesIO(photo)
                    repair_photo.name = 'problem.jpg'
                    bot.send_photo(
                        repair_chat, repair_photo, caption=full,
                        parse_mode='HTML',
                    )
                else:
                    bot.send_message(repair_chat, full, parse_mode='HTML')
        elif event in {'solution', 'returned', 'completed'}:
            full, short = progress_task_notification(
                event,
                task_type,
                task['club'],
                task['title'],
                message,
                actor=actor,
            )
            if CHATS.get('reports'):
                bot.send_message(
                    CHATS['reports'], f"#задачи\n\n{full}\n\n@OMGVR_Admin_Bot",
                    parse_mode='HTML',
                )
            if CHATS.get('main_group'):
                prefix = f'{mentions}\n\n' if event in {'returned', 'completed'} else ''
                bot.send_message(
                    CHATS['main_group'], f'{prefix}{short}', parse_mode='HTML',
                )
            if task_type == REPAIR_TASK_TYPE and CHATS.get('repair_extra'):
                bot.send_message(
                    CHATS['repair_extra'], full,
                    parse_mode='HTML',
                )
    except Exception as error:
        print(f'Ошибка уведомления Mini App Taskboard: {error}')
    return video_reference


def _initialize_problem_video_schema(db_path=DB_PATH):
    conn = sqlite3.connect(db_path)
    try:
        with conn:
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS task_videos (
                       task_id INTEGER PRIMARY KEY,
                       telegram_file_id TEXT NOT NULL,
                       telegram_file_unique_id TEXT,
                       mime_type TEXT NOT NULL,
                       file_size INTEGER NOT NULL
                   )'''
            )
    finally:
        conn.close()


def _delete_problem_after_failed_video(task_id, db_path=DB_PATH):
    conn = sqlite3.connect(db_path)
    try:
        with conn:
            tables = {
                row[0]
                for row in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                )
            }
            equipment_units = []
            if 'equipment_unit_tasks' in tables:
                equipment_units = [
                    row[0] for row in conn.execute(
                        'SELECT unit_id FROM equipment_unit_tasks WHERE task_id=?',
                        (task_id,),
                    )
                ]
            for table in (
                'equipment_unit_tasks', 'repair_case_locations', 'repair_events',
                'repair_cases', 'task_videos', 'task_events',
            ):
                if table in tables:
                    conn.execute(f'DELETE FROM {table} WHERE task_id=?', (task_id,))
            for unit_id in equipment_units:
                if not conn.execute(
                    'SELECT 1 FROM equipment_unit_tasks WHERE unit_id=?', (unit_id,)
                ).fetchone() and conn.execute(
                    '''SELECT 1 FROM equipment_units units
                       WHERE units.id=? AND units.generation=1
                         AND units.replaced_by_id IS NULL
                         AND NOT EXISTS (
                             SELECT 1 FROM equipment_events events
                             WHERE events.unit_id=units.id
                               AND events.event_type!='discovered'
                         )
                         AND NOT EXISTS (
                             SELECT 1 FROM equipment_events events
                             WHERE events.related_unit_id=units.id
                         )''',
                    (unit_id,),
                ).fetchone():
                    conn.execute('DELETE FROM equipment_events WHERE unit_id=?', (unit_id,))
                    conn.execute('DELETE FROM equipment_units WHERE id=?', (unit_id,))
            conn.execute('DELETE FROM tasks WHERE ID=?', (task_id,))
    finally:
        conn.close()


def _read_problem_video(upload):
    if not upload or not upload.filename:
        return None
    if upload.mimetype not in {
        'video/mp4', 'video/quicktime', 'video/webm',
    }:
        raise ValueError('Видео должно быть в формате MP4, MOV или WebM')
    content = upload.read(20 * 1024 * 1024 + 1)
    if len(content) > 20 * 1024 * 1024:
        raise ValueError('Видео должно быть не больше 20 МБ')
    return {
        'content': content,
        'filename': upload.filename,
        'mimetype': upload.mimetype,
    }


def _employee_explanation(row):
    shifts = float(row.get('shifts', 0) or 0)
    metric_specs = (
        ('reviews', 'Отзывы', 1 / 3),
        ('forms', 'Анкеты', 1 / 3),
        ('extensions', 'Продления', 1 / 3),
        ('certificates', 'Сертификаты', 0.10),
        ('subscriptions', 'Абонементы', 0.10),
    )
    metrics = []
    for key, label, weight in metric_specs:
        fact = float(row.get(key, 0) or 0)
        plan_per_shift = float(row.get(f'{key}_plan_per_shift', 0) or 0)
        target = float(
            row.get(f'{key}_target', shifts * plan_per_shift) or 0
        )
        ratio = float(row.get(f'{key}_pct', 0) or 0)
        metrics.append({
            'key': key,
            'label': label,
            'fact': fact,
            'plan_per_shift': plan_per_shift,
            'target': target,
            'needed': max(target - fact, 0),
            'ratio': ratio,
            'weight': weight,
            'contribution_pct': ratio * weight,
        })

    initiative_fact = float(row.get('initiatives', 0) or 0)
    initiative_ratio = float(row.get('initiatives_pct', 0) or 0)
    metrics.append({
        'key': 'initiatives',
        'label': 'Инициативы',
        'fact': initiative_fact,
        'plan_per_shift': None,
        'bonus_per_item': float(
            row.get('initiative_bonus_per_item', 0.10) or 0.10
        ),
        'target': None,
        'needed': 0,
        'ratio': initiative_ratio,
        'weight': 1.0,
        'contribution_pct': initiative_ratio,
    })

    for goal in row.get('custom_goals', []):
        fact = float(goal.get('fact', 0) or 0)
        target = goal.get('target')
        target = float(target) if target is not None else None
        metrics.append({
            'key': f"hashtag:{goal['hashtag']}",
            'label': goal['name'],
            'fact': fact,
            'plan_per_shift': (
                float(goal['value'])
                if goal['calculation_type'] == 'per_shift_target'
                else None
            ),
            'bonus_per_item': (
                float(goal['contribution_pct'])
                if goal['calculation_type'] == 'per_unit'
                else None
            ),
            'target': target,
            'needed': max(target - fact, 0) if target is not None else 0,
            'ratio': goal.get('ratio'),
            'weight': float(goal.get('shift_share', 1) or 0),
            'contribution_pct': float(
                goal.get('actual_contribution_pct', 0) or 0
            ),
            'profile_shifts': float(goal.get('profile_shifts', 0) or 0),
            'shift_share': float(goal.get('shift_share', 1) or 0),
            'calculation_type': goal['calculation_type'],
            'calculation_type_label': goal['calculation_type_label'],
            'audience_label': goal['audience_label'],
            'unit_label': goal['unit_label'],
        })

    metric_contribution = sum(
        metric['contribution_pct']
        for metric in metrics
    )
    penalty_impact = float(row.get('penalty_impact', 0) or 0)
    stream_bonus = float(row.get('stream_bonus', 0) or 0)
    total_pct = float(row.get('total_pct', 0) or 0)
    average_pct = float(row.get('average_pct', 0) or 0)
    return {
        'metrics': metrics,
        'metric_contribution_pct': metric_contribution,
        'penalty_impact_pct': penalty_impact,
        'stream_bonus_pct': stream_bonus,
        'total_pct': total_pct,
        'pace': {
            'available': shifts > 0,
            'shifts': shifts,
            'projected_pct': total_pct if shifts > 0 else 0,
            'green_threshold_pct': average_pct,
            'gap_to_green_pct': max(average_pct - total_pct, 0),
        },
    }


def _attention_reasons(row):
    reasons = []
    if row.get('zone') == '🔴':
        reasons.append({
            'key': 'red_zone',
            'label': 'Красная зона',
        })
    if int(row.get('penalties', 0) or 0) > 0:
        reasons.append({
            'key': 'penalty',
            'label': 'Активный штраф',
        })
    if float(row.get('kpi_change_7d', 0) or 0) <= -0.10:
        reasons.append({
            'key': 'drop',
            'label': 'Падение KPI за 7 дней',
        })
    metric_total = sum(
        float(row.get(field, 0) or 0)
        for field in (
            'reviews',
            'forms',
            'extensions',
            'certificates',
            'subscriptions',
            'initiatives',
        )
    )
    metric_total += sum(
        float(goal.get('fact', 0) or 0)
        for goal in row.get('custom_goals', [])
    )
    if float(row.get('shifts', 0) or 0) > 0 and metric_total == 0:
        reasons.append({
            'key': 'no_kpi',
            'label': 'Есть смены, но нет KPI-записей',
        })
    return reasons


def _snapshot_employee(row):
    fields = (
        'login',
        'nickname',
        'role',
        'role_name',
        'clubs',
        'shifts',
        'weighted_shifts',
        'reviews',
        'reviews_pct',
        'forms',
        'forms_pct',
        'extensions',
        'extensions_pct',
        'certificates',
        'certificates_pct',
        'subscriptions',
        'subscriptions_pct',
        'initiatives',
        'initiatives_pct',
        'custom_goals',
        'custom_goals_pct',
        'penalties',
        'penalty_impact',
        'stream',
        'stream_bonus',
        'total_pct',
        'weighted_pct',
        'rank',
        'average_pct',
        'zone',
    )
    return {
        field: row.get(field)
        for field in fields
    }


def _month_close_preview(month):
    selected_day = _default_day(month)
    active_logins = _active_employee_logins()
    rows = calculate_monthly_kpi(
        month,
        employee_logins=active_logins,
        period_end=selected_day,
    )
    metadata = _employee_metadata(active_logins, month)
    for row in rows:
        row.update(metadata.get(row['login'], {
            'role': None,
            'role_name': 'Неизвестно',
            'clubs': [],
        }))

    participants = [row for row in rows if float(row.get('shifts', 0) or 0) > 0]
    no_shifts = [row for row in rows if float(row.get('shifts', 0) or 0) <= 0]
    red_zone = [row for row in participants if row.get('zone') == '🔴']
    penalties = [
        row
        for row in rows
        if int(row.get('penalties', 0) or 0) > 0
    ]
    no_kpi = []
    for row in participants:
        metric_total = sum(
            float(row.get(field, 0) or 0)
            for field in (
                'reviews',
                'forms',
                'extensions',
                'certificates',
                'subscriptions',
                'initiatives',
            )
        )
        metric_total += sum(
            float(goal.get('fact', 0) or 0)
            for goal in row.get('custom_goals', [])
        )
        if metric_total == 0:
            no_kpi.append(row)

    warning_specs = (
        ('no_shifts', 'Активные сотрудники без смен', no_shifts),
        ('no_kpi', 'Есть смены, но нет KPI-записей', no_kpi),
        ('red_zone', 'Сотрудники в красной зоне', red_zone),
        ('penalties', 'Сотрудники с активными штрафами', penalties),
    )
    warnings = [
        {
            'key': key,
            'label': label,
            'count': len(items),
            'employees': [
                {
                    'login': item.get('login'),
                    'nickname': item.get('nickname'),
                }
                for item in items
            ],
        }
        for key, label, items in warning_specs
        if items
    ]
    average = (
        sum(float(row.get('total_pct', 0) or 0) for row in participants)
        / len(participants)
        if participants else 0
    )
    zones = {
        zone: sum(row.get('zone') == zone for row in participants)
        for zone in ('🟢', '🟡', '🔴')
    }
    summary = {
        'active_employees': len(rows),
        'participants': len(participants),
        'average_pct': average,
        'zones': zones,
        'warnings': sum(warning['count'] for warning in warnings),
    }
    return {
        'period_month': month,
        'date': selected_day,
        'generated_at': datetime.now(UTC).isoformat(),
        'summary': summary,
        'warnings': warnings,
        'employees': [_snapshot_employee(row) for row in rows],
    }


@app.after_request
def add_security_headers(response):
    response.headers['Cache-Control'] = (
        'public, max-age=300'
        if request.path.startswith('/static/')
        else 'no-store'
    )
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    return response


@app.errorhandler(ValueError)
def handle_value_error(error):
    return jsonify({'error': str(error)}), 400


@app.errorhandler(413)
def handle_upload_too_large(_error):
    return jsonify({
        'error': 'Отправка слишком большая. Уменьшите размер фотографий или видео.'
    }), 413


@app.get('/')
def index():
    return send_from_directory(STATIC_DIR, 'home.html')


@app.get('/kpi')
def kpi_index():
    return send_from_directory(STATIC_DIR, 'index.html')


@app.get('/problems')
def problems_index():
    return send_from_directory(STATIC_DIR, 'problems.html')


@app.get('/shift-config')
def shift_config_index():
    return send_from_directory(STATIC_DIR, 'shift_config.html')


@app.get('/shift')
def shift_index():
    return send_from_directory(STATIC_DIR, 'shift.html')


@app.get('/shift-report')
@app.get('/shift-test')
def shift_test_index():
    return send_from_directory(STATIC_DIR, 'shift_test.html')


@app.get('/camera-test')
def camera_test_index():
    return send_from_directory(STATIC_DIR, 'camera_test.html')


@app.get('/static/<path:filename>')
def static_file(filename):
    return send_from_directory(STATIC_DIR, filename)


@app.get('/health')
def health():
    return jsonify({'status': 'ok'})


@app.get('/api/me')
@require_user
def api_me():
    role = int(g.kpi_user['status'])
    return jsonify({
        'telegram_id': g.kpi_user['chatid'],
        'login': g.kpi_user['login'],
        'name': (
            g.kpi_user.get('nick_name')
            or g.kpi_user.get('first_name')
            or g.kpi_user['login']
        ),
        'role': role,
        'role_name': ROLE_NAMES[role],
        'can_manage': role >= ROLE_MANAGER,
        'can_edit_settings': role == ROLE_OWNER,
    })


@app.get('/api/home')
@require_user
def api_home():
    today = _moscow_today()
    month = today.replace(day=1).isoformat()
    selected_day = today.isoformat()
    role = int(g.kpi_user['status'])
    login = _actor_login()
    payload = {
        'date': selected_day,
        'role': role,
        'upcoming_shifts': [],
        'personal_kpi': None,
        'management': None,
        'clubs': [],
    }

    if role != ROLE_OWNER:
        payload['upcoming_shifts'] = _upcoming_shifts(login)
        personal_rows = calculate_monthly_kpi(
            month,
            employee_logins=[login],
            period_end=selected_day,
        )
        if personal_rows:
            row = personal_rows[0]
            payload['personal_kpi'] = {
                'shifts': float(row.get('shifts', 0) or 0),
                'total_pct': float(row.get('total_pct', 0) or 0),
                'zone': row.get('zone'),
            }

    if role >= ROLE_MANAGER:
        team_rows, management = _team_snapshot(month, selected_day)
        management['problems'] = _task_counts()
        payload['management'] = management
        payload['clubs'] = _club_dashboard(
            team_rows,
            _problem_counts_by_club(),
        )
    return jsonify(payload)


@app.get('/api/bookings/today')
@require_user
def api_today_bookings():
    today = _moscow_today()
    role = int(g.kpi_user['status'])
    freshness = booking_freshness(db_path=DB_PATH)
    payload = {
        'date': today.isoformat(),
        'last_synced_at': freshness['last_synced_at'],
        'age_minutes': freshness['age_minutes'],
        'stale': freshness['stale'],
        'mode': 'clubs',
        'groups': [],
        'bookings': [],
    }

    if role >= ROLE_MANAGER:
        payload['mode'] = 'management'
        payload['groups'] = _club_booking_groups(
            list(BUKZA_CLUB_CODES.values()),
            today,
        )
        return jsonify(payload)

    shift_clubs = _today_shift_clubs(_actor_login())
    if 'Коллцентр' in shift_clubs:
        payload['mode'] = 'callcenter'
        payload['bookings'] = [
            _public_booking(order, include_order=True)
            for order in upcoming_unpaid_orders(db_path=DB_PATH)
        ]
        return jsonify(payload)

    physical_clubs = set(BUKZA_CLUB_CODES.values())
    payload['groups'] = _club_booking_groups(
        [club for club in shift_clubs if club in physical_clubs],
        today,
    )
    return jsonify(payload)


@app.get('/api/shift-config')
@require_manager
def api_shift_config():
    return jsonify(get_editor_config(DB_PATH))


@app.get('/api/shift')
@require_user
def api_shift():
    role = int(g.kpi_user['status'])
    today = _moscow_today()
    dashboard = None
    if role != ROLE_OWNER:
        dashboard = {
            'today': today.isoformat(),
            'upcoming_shifts': _upcoming_shifts(_actor_login(), limit=4),
            'month_summary': _shift_month_summary(
                _actor_login(), today.strftime('%Y-%m'),
            ),
        }
    return jsonify({
        'external_url': OMG_SHIFT_URL,
        'user_name': (
            g.kpi_user.get('nick_name')
            or g.kpi_user.get('first_name')
            or g.kpi_user['login']
        ),
        'role_name': ROLE_NAMES[role],
        'can_manage': role >= ROLE_MANAGER,
        'camera_test_available': bool(
            TELEGRAM_API_KEY and CAMERA_TEST_RECIPIENT_CHAT_ID
        ),
        'shift_report_available': bool(
            TELEGRAM_API_KEY and CHATS.get('reports') and CHATS.get('main_group')
        ),
        'employee_dashboard': dashboard,
    })


@app.get('/api/shift-test/scenario')
@require_user
def api_shift_test_scenario():
    action = str(request.args.get('action') or '').strip()
    variant = request.args.get('variant')
    requested_club = str(request.args.get('club') or '').strip() or None
    if (
        int(g.kpi_user['status']) == ROLE_OWNER
        and not requested_club
        and not _select_shift_report_test_shift(_actor_login())
    ):
        clubs = _shift_report_test_owner_clubs(action)
        if not clubs:
            raise ValueError('Не настроено ни одного клуба с таким сценарием')
        return jsonify({
            'production_mode': True,
            'requires_club_selection': True,
            'action': action,
            'action_label': 'Открытие' if action == 'open' else 'Закрытие',
            'clubs': clubs,
        })
    scenario = _shift_report_test_scenario(
        action,
        variant_index=variant if variant not in (None, '') else None,
        requested_club=requested_club,
    )
    return jsonify(scenario)


@app.post('/api/shift-test/start')
@require_user
def api_shift_test_start():
    if (
        not TELEGRAM_API_KEY
        or not CHATS.get('reports')
        or not CHATS.get('main_group')
    ):
        raise ValueError('Рабочие чаты открытия и закрытия не настроены')
    payload = request.get_json(silent=True) or {}
    action = str(payload.get('action') or '').strip()
    if payload.get('variant_index') is None:
        raise ValueError('В отчёте отсутствует набор сценария')
    scenario = _shift_report_test_scenario(
        action,
        variant_index=payload.get('variant_index'),
        requested_club=str(payload.get('club') or '').strip() or None,
    )
    if str(payload.get('version') or '') != scenario['version']:
        raise ValueError('Сценарий смены изменился. Начните заново')
    try:
        result = _start_shift_report_run(
            scenario,
            str(payload.get('run_id') or ''),
            early_confirmed=payload.get('early_confirmed') is True,
        )
    except RuntimeError as error:
        return jsonify({'error': str(error)}), 502
    if result.get('requires_early_confirmation'):
        return jsonify({
            'error': 'Закрытие начинается раньше установленного времени.',
            'code': 'early_close_confirmation_required',
        }), 409
    return jsonify(result)


@app.post('/api/shift-test/submit')
@require_user
def api_shift_test_submit():
    if not CHATS.get('reports') or not CHATS.get('main_group'):
        raise ValueError('Рабочие чаты открытия и закрытия не настроены')
    try:
        payload = json.loads(request.form.get('report') or '{}')
    except json.JSONDecodeError as error:
        raise ValueError('Отчёт передан неверно') from error
    action = str(payload.get('action') or '').strip()
    if payload.get('variant_index') is None:
        raise ValueError('В отчёте отсутствует набор сценария')
    scenario = _shift_report_test_scenario(
        action,
        variant_index=payload.get('variant_index'),
        requested_club=str(payload.get('club') or '').strip() or None,
    )
    answers, photos = _shift_report_test_data(scenario, payload)

    actor = f'{_actor_login()}:{action}'
    now = time.monotonic()
    with _shift_report_test_lock:
        previous = _shift_report_test_sent_at.get(actor, 0)
        if now - previous < CAMERA_TEST_COOLDOWN_SECONDS:
            return jsonify({
                'error': 'Подождите несколько секунд перед повторной отправкой.'
            }), 429
    try:
        result = _complete_shift_report_run(scenario, payload, answers, photos)
    except Exception as error:
        print(f'Ошибка завершения отчёта смены: {error}')
        return jsonify({
            'error': f'Не удалось завершить отчёт: {error}'
        }), 502
    with _shift_report_test_lock:
        _shift_report_test_sent_at[actor] = now
    return jsonify({
        'sent': True,
        'completed': result['completed'],
        'already_completed': result['already_completed'],
        'photos': len(photos),
    })


@app.post('/api/camera-test')
@require_user
def api_camera_test():
    if request.form.get('consent') != 'yes':
        raise ValueError('Подтвердите отправку тестового материала Павлу')
    if not CAMERA_TEST_RECIPIENT_CHAT_ID:
        raise ValueError('Получатель теста камеры не настроен')
    try:
        diagnostics = json.loads(request.form.get('diagnostics') or '{}')
    except json.JSONDecodeError as error:
        raise ValueError('Некорректная диагностика камеры') from error
    if not isinstance(diagnostics, dict):
        raise ValueError('Некорректная диагностика камеры')

    actor = _actor_login()
    now = time.monotonic()
    with _camera_test_lock:
        previous = _camera_test_sent_at.get(actor, 0)
        if now - previous < CAMERA_TEST_COOLDOWN_SECONDS:
            return jsonify({
                'error': 'Подождите несколько секунд перед следующим тестом.'
            }), 429

    media = _camera_test_upload()
    media_type = media['media_type'] if media else 'Только диагностика'
    media_size = len(media['content']) if media else 0
    caption = _camera_test_caption(diagnostics, media_type, media_size)
    bot = _notification_bot()
    if not bot:
        return jsonify({'error': 'Telegram-бот временно недоступен.'}), 503

    delivery = 'text'
    try:
        if not media:
            bot.send_message(
                CAMERA_TEST_RECIPIENT_CHAT_ID, caption, parse_mode='HTML',
            )
        else:
            media_file = io.BytesIO(media['content'])
            media_file.name = media['filename']
            if media['mimetype'] in {'image/jpeg', 'image/png', 'image/webp'}:
                bot.send_photo(
                    CAMERA_TEST_RECIPIENT_CHAT_ID,
                    media_file,
                    caption=caption,
                    parse_mode='HTML',
                )
                delivery = 'photo'
            elif media['mimetype'] == 'video/mp4':
                bot.send_video(
                    CAMERA_TEST_RECIPIENT_CHAT_ID,
                    media_file,
                    caption=caption,
                    parse_mode='HTML',
                    supports_streaming=True,
                )
                delivery = 'video'
            else:
                bot.send_document(
                    CAMERA_TEST_RECIPIENT_CHAT_ID,
                    media_file,
                    caption=caption,
                    parse_mode='HTML',
                )
                delivery = 'document'
    except Exception as error:
        print(f'Ошибка отправки теста камеры: {error}')
        return jsonify({
            'error': 'Не удалось отправить тест Павлу через Telegram.'
        }), 502

    with _camera_test_lock:
        _camera_test_sent_at[actor] = now
    return jsonify({'sent': True, 'delivery': delivery})


@app.put('/api/shift-config')
@require_manager
def api_save_shift_config():
    payload = request.get_json(silent=True) or {}
    try:
        saved = save_editor_config(DB_PATH, payload, _actor_login())
    except RuntimeError as error:
        return jsonify({'error': str(error)}), 409
    return jsonify(saved)


@app.get('/api/shift-config/history')
@require_manager
def api_shift_config_history():
    return jsonify({'versions': list_versions(DB_PATH)})


@app.post('/api/shift-config/history/<int:version_id>/rollback')
@require_manager
def api_shift_config_rollback(version_id):
    payload = request.get_json(silent=True) or {}
    try:
        saved = rollback_version(
            DB_PATH,
            version_id,
            payload.get('version'),
            _actor_login(),
        )
    except RuntimeError as error:
        return jsonify({'error': str(error)}), 409
    except LookupError as error:
        return jsonify({'error': str(error)}), 404
    return jsonify(saved)


def _task_payload(row, include_description=True):
    result = {
        'id': int(row['ID']),
        'date': row['dtrep'],
        'type': _normalize_legacy_text(row['type']),
        'club': row['club'],
        'title': _normalize_legacy_text(row['title']),
        'status': _normalize_legacy_text(row['status']),
        'closed_at': row['dtfb'],
        'has_photo': row['photo'] is not None,
        'has_video': 'has_video' in row.keys() and bool(row['has_video']),
    }
    if include_description:
        result.update({
            'description': _normalize_legacy_text(row['desc']),
            'feedback': _plain_task_feedback(
                _normalize_legacy_text(row['feedback'])
            ),
        })
    return result


@app.get('/api/problems')
@require_user
def api_problems():
    _initialize_problem_video_schema(DB_PATH)
    requested_status = str(request.args.get('status') or 'work').strip().lower()
    status_map = {
        'work': ('В работе',),
        'review': ('На проверке',),
        'done': ('Выполнено',),
    }
    if requested_status not in status_map:
        raise ValueError('Неизвестный фильтр статуса')
    placeholders, statuses = _status_sql(status_map[requested_status])
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            f'''SELECT tasks.*,
                       EXISTS (
                           SELECT 1 FROM task_videos videos
                           WHERE videos.task_id=tasks.ID
                       ) AS has_video
                FROM tasks WHERE status IN ({placeholders})
                ORDER BY date(dtrep) DESC, ID DESC LIMIT 200''',
            statuses,
        ).fetchall()
    finally:
        conn.close()
    return jsonify({
        'status': requested_status,
        'tasks': [_task_payload(row, include_description=False) for row in rows],
        'counts': _task_counts(),
    })


@app.get('/api/problems-meta')
@require_user
def api_problems_meta():
    return jsonify({
        'clubs': list(get_clubs()),
        'types': [
            GENERAL_TASK_TYPE,
            REPAIR_TASK_TYPE,
            BOT_TASK_TYPE,
        ],
        'can_process': int(g.kpi_user['status']) >= ROLE_TECHNICIAN,
        'can_view_analytics': int(g.kpi_user['status']) >= ROLE_TECHNICIAN,
        'can_export_analytics': int(g.kpi_user['status']) >= ROLE_MANAGER,
        'can_view_equipment': int(g.kpi_user['status']) >= ROLE_TECHNICIAN,
        'can_edit_repair_catalog': int(g.kpi_user['status']) >= ROLE_MANAGER,
        'repair_clubs': list(ZONE_COUNTS),
    })


@app.get('/api/problems/analytics')
@require_technician
def api_problem_analytics():
    mode = str(request.args.get('mode') or 'month').strip().lower()
    return jsonify(build_task_analytics(
        DB_PATH,
        mode=mode,
        month=request.args.get('month'),
        year=request.args.get('year'),
    ))


def _problem_report_from_request():
    return build_task_report(
        DB_PATH,
        mode=str(request.args.get('mode') or 'month').strip().lower(),
        month=request.args.get('month'),
        year=request.args.get('year'),
    )


@app.post('/api/problems/export/text')
@require_manager
def api_problem_report_text():
    report = _problem_report_from_request()
    chunks = format_task_report_html(report)
    bot = _notification_bot()
    if not bot:
        return jsonify({'error': 'Telegram-бот временно недоступен.'}), 503
    try:
        for chunk in chunks:
            bot.send_message(
                g.kpi_user['chatid'],
                chunk,
                parse_mode='HTML',
                disable_web_page_preview=True,
            )
    except Exception as error:
        print(f'Ошибка отправки отчёта Taskboard в Telegram: {error}')
        return jsonify({'error': 'Не удалось отправить отчёт в чат с ботом.'}), 502
    return jsonify({'sent': True, 'messages': len(chunks)})


@app.post('/api/problems/export/excel')
@require_manager
def api_problem_report_excel():
    report = _problem_report_from_request()
    workbook = _task_report_workbook(report)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    period = (
        report['period']['value']
        if report['period']['mode'] != 'all' else 'all'
    )
    filename = f'Taskboard_{period}.xlsx'
    output.name = filename
    bot = _notification_bot()
    if not bot:
        output.close()
        return jsonify({'error': 'Telegram-бот временно недоступен.'}), 503
    try:
        bot.send_document(
            g.kpi_user['chatid'],
            output,
            caption=f"🚩 Доска проблем · {report['period']['label']}",
        )
    except Exception as error:
        print(f'Ошибка отправки Excel Taskboard в Telegram: {error}')
        return jsonify({'error': 'Не удалось отправить Excel в чат с ботом.'}), 502
    finally:
        output.close()
    return jsonify({'sent': True, 'filename': filename})


@app.get('/api/repairs/catalog')
@require_user
def api_repair_catalog():
    club = str(request.args.get('club') or '').strip()
    if club and club not in get_clubs():
        raise ValueError('Выберите клуб')
    include_inactive = (
        request.args.get('include_inactive') == '1'
        and int(g.kpi_user['status']) >= ROLE_MANAGER
    )
    return jsonify(catalog_payload(DB_PATH, club, include_inactive))


@app.get('/api/equipment')
@require_technician
def api_equipment():
    return jsonify(equipment_list_payload(DB_PATH))


@app.get('/api/equipment/<int:unit_id>')
@require_technician
def api_equipment_unit(unit_id):
    initialize_repair_schema(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        result = equipment_payload(conn, unit_id)
    finally:
        conn.close()
    if not result:
        return jsonify({'error': 'Оборудование не найдено.'}), 404
    return jsonify(result)


@app.post('/api/equipment/<int:unit_id>/replace')
@require_technician
def api_replace_equipment(unit_id):
    payload = request.get_json(silent=True) or {}
    message = str(payload.get('message') or '').strip()
    if len(message) > 1000:
        raise ValueError('Комментарий должен быть не длиннее 1000 символов')
    close_multi_location_tasks = bool(payload.get('close_multi_location_tasks'))
    actor = task_actor_snapshot(g.kpi_user)
    current = datetime.now(ZoneInfo('Europe/Moscow'))
    event_at = current.isoformat(timespec='seconds')
    feedback_message = message or 'Оборудование заменено на новое'
    feedback_entry = (
        f'<b>[{current.strftime("%d.%m")}] Замена оборудования:</b> '
        f'{html.escape(feedback_message)}'
    )
    initialize_repair_schema(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        with conn:
            result = replace_equipment_unit(
                conn,
                unit_id,
                actor,
                message,
                close_multi_location_tasks,
                event_at,
                feedback_entry,
            )
    finally:
        conn.close()
    if not result:
        return jsonify({'error': 'Оборудование уже заменено или не найдено.'}), 409
    if result.get('requires_confirmation'):
        return jsonify(result), 409
    closed_tasks = result.pop('closed_tasks', [])
    result['closed_task_count'] = len(closed_tasks)
    for task in closed_tasks:
        _send_problem_notification(
            'completed', task, message=feedback_message, actor=actor,
        )
    return jsonify(result)


@app.post('/api/repairs/catalog/items')
@require_manager
def api_add_repair_item():
    payload = request.get_json(silent=True) or {}
    name = str(payload.get('name') or '').strip()
    if not name or len(name) > 80:
        raise ValueError('Название оборудования должно быть не длиннее 80 символов')
    initialize_repair_schema(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    try:
        with conn:
            cursor = conn.execute(
                '''INSERT INTO repair_item_types(name, sort_order)
                   VALUES (?, (SELECT COALESCE(MAX(sort_order), 0) + 1
                               FROM repair_item_types))''',
                (name,),
            )
    except sqlite3.IntegrityError as error:
        raise ValueError('Такое оборудование уже есть в списке') from error
    finally:
        conn.close()
    return jsonify({'id': cursor.lastrowid, 'name': name}), 201


@app.post('/api/repairs/catalog/details')
@require_manager
def api_add_repair_detail():
    payload = request.get_json(silent=True) or {}
    name = str(payload.get('name') or '').strip()
    try:
        item_id = int(payload.get('item_id'))
    except (TypeError, ValueError) as error:
        raise ValueError('Выберите оборудование') from error
    if not name or len(name) > 80:
        raise ValueError('Уточнение должно быть не длиннее 80 символов')
    initialize_repair_schema(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    try:
        with conn:
            if not conn.execute(
                'SELECT 1 FROM repair_item_types WHERE id=?', (item_id,)
            ).fetchone():
                raise ValueError('Оборудование не найдено')
            cursor = conn.execute(
                '''INSERT INTO repair_item_details(item_type_id, name, sort_order)
                   VALUES (?, ?, (SELECT COALESCE(MAX(sort_order), 0) + 1
                                   FROM repair_item_details
                                   WHERE item_type_id=?))''',
                (item_id, name, item_id),
            )
    except sqlite3.IntegrityError as error:
        raise ValueError('Такое уточнение уже есть в списке') from error
    finally:
        conn.close()
    return jsonify({'id': cursor.lastrowid, 'name': name}), 201


@app.post('/api/repairs/catalog/locations')
@require_manager
def api_add_repair_location():
    payload = request.get_json(silent=True) or {}
    club = str(payload.get('club') or '').strip()
    name = str(payload.get('name') or '').strip()
    if club not in get_clubs():
        raise ValueError('Выберите клуб')
    if not name or len(name) > 80:
        raise ValueError('Название места должно быть не длиннее 80 символов')
    initialize_repair_schema(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    try:
        with conn:
            cursor = conn.execute(
                '''INSERT INTO repair_locations(club, name, kind, sort_order)
                   VALUES (?, ?, 'other',
                           (SELECT COALESCE(MAX(sort_order), 0) + 1
                            FROM repair_locations WHERE club=?))''',
                (club, name, club),
            )
    except sqlite3.IntegrityError as error:
        raise ValueError('Такое место уже есть в клубе') from error
    finally:
        conn.close()
    return jsonify({'id': cursor.lastrowid, 'name': name}), 201


@app.patch('/api/repairs/catalog/<kind>/<int:entry_id>')
@require_manager
def api_toggle_repair_catalog_entry(kind, entry_id):
    table_map = {
        'items': 'repair_item_types',
        'details': 'repair_item_details',
        'locations': 'repair_locations',
    }
    if kind not in table_map:
        raise ValueError('Неизвестный раздел справочника')
    payload = request.get_json(silent=True) or {}
    active = 1 if bool(payload.get('active')) else 0
    initialize_repair_schema(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    try:
        with conn:
            cursor = conn.execute(
                f'UPDATE {table_map[kind]} SET active=? WHERE id=?',
                (active, entry_id),
            )
            if cursor.rowcount != 1:
                return jsonify({'error': 'Запись справочника не найдена.'}), 404
    finally:
        conn.close()
    return jsonify({'id': entry_id, 'active': bool(active)})


@app.get('/api/repairs/migration-review')
@require_manager
def api_repair_migration_review():
    return jsonify(migration_review_payload(DB_PATH))


@app.post('/api/repairs/<int:task_id>/mapping')
@require_manager
def api_map_legacy_repair(task_id):
    payload = request.get_json(silent=True) or {}
    try:
        item_id = int(payload.get('item_id'))
        detail_id = int(payload['detail_id']) if payload.get('detail_id') else None
        location_ids = payload.get('location_ids') or []
        if not isinstance(location_ids, list):
            raise ValueError
    except (TypeError, ValueError) as error:
        raise ValueError('Выберите оборудование и хотя бы одно место') from error
    initialize_repair_schema(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        with conn:
            task = conn.execute(
                '''SELECT ID, club FROM tasks
                   WHERE ID=? AND lower(type)=lower('Ремонт')''',
                (task_id,),
            ).fetchone()
            if not task:
                return jsonify({'error': 'Ремонтная заявка не найдена.'}), 404
            if conn.execute(
                'SELECT 1 FROM repair_cases WHERE task_id=?', (task_id,)
            ).fetchone():
                return jsonify({'error': 'Заявка уже сопоставлена.'}), 409
            create_repair_case(
                conn, task_id, task['club'], item_id, detail_id, location_ids,
            )
            conn.execute(
                "UPDATE repair_cases SET mapping_source='legacy-manual' WHERE task_id=?",
                (task_id,),
            )
            conn.execute('DELETE FROM repair_events WHERE task_id=?', (task_id,))
    finally:
        conn.close()
    return jsonify({'task_id': task_id, 'mapped': True})


@app.get('/api/problems/<int:task_id>')
@require_user
def api_problem(task_id):
    initialize_repair_schema(DB_PATH)
    initialize_task_analytics_schema(DB_PATH)
    _initialize_problem_video_schema(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute(
            '''SELECT tasks.*,
                      EXISTS (
                          SELECT 1 FROM task_videos videos
                          WHERE videos.task_id=tasks.ID
                      ) AS has_video
               FROM tasks WHERE ID=?''',
            (task_id,),
        ).fetchone()
        repair = repair_payload(conn, task_id) if row else None
        activity = task_activity_payload(conn, task_id) if row else []
    finally:
        conn.close()
    if not row:
        return jsonify({'error': 'Проблема не найдена.'}), 404
    result = _task_payload(row)
    result['repair'] = repair
    result['activity'] = activity
    return jsonify(result)


@app.get('/api/problems/<int:task_id>/photo')
@require_user
def api_problem_photo(task_id):
    conn = sqlite3.connect(DB_PATH)
    try:
        row = conn.execute('SELECT photo FROM tasks WHERE ID=?', (task_id,)).fetchone()
    finally:
        conn.close()
    if not row or row[0] is None:
        return jsonify({'error': 'Фотография не найдена.'}), 404
    photo = row[0]
    if photo.startswith(b'\x89PNG\r\n\x1a\n'):
        mimetype = 'image/png'
    elif photo.startswith(b'RIFF') and photo[8:12] == b'WEBP':
        mimetype = 'image/webp'
    else:
        mimetype = 'image/jpeg'
    return send_file(io.BytesIO(photo), mimetype=mimetype, max_age=0)


@app.get('/api/problems/<int:task_id>/video')
@require_user
def api_problem_video(task_id):
    _initialize_problem_video_schema(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    try:
        row = conn.execute(
            '''SELECT telegram_file_id, mime_type FROM task_videos
               WHERE task_id=?''',
            (task_id,),
        ).fetchone()
    finally:
        conn.close()
    if not row:
        return jsonify({'error': 'Видео не найдено.'}), 404
    bot = _notification_bot()
    if not bot:
        return jsonify({'error': 'Видео временно недоступно.'}), 503
    try:
        file_info = bot.get_file(row[0])
        video = bot.download_file(file_info.file_path)
    except Exception as error:
        print(f'Ошибка загрузки видео Taskboard: {error}')
        return jsonify({'error': 'Не удалось загрузить видео.'}), 502
    return send_file(io.BytesIO(video), mimetype=row[1], max_age=0)


@app.post('/api/problems')
@require_user
def api_create_problem():
    task_type = str(request.form.get('type') or '').strip()
    club = str(request.form.get('club') or '').strip()
    title = str(request.form.get('title') or '').strip()
    description = str(request.form.get('description') or '').strip()
    allowed_types = {GENERAL_TASK_TYPE, REPAIR_TASK_TYPE, BOT_TASK_TYPE}
    if task_type not in allowed_types:
        raise ValueError('Выберите тип обращения')
    if club not in get_clubs():
        raise ValueError('Выберите клуб')
    if task_type == REPAIR_TASK_TYPE:
        title = 'Ремонт'
    if not title or len(title) > 50 or title.isnumeric():
        raise ValueError('Название должно содержать текст и быть не длиннее 50 символов')
    if not description or len(description) > 1000:
        raise ValueError('Описание должно быть не длиннее 1000 символов')
    item_id = detail_id = None
    location_ids = []
    if task_type == REPAIR_TASK_TYPE:
        try:
            item_id = int(request.form.get('repair_item_id'))
            detail_id = (
                int(request.form.get('repair_detail_id'))
                if request.form.get('repair_detail_id') else None
            )
            location_ids = json.loads(request.form.get('repair_location_ids') or '[]')
            if not isinstance(location_ids, list):
                raise ValueError
        except (TypeError, ValueError, json.JSONDecodeError) as error:
            raise ValueError('Выберите оборудование и хотя бы одно место') from error
    upload = request.files.get('photo')
    video_upload = request.files.get('video')
    photo = None
    if upload and upload.filename:
        if upload.mimetype not in {'image/jpeg', 'image/png', 'image/webp'}:
            raise ValueError('Фото должно быть в формате JPG, PNG или WebP')
        photo = upload.read(6 * 1024 * 1024 + 1)
        if len(photo) > 6 * 1024 * 1024:
            raise ValueError('Фото должно быть не больше 6 МБ')
    video = _read_problem_video(video_upload)
    if video:
        if photo is not None:
            raise ValueError('Прикрепите либо фото, либо видео')

    if task_type == REPAIR_TASK_TYPE:
        initialize_repair_schema(DB_PATH)
    _initialize_problem_video_schema(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    actor = task_actor_snapshot(g.kpi_user) if task_type == REPAIR_TASK_TYPE else None
    try:
        with conn:
            cursor = conn.execute(
                '''INSERT INTO tasks (
                       dtrep, type, club, title, photo, desc, status
                   ) VALUES (?, ?, ?, ?, ?, ?, 'В работе')''',
                (_moscow_today().isoformat(), task_type, club, title, photo, description),
            )
            task_id = cursor.lastrowid
            record_task_event(conn, task_id, 'created', actor=actor)
            if task_type == REPAIR_TASK_TYPE:
                title = create_repair_case(
                    conn, task_id, club, item_id, detail_id, location_ids,
                )
                conn.execute('UPDATE tasks SET title=? WHERE ID=?', (title, task_id))
            task = {
                'title': title,
                'type': task_type,
                'club': club,
                'description': description,
            }
    finally:
        conn.close()
    if video:
        video_reference = _send_problem_notification(
            'created', task, video=video, actor=actor,
        )
        if not video_reference:
            _delete_problem_after_failed_video(task_id, DB_PATH)
            raise ValueError(
                'Не удалось сохранить видео в Telegram. Попробуйте ещё раз.'
            )
        conn = sqlite3.connect(DB_PATH)
        try:
            with conn:
                conn.execute(
                    '''INSERT INTO task_videos (
                           task_id, telegram_file_id, telegram_file_unique_id,
                           mime_type, file_size
                       ) VALUES (?, ?, ?, ?, ?)''',
                    (
                        task_id,
                        video_reference['file_id'],
                        video_reference.get('file_unique_id'),
                        video_reference['mimetype'],
                        video_reference['file_size'],
                    ),
                )
        finally:
            conn.close()
    else:
        _send_problem_notification('created', task, photo=photo, actor=actor)
    return jsonify({'id': task_id, 'status': 'В работе'}), 201


def _change_problem_status(
    task_id, expected_status, new_status, entry=None, actor=None,
):
    initialize_repair_schema(DB_PATH)
    placeholders, statuses = _status_sql((expected_status,))
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        with conn:
            task = conn.execute(
                f'''SELECT * FROM tasks WHERE ID=?
                    AND status IN ({placeholders})''',
                (task_id, *statuses),
            ).fetchone()
            if not task:
                return None
            feedback = str(task['feedback'] or '').strip()
            if entry:
                feedback = f'{feedback}\n\n{entry}'.strip()
            status_date = (
                _moscow_today().isoformat()
                if new_status in {'На проверке', 'Выполнено'} else None
            )
            cursor = conn.execute(
                f'''UPDATE tasks SET status=?, feedback=?, dtfb=?
                    WHERE ID=? AND status IN ({placeholders})''',
                (new_status, feedback, status_date, task_id, *statuses),
            )
            if cursor.rowcount != 1:
                return None
            event_types = {
                'На проверке': 'solution',
                'В работе': 'returned',
                'Выполнено': 'confirmed',
            }
            add_repair_event(
                conn, task_id, event_types.get(new_status, 'status_changed'),
                _plain_task_feedback(entry) if entry else None,
            )
            record_task_event(
                conn,
                task_id,
                event_types.get(new_status, 'solution'),
                actor=actor,
            )
            return dict(task)
    finally:
        conn.close()


@app.post('/api/problems/<int:task_id>/solution')
@require_user
def api_problem_solution(task_id):
    if int(g.kpi_user['status']) < ROLE_TECHNICIAN:
        return jsonify({'error': 'Обрабатывать проблемы может ремонтник или менеджер.'}), 403
    payload = request.get_json(silent=True) or {}
    solution = str(payload.get('message') or '').strip()
    if not solution or len(solution) > 1000:
        raise ValueError('Решение должно быть не длиннее 1000 символов')
    today_short = datetime.now(ZoneInfo('Europe/Moscow')).strftime('%d.%m')
    actor = task_actor_snapshot(g.kpi_user)
    task = _change_problem_status(
        task_id,
        'В работе',
        'На проверке',
        f'<b>[{today_short}] Админ:</b> {html.escape(solution)}',
        actor=actor,
    )
    if not task:
        return jsonify({'error': 'Статус проблемы уже изменился.'}), 409
    _send_problem_notification('solution', task, message=solution, actor=actor)
    return jsonify({'status': 'На проверке'})


@app.post('/api/problems/<int:task_id>/confirm')
@require_user
def api_problem_confirm(task_id):
    actor = task_actor_snapshot(g.kpi_user)
    task = _change_problem_status(
        task_id,
        'На проверке',
        'Выполнено',
        actor=actor,
    )
    if not task:
        return jsonify({'error': 'Статус проблемы уже изменился.'}), 409
    _send_problem_notification('completed', task, actor=actor)
    return jsonify({'status': 'Выполнено'})


@app.post('/api/problems/<int:task_id>/return')
@require_user
def api_problem_return(task_id):
    payload = request.get_json(silent=True) or {}
    reason = str(payload.get('message') or '').strip()
    if not reason or len(reason) > 1000:
        raise ValueError('Укажите причину возврата не длиннее 1000 символов')
    today_short = datetime.now(ZoneInfo('Europe/Moscow')).strftime('%d.%m')
    actor = task_actor_snapshot(g.kpi_user)
    task = _change_problem_status(
        task_id,
        'На проверке',
        'В работе',
        f'<b>[{today_short}] Сотрудник:</b> {html.escape(reason)}',
        actor=actor,
    )
    if not task:
        return jsonify({'error': 'Статус проблемы уже изменился.'}), 409
    _send_problem_notification('returned', task, message=reason, actor=actor)
    return jsonify({'status': 'В работе'})


@app.get('/api/kpi')
@require_user
def api_kpi():
    month = _validate_month(
        request.args.get('month') or date.today().strftime('%Y-%m')
    )
    selected_day = _validate_day(
        request.args.get('date') or _default_day(month),
        month,
    )
    current_login = _actor_login()
    cache_key = (month, selected_day, current_login)
    with _kpi_cache_lock:
        cached = _kpi_cache.get(cache_key)
    if cached and time.monotonic() - cached['created_at'] < KPI_CACHE_SECONDS:
        return jsonify(cached['payload'])

    active_logins = _active_employee_logins()
    employee_logins = active_logins
    rows = calculate_monthly_kpi(
        month,
        employee_logins=employee_logins,
        period_end=selected_day,
    )
    previous_rows = []
    selected_date = datetime.strptime(selected_day, '%Y-%m-%d').date()
    if selected_date.day > 1:
        previous_rows = calculate_monthly_kpi(
            month,
            employee_logins=employee_logins,
            period_end=selected_date - timedelta(days=1),
        )
    previous_by_login = {row['login']: row for row in previous_rows}
    comparison_date = max(
        datetime.strptime(month, '%Y-%m-%d').date(),
        selected_date - timedelta(days=7),
    )
    week_rows = []
    if comparison_date < selected_date:
        week_rows = calculate_monthly_kpi(
            month,
            employee_logins=employee_logins,
            period_end=comparison_date,
        )
    week_by_login = {row['login']: row for row in week_rows}
    metadata = _employee_metadata(employee_logins, month)
    hashtag_summaries = get_hashtag_summaries(
        employee_logins,
        month,
        period_end=selected_day,
        db_path=DB_PATH,
    )
    for row in rows:
        previous = previous_by_login.get(row['login'])
        week_previous = week_by_login.get(row['login'])
        row['kpi_change'] = (
            row['total_pct'] - previous['total_pct']
            if previous else None
        )
        row['rank_change'] = (
            previous['rank'] - row['rank']
            if (
                previous
                and previous['rank'] is not None
                and row['rank'] is not None
            )
            else None
        )
        row['kpi_change_7d'] = (
            row['total_pct'] - week_previous['total_pct']
            if week_previous else None
        )
        row.update(metadata.get(row['login'], {
            'role': None,
            'role_name': 'Неизвестно',
            'clubs': [],
        }))
        row['attention_reasons'] = _attention_reasons(row)
        row['needs_attention'] = bool(row['attention_reasons'])
        goal_hashtags = {
            goal['hashtag'] for goal in row.get('custom_goals', [])
        }
        row['extra_hashtags'] = [
            item for item in hashtag_summaries.get(row['login'], [])
            if item.get('hashtag') not in goal_hashtags
        ]
        row['explanation'] = _employee_explanation(row)
    active_login_set = set(active_logins)
    penalties = [
        item
        for item in list_penalties(month)
        if str(item.get('employee_login') or '').strip().lower()
        in active_login_set
    ]
    current_employee = next(
        (
            row
            for row in rows
            if str(row.get('login') or '').strip().lower() == current_login
        ),
        None,
    )
    if current_employee is None and current_login in set(active_logins):
        personal_rows = calculate_monthly_kpi(
            month,
            employee_logins=[current_login],
            period_end=selected_day,
        )
        if personal_rows:
            current_employee = personal_rows[0]
            current_employee['kpi_change'] = None
            current_employee['rank_change'] = None
            goal_hashtags = {
                goal['hashtag']
                for goal in current_employee.get('custom_goals', [])
            }
            current_employee['extra_hashtags'] = [
                item for item in hashtag_summaries.get(
                    current_employee['login'], []
                )
                if item.get('hashtag') not in goal_hashtags
            ]
            current_employee['explanation'] = _employee_explanation(
                current_employee
            )
    freshness = get_kpi_freshness(month, period_end=selected_day)
    freshness['calculated_at'] = datetime.now(UTC).isoformat()
    payload = {
        'month': month[:7],
        'date': selected_day,
        'month_status': get_month_status(month),
        'employees': rows,
        'penalties': penalties,
        'my_kpi': current_employee,
        'freshness': freshness,
    }
    with _kpi_cache_lock:
        expired_keys = [
            key
            for key, value in _kpi_cache.items()
            if time.monotonic() - value['created_at'] >= KPI_CACHE_SECONDS
        ]
        for key in expired_keys:
            _kpi_cache.pop(key, None)
        _kpi_cache[cache_key] = {
            'created_at': time.monotonic(),
            'payload': payload,
        }
    return jsonify(payload)


@app.post('/api/kpi/export')
@require_manager
def api_kpi_export():
    month = _validate_month(
        request.args.get('month') or date.today().strftime('%Y-%m')
    )
    month_start = datetime.strptime(month, '%Y-%m-%d').date()
    month_end = month_start.replace(
        day=calendar.monthrange(month_start.year, month_start.month)[1],
    ).isoformat()
    employee_logins = _employee_logins_with_month_shifts(
        _active_employee_logins(), month,
    )
    rows = calculate_monthly_kpi(
        month,
        employee_logins=employee_logins,
        period_end=month_end,
    )
    workbook = _kpi_export_workbook(rows, month)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    filename = f'KPI_{month[:7]}.xlsx'
    output.name = filename
    bot = _notification_bot()
    if not bot:
        output.close()
        return jsonify({'error': 'Telegram-бот временно недоступен.'}), 503
    try:
        bot.send_document(
            g.kpi_user['chatid'],
            output,
            caption=(
                f'📊 KPI за '
                f'{KPI_EXPORT_MONTH_NAMES[month_start.month - 1]} '
                f'{month_start.year}'
            ),
        )
    except Exception as error:
        print(f'Ошибка отправки Excel KPI в Telegram: {error}')
        return jsonify({'error': 'Не удалось отправить Excel в чат с ботом.'}), 502
    finally:
        output.close()
    return jsonify({'sent': True, 'filename': filename})


@app.get('/api/kpi/details')
@require_user
def api_kpi_details():
    month = _validate_month(request.args.get('month'))
    selected_day = _validate_day(request.args.get('date'), month)
    employee_login = str(request.args.get('employee_login') or '').strip()
    metric = str(request.args.get('metric') or '').strip()
    if employee_login.lower() not in set(_active_employee_logins()):
        raise ValueError('Сотрудник не найден или неактивен')
    if metric.startswith('hashtag:'):
        entries = get_hashtag_entries(
            employee_login,
            month,
            metric.removeprefix('hashtag:'),
            period_end=selected_day,
            db_path=DB_PATH,
        )
    else:
        entries = get_metric_entries(
            employee_login,
            month,
            metric,
            period_end=selected_day,
        )
    return jsonify({
        'employee_login': employee_login.lower(),
        'metric': metric,
        'date': selected_day,
        'entries': entries,
    })


@app.get('/api/kpi/analytics')
@require_user
def api_kpi_analytics():
    month = _validate_month(
        request.args.get('month') or date.today().strftime('%Y-%m')
    )
    mode = str(request.args.get('mode') or 'daily').strip().lower()
    if mode not in {'daily', 'monthly'}:
        raise ValueError('Режим аналитики должен быть daily или monthly')

    initialize_kpi_calculation_schema()
    selected_end = (
        _validate_day(
            request.args.get('date') or _default_day(month),
            month,
        )
        if mode == 'daily' else ''
    )
    cache_key = (
        mode,
        month,
        selected_end,
    )
    with _analytics_cache_lock:
        cached = _analytics_cache.get(cache_key)
    if cached and time.monotonic() - cached['created_at'] < ANALYTICS_CACHE_SECONDS:
        return jsonify(cached['payload'])

    active_logins = _active_employee_logins()
    periods = []
    if mode == 'daily':
        end_date = datetime.strptime(selected_end, '%Y-%m-%d').date()
        periods = [
            (month, date(month_date.year, month_date.month, day).isoformat())
            for month_date in [datetime.strptime(month, '%Y-%m-%d').date()]
            for day in range(1, end_date.day + 1)
        ]
    else:
        periods = [
            (period_month, calendar.monthrange(
                int(period_month[:4]),
                int(period_month[5:7]),
            )[1])
            for period_month in (
                _shift_month(month, offset)
                for offset in range(-11, 1)
            )
        ]
        periods = [
            (
                period_month,
                f'{period_month[:7]}-{month_day:02d}',
            )
            for period_month, month_day in periods
        ]

    points = []
    employees = {}
    if mode == 'daily':
        employee_logins = _employee_logins_with_month_shifts(
            active_logins,
            month,
        )
        daily_series = calculate_daily_kpi_series(
            month,
            periods[-1][1],
            employee_logins=employee_logins,
            ensure_schema=False,
        )
        period_rows = [
            (snapshot['date'], snapshot['employees'])
            for snapshot in daily_series
        ]
    else:
        period_rows = []
        for period_month, period_end in periods:
            employee_logins = _employee_logins_with_month_shifts(
                active_logins,
                period_month,
            )
            rows = calculate_monthly_kpi(
                period_month,
                employee_logins=employee_logins,
                period_end=period_end,
                ensure_schema=False,
            )
            period_rows.append((period_end, rows))

    for period_end, rows in period_rows:
        for row in rows:
            employees[row['login']] = {
                'login': row['login'],
                'nickname': row['nickname'],
            }
        points.append(_analytics_point(period_end, rows))

    payload = {
        'mode': mode,
        'month': month[:7],
        'employees': sorted(
            employees.values(),
            key=lambda employee: employee['nickname'].lower(),
        ),
        'points': points,
    }
    with _analytics_cache_lock:
        expired_keys = [
            key
            for key, value in _analytics_cache.items()
            if time.monotonic() - value['created_at'] >= ANALYTICS_CACHE_SECONDS
        ]
        for key in expired_keys:
            _analytics_cache.pop(key, None)
        _analytics_cache[cache_key] = {
            'created_at': time.monotonic(),
            'payload': payload,
        }
        while len(_analytics_cache) > 12:
            oldest_key = min(
                _analytics_cache,
                key=lambda key: _analytics_cache[key]['created_at'],
            )
            _analytics_cache.pop(oldest_key, None)
    return jsonify(payload)


@app.get('/api/kpi/settings')
@require_owner
def api_kpi_settings():
    month = _validate_month(
        request.args.get('month') or date.today().strftime('%Y-%m')
    )
    return jsonify(get_kpi_settings(month))


@app.put('/api/kpi/settings')
@require_owner
def api_save_kpi_settings():
    payload = request.get_json(silent=True) or {}
    month = _validate_month(payload.get('month'))
    raw_metrics = payload.get('metrics')
    raw_clubs = payload.get('clubs')
    if not isinstance(raw_metrics, dict) or not isinstance(raw_clubs, dict):
        raise ValueError('Передайте нормы показателей и веса клубов')

    metrics = {
        str(metric): value
        for metric, value in raw_metrics.items()
    }
    clubs = {}
    for club, weights in raw_clubs.items():
        if not isinstance(weights, dict):
            raise ValueError('Для каждого клуба нужны веса будней и выходных')
        clubs[str(club)] = (
            weights.get('weekday_weight'),
            weights.get('weekend_weight'),
        )

    settings = save_kpi_settings(
        month,
        metrics,
        clubs,
        _actor_login(),
    )
    _clear_analytics_cache()
    return jsonify(settings)


@app.post('/api/kpi/goals')
@require_owner
def api_save_custom_goal():
    payload = request.get_json(silent=True) or {}
    month = _validate_month(payload.get('month'))
    goal_payload = dict(payload)
    goal_payload.pop('month', None)
    goal_key = save_custom_goal(month, goal_payload, _actor_login())
    _clear_analytics_cache()
    settings = get_kpi_settings(month)
    settings['saved_goal_key'] = goal_key
    return jsonify(settings)


@app.post('/api/kpi/goals/<goal_key>/disable')
@require_owner
def api_disable_custom_goal(goal_key):
    payload = request.get_json(silent=True) or {}
    month = _validate_month(payload.get('month'))
    disable_custom_goal(goal_key, month, _actor_login())
    _clear_analytics_cache()
    return jsonify(get_kpi_settings(month))


@app.post('/api/penalties')
@require_manager
def api_add_penalty():
    payload = request.get_json(silent=True) or {}
    month = _validate_month(payload.get('month'))
    employee_login = str(payload.get('employee_login') or '').strip()
    reason = str(payload.get('reason') or '').strip()
    if not employee_login:
        raise ValueError('Выберите сотрудника')
    if employee_login.lower() not in set(_active_employee_logins()):
        raise ValueError('Сотрудник не найден или неактивен')
    penalty_id = add_penalty(
        employee_login,
        month,
        reason,
        _actor_login(),
        source='telegram_mini_app',
    )
    _clear_analytics_cache()
    return jsonify({'id': penalty_id, 'impact_pct': 0.10}), 201


@app.post('/api/penalties/<int:penalty_id>/cancel')
@require_manager
def api_cancel_penalty(penalty_id):
    payload = request.get_json(silent=True) or {}
    reason = str(payload.get('reason') or '').strip()
    if not cancel_penalty(penalty_id, _actor_login(), reason):
        return jsonify({'error': 'Активный штраф не найден.'}), 404
    _clear_analytics_cache()
    return jsonify({'status': 'cancelled'})


@app.post('/api/streams')
@require_manager
def api_set_stream():
    payload = request.get_json(silent=True) or {}
    month = _validate_month(payload.get('month'))
    employee_login = str(payload.get('employee_login') or '').strip()
    if not employee_login:
        raise ValueError('Выберите сотрудника')
    if employee_login.lower() not in set(_active_employee_logins()):
        raise ValueError('Сотрудник не найден или неактивен')
    set_monthly_stream(
        employee_login,
        month,
        bool(payload.get('enabled')),
        _actor_login(),
        source='telegram_mini_app',
    )
    _clear_analytics_cache()
    return jsonify({'status': 'saved'})


@app.post('/api/month-status')
@require_manager
def api_month_status():
    payload = request.get_json(silent=True) or {}
    month = _validate_month(payload.get('month'))
    is_closed = bool(payload.get('is_closed'))
    snapshot = _month_close_preview(month) if is_closed else None
    status = set_month_status(
        month,
        is_closed,
        _actor_login(),
        snapshot=snapshot,
    )
    _clear_analytics_cache()
    return jsonify(status)


@app.get('/api/month-close-check')
@require_manager
def api_month_close_check():
    month = _validate_month(request.args.get('month'))
    preview = _month_close_preview(month)
    return jsonify({
        'period_month': preview['period_month'],
        'date': preview['date'],
        'generated_at': preview['generated_at'],
        'summary': preview['summary'],
        'warnings': preview['warnings'],
    })


def main():
    initialize_kpi_calculation_schema()
    port = int(os.getenv('KPI_WEBAPP_PORT', '8080'))
    app.run(host='0.0.0.0', port=port, debug=False)


if __name__ == '__main__':
    main()
