import hashlib
import hmac
import json
import sqlite3
import tempfile
import time
import unittest
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import urlencode
from unittest.mock import Mock, patch
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

import kpi_web


BOT_TOKEN = '123456:test-token'


def signed_init_data(telegram_id=1001):
    values = {
        'auth_date': str(int(time.time())),
        'query_id': 'test-query',
        'user': json.dumps(
            {'id': telegram_id, 'first_name': 'Test'},
            separators=(',', ':'),
        ),
    }
    data_check_string = '\n'.join(
        f'{key}={values[key]}'
        for key in sorted(values)
    )
    secret = hmac.new(
        b'WebAppData',
        BOT_TOKEN.encode(),
        hashlib.sha256,
    ).digest()
    values['hash'] = hmac.new(
        secret,
        data_check_string.encode(),
        hashlib.sha256,
    ).hexdigest()
    return urlencode(values)


def user(role):
    return {
        'chatid': '1001',
        'login': '@tester',
        'nick_name': 'Тестер',
        'first_name': 'Test',
        'status': role,
    }


class KpiWebTest(unittest.TestCase):
    def setUp(self):
        kpi_web.app.config['TESTING'] = True
        kpi_web._kpi_cache.clear()
        self.client = kpi_web.app.test_client()
        self.headers = {'X-Telegram-Init-Data': signed_init_data()}
        self.membership_patch = patch.object(
            kpi_web,
            'is_main_group_member',
            return_value=True,
        )
        self.membership_patch.start()

    def tearDown(self):
        self.membership_patch.stop()

    def test_init_data_signature_and_age_are_validated(self):
        self.assertEqual(
            kpi_web._validate_init_data(
                signed_init_data(777),
                BOT_TOKEN,
            )['telegram_id'],
            777,
        )
        self.assertIsNone(kpi_web._validate_init_data('hash=wrong', BOT_TOKEN))

    def test_active_user_outside_main_group_cannot_use_api(self):
        self.membership_patch.stop()
        self.membership_patch = patch.object(
            kpi_web,
            'is_main_group_member',
            return_value=False,
        )
        self.membership_patch.start()
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(3)),
        ):
            response = self.client.get('/api/me', headers=self.headers)

        self.assertEqual(response.status_code, 401)

    def test_module_pages_load_shared_swipe_navigation(self):
        for path in ('/kpi', '/problems', '/shift', '/shift-config'):
            response = self.client.get(path)
            try:
                self.assertEqual(response.status_code, 200)
                self.assertIn(
                    b'/static/swipe_navigation.js', response.data,
                )
            finally:
                response.close()

        script = self.client.get('/static/swipe_navigation.js')
        try:
            self.assertEqual(script.status_code, 200)
            self.assertEqual(
                script.headers['Cache-Control'],
                'public, max-age=300',
            )
            self.assertIn(b"navigate('/')", script.data)
        finally:
            script.close()

        page = self.client.get('/kpi')
        try:
            self.assertEqual(page.headers['Cache-Control'], 'no-store')
        finally:
            page.close()

    def test_kpi_and_taskboard_mobile_controls_use_full_width_layouts(self):
        response = self.client.get('/kpi')
        try:
            kpi_html = response.get_data(as_text=True)
            self.assertNotIn('class="home-link"', kpi_html)
            self.assertIn('id="kpiUserName"', kpi_html)
            self.assertLess(
                kpi_html.index('id="dateDisplay"'),
                kpi_html.index('id="managerFilters"'),
            )
            self.assertLess(
                kpi_html.index('id="managerFilters"'),
                kpi_html.index('id="summary"'),
            )
            self.assertLess(
                kpi_html.index('class="sort-controls"'),
                kpi_html.index('id="searchInput"'),
            )
            self.assertLess(
                kpi_html.index('class="settings-card custom-goals-settings"'),
                kpi_html.index('id="settingsMetrics"'),
            )
            self.assertNotIn(
                'Цели версионируются с выбранного месяца', kpi_html,
            )
            self.assertIn('id="analyticsMonthDisplay"', kpi_html)
            self.assertIn('id="settingsMonthDisplay"', kpi_html)
            self.assertIn('id="exportKpiExcel"', kpi_html)
            self.assertLess(
                kpi_html.index('id="managerFilters"'),
                kpi_html.index('id="exportKpiExcel"'),
            )
        finally:
            response.close()

        response = self.client.get('/problems')
        try:
            taskboard_html = response.get_data(as_text=True)
            self.assertNotIn('class="back-link"', taskboard_html)
            self.assertIn('id="problemUserName"', taskboard_html)
            self.assertIn('id="problemUserBadge"', taskboard_html)
            self.assertLess(
                taskboard_html.index('id="newProblem"'),
                taskboard_html.index('id="repairCatalog"'),
            )
            self.assertIn('id="boardViewTabs"', taskboard_html)
            self.assertIn('id="analyticsView"', taskboard_html)
            self.assertIn('id="sendProblemReport"', taskboard_html)
            self.assertIn('id="downloadProblemExcel"', taskboard_html)
            self.assertLess(
                taskboard_html.index('id="analyticsSummary"'),
                taskboard_html.index('id="analyticsTypes"'),
            )
            self.assertLess(
                taskboard_html.index('id="analyticsTypes"'),
                taskboard_html.index('id="analyticsClubs"'),
            )
        finally:
            response.close()

        response = self.client.get('/shift')
        try:
            shift_html = response.get_data(as_text=True)
            self.assertNotIn('class="shift-back"', shift_html)
            self.assertIn('id="shiftUserName"', shift_html)
            self.assertIn('id="shiftRole"', shift_html)
        finally:
            response.close()

        response = self.client.get('/static/app.js')
        try:
            app_script = response.get_data(as_text=True)
            self.assertIn(
                'Команда OMG VR · ${state.me.name}', app_script,
            )
            self.assertIn('owner-settings-button', app_script)
            self.assertIn("state.me.role_name", app_script)
            self.assertIn('omg-kpi-manager-filters', app_script)
        finally:
            response.close()

    def test_employee_role_cannot_see_manager_controls(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
        ):
            response = self.client.get('/api/me', headers=self.headers)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.get_json()['can_manage'])

    def test_manager_sends_full_month_kpi_excel_to_bot_chat(self):
        rows = [
            {
                'login': '@second', 'nickname': 'Дарья В', 'shifts': 21,
                'total_pct': 0.24, 'rank': 2, 'birthdays': 3, 'zone': '🔴',
            },
            {
                'login': '@first', 'nickname': 'Мося', 'shifts': 26,
                'total_pct': 1.28, 'rank': 1, 'birthdays': 2, 'zone': '🟢',
            },
            {
                'login': '@zero', 'nickname': 'Без смен', 'shifts': 0,
                'total_pct': 0, 'rank': None, 'birthdays': 1, 'zone': '⚪',
            },
        ]
        sent_document = {}
        bot = Mock()

        def capture_document(chat_id, document, **kwargs):
            sent_document['chat_id'] = chat_id
            sent_document['filename'] = document.name
            sent_document['content'] = document.getvalue()
            sent_document['caption'] = kwargs.get('caption')

        bot.send_document.side_effect = capture_document
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(2)),
            patch.object(kpi_web, '_notification_bot', return_value=bot),
            patch.object(
                kpi_web, '_active_employee_logins',
                return_value=['@first', '@second', '@zero'],
            ),
            patch.object(
                kpi_web, '_employee_logins_with_month_shifts',
                return_value=['@first', '@second', '@zero'],
            ),
            patch.object(
                kpi_web, 'calculate_monthly_kpi', return_value=rows,
            ) as calculate,
        ):
            response = self.client.post(
                '/api/kpi/export?month=2026-08', headers=self.headers,
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get_json(),
            {'sent': True, 'filename': 'KPI_2026-08.xlsx'},
        )
        calculate.assert_called_once_with(
            '2026-08-01',
            employee_logins=['@first', '@second', '@zero'],
            period_end='2026-08-31',
        )
        self.assertEqual(sent_document['chat_id'], '1001')
        self.assertEqual(sent_document['filename'], 'KPI_2026-08.xlsx')
        self.assertEqual(sent_document['caption'], '📊 KPI за Август 2026')
        workbook = load_workbook(BytesIO(sent_document['content']))
        sheet = workbook['KPI Август 2026']
        self.assertIn('C1:E1', {str(item) for item in sheet.merged_cells.ranges})
        self.assertEqual(
            [sheet.cell(2, column).value for column in range(1, 6)],
            ['Ник', 'По 6 ч', '%', 'Рейтинг', 'ДРшки'],
        )
        self.assertEqual(
            [sheet.cell(3, column).value for column in range(1, 6)],
            ['Мося', 26, 1.28, 1, 1000],
        )
        self.assertEqual(sheet['A4'].value, 'Дарья В')
        self.assertEqual(sheet['E4'].value, 1500)
        self.assertEqual(sheet.max_row, 4)
        self.assertEqual(sheet['B3'].number_format, '0')
        self.assertEqual(sheet['C3'].number_format, '0%')
        self.assertEqual(sheet['A1'].font.name, 'Comfortaa')
        self.assertEqual(sheet['A3'].font.name, 'Comfortaa')
        self.assertEqual(sheet['A3'].fill.fgColor.rgb, '008EC37D')
        self.assertEqual(sheet['A4'].fill.fgColor.rgb, '00F3C4C6')
        workbook.close()

    def test_employee_cannot_export_kpi_excel(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, 'calculate_monthly_kpi') as calculate,
        ):
            response = self.client.post(
                '/api/kpi/export?month=2026-08', headers=self.headers,
            )

        self.assertEqual(response.status_code, 403)
        calculate.assert_not_called()

    def test_camera_prototype_is_linked_from_shift(self):
        shift_response = self.client.get('/shift')
        camera_response = self.client.get('/camera-test')
        shift_test_response = self.client.get('/shift-report?action=open')
        try:
            self.assertEqual(shift_response.status_code, 200)
            shift_html = shift_response.get_data(as_text=True)
            self.assertIn('id="openCameraTest"', shift_html)
            self.assertIn('id="shiftReportTest"', shift_html)
            self.assertIn('/shift-report?action=open', shift_html)
            self.assertIn('/shift-report?action=close', shift_html)
            self.assertEqual(camera_response.status_code, 200)
            camera_html = camera_response.get_data(as_text=True)
            self.assertIn('id="cameraStage"', camera_html)
            self.assertIn('id="sendCapture"', camera_html)
            self.assertIn('/static/camera_test.js', camera_html)
            self.assertEqual(shift_test_response.status_code, 200)
            shift_test_html = shift_test_response.get_data(as_text=True)
            self.assertIn('id="ownerClubStage"', shift_test_html)
            self.assertIn('id="earlyCloseDialog"', shift_test_html)
            self.assertIn('id="checklistStage"', shift_test_html)
            self.assertIn('id="questionStage"', shift_test_html)
            self.assertIn('id="cameraStage"', shift_test_html)
            self.assertIn('id="reviewStage"', shift_test_html)
            self.assertIn('/static/shift_test.js', shift_test_html)
            self.assertNotIn('Без рабочих изменений', shift_test_html)
            self.assertNotIn('Отправить Павлу', shift_test_html)
        finally:
            shift_response.close()
            camera_response.close()
            shift_test_response.close()

    def test_shift_test_selects_active_shift_then_nearest_shift(self):
        shifts = [
            {
                'date': '2026-08-10', 'club': 'Первый', 'duration': 4,
                'start': '08:00', 'end': '12:00',
            },
            {
                'date': '2026-08-10', 'club': 'Второй', 'duration': 4,
                'start': '12:30', 'end': '16:30',
            },
        ]
        with patch.object(kpi_web, '_upcoming_shifts', return_value=shifts):
            active = kpi_web._select_shift_report_test_shift(
                '@tester',
                now=datetime(2026, 8, 10, 13, 0, tzinfo=ZoneInfo('Europe/Moscow')),
            )
            nearest = kpi_web._select_shift_report_test_shift(
                '@tester',
                now=datetime(2026, 8, 10, 12, 10, tzinfo=ZoneInfo('Europe/Moscow')),
            )
            pinned = kpi_web._select_shift_report_test_shift(
                '@tester',
                now=datetime(2026, 8, 10, 13, 0, tzinfo=ZoneInfo('Europe/Moscow')),
                requested_club='Первый',
            )

        self.assertEqual(active['club'], 'Второй')
        self.assertEqual(nearest['club'], 'Второй')
        self.assertEqual(pinned['club'], 'Первый')

    def test_shift_test_scenario_uses_today_shift_and_editor_variant(self):
        today = kpi_web._moscow_today().isoformat()
        clubs = {
            'Тестовый клуб': {
                'shift_name': 'Клуб в расписании',
                'questions': {
                    '✅ Открыть смену': [[
                        {'text': 'Сколько наличных?', 'type': 'num', 'checklist': 'Включить оборудование'},
                        {'text': 'Фото клуба', 'type': 'photo'},
                        {'text': 'Фото оборудования', 'type': 'photo'},
                    ]],
                },
            },
        }
        shifts = [{
            'date': today, 'club': 'Клуб в расписании', 'duration': 12,
            'start': '10:00', 'end': '22:00',
        }]
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, 'get_clubs', return_value=clubs),
            patch.object(kpi_web, '_upcoming_shifts', return_value=shifts),
            patch.object(kpi_web.random, 'randrange', return_value=0),
        ):
            response = self.client.get(
                '/api/shift-test/scenario?action=open', headers=self.headers,
            )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload['production_mode'])
        self.assertEqual(payload['club'], 'Тестовый клуб')
        self.assertEqual(payload['variant_label'], 'A')
        self.assertEqual(payload['checklist'], ['Включить оборудование'])
        self.assertEqual(
            [question['type'] for question in payload['questions']],
            ['num', 'photo', 'photo'],
        )

    def test_shift_test_is_blocked_for_manager_when_today_shift_is_missing(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(2)),
            patch.object(kpi_web, '_upcoming_shifts', return_value=[]),
        ):
            response = self.client.get(
                '/api/shift-test/scenario?action=close', headers=self.headers,
            )

        self.assertEqual(response.status_code, 400)
        self.assertIn('На сегодня смена', response.get_json()['error'])

    def test_shift_test_owner_without_shift_can_select_configured_club(self):
        clubs = {
            'Тестовый клуб': {
                'shift_name': 'Клуб в расписании',
                'questions': {
                    '✅ Открыть смену': [[
                        {'text': 'Комментарий', 'type': 'text'},
                    ]],
                },
            },
            'Без открытия': {
                'shift_name': 'Другой клуб',
                'questions': {},
            },
        }
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(3)),
            patch.object(kpi_web, 'get_clubs', return_value=clubs),
            patch.object(kpi_web, '_upcoming_shifts', return_value=[]),
            patch.object(kpi_web.random, 'randrange', return_value=0),
        ):
            selection_response = self.client.get(
                '/api/shift-test/scenario?action=open', headers=self.headers,
            )
            scenario_response = self.client.get(
                '/api/shift-test/scenario?action=open&club=Тестовый+клуб',
                headers=self.headers,
            )

        self.assertEqual(selection_response.status_code, 200)
        selection = selection_response.get_json()
        self.assertTrue(selection['requires_club_selection'])
        self.assertEqual(selection['clubs'], ['Тестовый клуб'])
        self.assertEqual(scenario_response.status_code, 200)
        scenario = scenario_response.get_json()
        self.assertEqual(scenario['club'], 'Тестовый клуб')
        self.assertTrue(scenario['shift']['manual'])
        self.assertIsNone(scenario['shift']['start'])

    def test_shift_test_manager_cannot_select_club_without_shift(self):
        clubs = {
            'Тестовый клуб': {
                'shift_name': 'Тестовый клуб',
                'questions': {
                    '✅ Открыть смену': [[
                        {'text': 'Комментарий', 'type': 'text'},
                    ]],
                },
            },
        }
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(2)),
            patch.object(kpi_web, 'get_clubs', return_value=clubs),
            patch.object(kpi_web, '_upcoming_shifts', return_value=[]),
        ):
            response = self.client.get(
                '/api/shift-test/scenario?action=open&club=Тестовый+клуб',
                headers=self.headers,
            )

        self.assertEqual(response.status_code, 400)
        self.assertIn('На сегодня смена', response.get_json()['error'])

    def test_shift_report_requires_confirmation_before_early_close(self):
        today = kpi_web._moscow_today().isoformat()
        clubs = {
            'Тестовый клуб': {
                'shift_name': 'Тестовый клуб',
                'questions': {
                    '🚫 Закрыть смену': [[
                        {'text': 'Фото клуба', 'type': 'photo'},
                    ]],
                },
            },
        }
        shifts = [{
            'date': today, 'club': 'Тестовый клуб', 'duration': 12,
            'start': '10:00', 'end': '22:00',
        }]
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / 'early-close.sqlite3'
            conn = sqlite3.connect(db_path)
            conn.execute(
                'CREATE TABLE clubs (club TEXT PRIMARY KEY, status TEXT)'
            )
            conn.execute(
                "INSERT INTO clubs VALUES ('Тестовый клуб', 'Открыт')"
            )
            conn.commit()
            conn.close()
            bot = Mock()
            with (
                patch.object(kpi_web, 'DB_PATH', str(db_path)),
                patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
                patch.object(kpi_web, 'get_user', return_value=user(0)),
                patch.object(kpi_web, 'get_clubs', return_value=clubs),
                patch.object(kpi_web, '_upcoming_shifts', return_value=shifts),
                patch.object(kpi_web, '_shift_report_is_early_close', return_value=True),
                patch.object(kpi_web, '_notification_bot', return_value=bot),
                patch.object(kpi_web, 'refresh_club_status_dashboard', return_value=True),
                patch.dict(kpi_web.CHATS, {
                    'reports': '-100-reports',
                    'main_group': '-100-main',
                }),
            ):
                scenario = self.client.get(
                    '/api/shift-test/scenario?action=close&variant=0',
                    headers=self.headers,
                ).get_json()
                payload = {
                    'run_id': '2026-08-10:close:early-run',
                    'action': 'close',
                    'club': 'Тестовый клуб',
                    'variant_index': 0,
                    'version': scenario['version'],
                }
                blocked = self.client.post(
                    '/api/shift-test/start', headers=self.headers, json=payload,
                )
                conn = sqlite3.connect(db_path)
                status_before = conn.execute(
                    "SELECT status FROM clubs WHERE club='Тестовый клуб'"
                ).fetchone()[0]
                conn.close()
                confirmed = self.client.post(
                    '/api/shift-test/start',
                    headers=self.headers,
                    json={**payload, 'early_confirmed': True},
                )
                conn = sqlite3.connect(db_path)
                status_after = conn.execute(
                    "SELECT status FROM clubs WHERE club='Тестовый клуб'"
                ).fetchone()[0]
                conn.close()

        self.assertEqual(blocked.status_code, 409)
        self.assertEqual(
            blocked.get_json()['code'], 'early_close_confirmation_required'
        )
        self.assertEqual(status_before, 'Открыт')
        self.assertEqual(confirmed.status_code, 200)
        self.assertEqual(status_after, 'Закрыт')

    def test_shift_report_updates_working_records_and_sends_album(self):
        today = kpi_web._moscow_today().isoformat()
        clubs = {
            'Тестовый клуб': {
                'shift_name': 'Тестовый клуб',
                'questions': {
                    '✅ Открыть смену': [[
                        {
                            'text': 'Сколько наличных?', 'type': 'num',
                            'checklist': 'Проверить оборудование',
                        },
                        {'text': 'Фото клуба', 'type': 'photo'},
                        {'text': 'Фото оборудования', 'type': 'photo'},
                    ]],
                },
            },
        }
        shifts = [{
            'date': today, 'club': 'Тестовый клуб', 'duration': 12,
            'start': '10:00', 'end': '22:00',
        }]
        bot = Mock()
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / 'shift-report.sqlite3'
            conn = sqlite3.connect(db_path)
            conn.executescript(
                '''CREATE TABLE clubs (club TEXT PRIMARY KEY, status TEXT);
                   INSERT INTO clubs VALUES ('Тестовый клуб', 'Закрыт');
                   CREATE TABLE activity (
                       ID INTEGER PRIMARY KEY AUTOINCREMENT,
                       dtrep TEXT, login TEXT, club TEXT, action TEXT
                   );
                   CREATE TABLE nal (
                       ID INTEGER PRIMARY KEY AUTOINCREMENT,
                       drep TEXT, club TEXT, amount INTEGER
                   );'''
            )
            conn.commit()
            conn.close()
            with (
                patch.object(kpi_web, 'DB_PATH', str(db_path)),
                patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
                patch.object(kpi_web, 'get_user', return_value=user(0)),
                patch.object(kpi_web, 'get_clubs', return_value=clubs),
                patch.object(kpi_web, '_upcoming_shifts', return_value=shifts),
                patch.object(kpi_web, '_notification_bot', return_value=bot),
                patch.object(kpi_web, 'refresh_club_status_dashboard', return_value=True),
                patch.object(kpi_web, 'update_table_open') as update_sheet,
                patch.object(kpi_web, '_shift_report_test_sent_at', {}),
                patch.dict(kpi_web.CHATS, {
                    'reports': '-100-reports',
                    'main_group': '-100-main',
                }),
            ):
                scenario_response = self.client.get(
                    '/api/shift-test/scenario?action=open&variant=0',
                    headers=self.headers,
                )
                scenario = scenario_response.get_json()
                run_id = '2026-08-10:open:test-run'
                start_response = self.client.post(
                    '/api/shift-test/start',
                    headers=self.headers,
                    json={
                        'run_id': run_id,
                        'action': 'open',
                        'club': 'Тестовый клуб',
                        'variant_index': 0,
                        'version': scenario['version'],
                    },
                )
                incomplete_response = self.client.post(
                    '/api/shift-test/submit',
                    headers=self.headers,
                    data={
                        'report': json.dumps({
                            'run_id': run_id,
                            'action': 'open',
                            'club': 'Тестовый клуб',
                            'variant_index': 0,
                            'version': scenario['version'],
                            'answers': {'q1': '1000'},
                            'photo_ids': ['q2', 'q3'],
                        }),
                    },
                )
                response = self.client.post(
                    '/api/shift-test/submit',
                    headers=self.headers,
                    data={
                        'report': json.dumps({
                            'run_id': run_id,
                            'action': 'open',
                            'club': 'Тестовый клуб',
                            'variant_index': 0,
                            'version': scenario['version'],
                            'answers': {'q1': '1000'},
                            'photo_ids': ['q2', 'q3'],
                        }),
                        'photos': [
                            (BytesIO(b'jpeg-one'), 'q2.jpg', 'image/jpeg'),
                            (BytesIO(b'jpeg-two'), 'q3.jpg', 'image/jpeg'),
                        ],
                    },
                )
                conn = sqlite3.connect(db_path)
                status = conn.execute(
                    "SELECT status FROM clubs WHERE club='Тестовый клуб'"
                ).fetchone()[0]
                activity = conn.execute(
                    'SELECT login, club, action FROM activity'
                ).fetchone()
                cash = conn.execute(
                    'SELECT club, amount FROM nal'
                ).fetchone()
                conn.close()

        self.assertEqual(start_response.status_code, 200)
        self.assertEqual(incomplete_response.status_code, 400)
        self.assertIn('Количество фотографий', incomplete_response.get_json()['error'])
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()['photos'], 2)
        self.assertEqual(status, 'Открыт')
        self.assertEqual(activity, ('@tester', 'Тестовый клуб', '✅ Открыть смену'))
        self.assertEqual(cash, ('Тестовый клуб', 1000))
        update_sheet.assert_called_once()
        bot.send_media_group.assert_called_once()
        self.assertEqual(bot.send_media_group.call_args.args[0], '-100-reports')
        album = bot.send_media_group.call_args.kwargs['media']
        self.assertEqual(len(album), 2)
        self.assertIsNone(album[0].caption)
        self.assertIn('🌅 <b>Открытие смены</b>', album[1].caption)
        self.assertNotIn('Набор:', album[1].caption)
        self.assertNotIn('Фотографий:', album[1].caption)
        sent_chats = [call.args[0] for call in bot.send_message.call_args_list]
        self.assertIn('-100-reports', sent_chats)
        self.assertIn('-100-main', sent_chats)

        script = self.client.get('/static/shift_test.js')
        try:
            self.assertIn(b'indexedDB.open', script.data)
            self.assertIn(b'audio: false', script.data)
            self.assertIn(b"form.append('photos'", script.data)
        finally:
            script.close()

    def test_shift_test_sends_one_photo_without_invalid_one_item_album(self):
        bot = Mock()
        scenario = {
            'action_label': 'Закрытие',
            'club': 'Тестовый клуб',
            'variant_label': 'A',
            'shift': {
                'date': '2026-08-10', 'start': '10:00', 'end': '22:00',
            },
            'questions': [{
                'id': 'q1', 'position': 1, 'text': 'Фото клуба', 'type': 'photo',
            }],
        }
        photos = [{
            'content': b'jpeg',
            'filename': 'q1.jpg',
            'question': scenario['questions'][0],
        }]
        with (
            kpi_web.app.test_request_context(),
            patch.object(kpi_web, 'CAMERA_TEST_RECIPIENT_CHAT_ID', '592831529'),
            patch.object(kpi_web, '_notification_bot', return_value=bot),
        ):
            kpi_web.g.kpi_user = user(0)
            kpi_web._send_shift_report_test(scenario, {}, photos)

        bot.send_photo.assert_called_once()
        bot.send_message.assert_not_called()
        bot.send_media_group.assert_not_called()
        self.assertIn(
            '🌙 <b>Закрытие смены</b>',
            bot.send_photo.call_args.kwargs['caption'],
        )

    def test_shift_test_falls_back_when_unified_photo_send_fails(self):
        bot = Mock()
        bot.send_photo.side_effect = [RuntimeError('caption rejected'), None]
        scenario = {
            'action_label': 'Открытие',
            'club': 'Тестовый клуб',
            'variant_label': 'A',
            'shift': {
                'date': '2026-08-10', 'start': '10:00', 'end': '22:00',
            },
            'questions': [{
                'id': 'q1', 'position': 1, 'text': 'Фото клуба', 'type': 'photo',
            }],
        }
        photos = [{
            'content': b'jpeg',
            'filename': 'q1.jpg',
            'question': scenario['questions'][0],
        }]
        with (
            kpi_web.app.test_request_context(),
            patch.object(kpi_web, 'CAMERA_TEST_RECIPIENT_CHAT_ID', '592831529'),
            patch.object(kpi_web, '_notification_bot', return_value=bot),
        ):
            kpi_web.g.kpi_user = user(0)
            kpi_web._send_shift_report_test(scenario, {}, photos)

        self.assertEqual(bot.send_photo.call_count, 2)
        bot.send_message.assert_called_once()
        self.assertIn(
            '🌅 <b>Открытие смены</b>',
            bot.send_message.call_args.args[1],
        )

    def test_camera_photo_is_sent_to_configured_private_chat(self):
        bot = Mock()
        diagnostics = {
            'platform': 'ios',
            'telegram_version': '9.0',
            'capture_method': 'Короткое нажатие в Mini App',
            'mime_type': 'image/jpeg',
            'camera_status': 'Работает',
        }
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, 'CAMERA_TEST_RECIPIENT_CHAT_ID', '592831529'),
            patch.object(kpi_web, '_notification_bot', return_value=bot),
            patch.object(kpi_web, '_camera_test_sent_at', {}),
        ):
            response = self.client.post(
                '/api/camera-test',
                headers=self.headers,
                data={
                    'consent': 'yes',
                    'diagnostics': json.dumps(diagnostics),
                    'media': (BytesIO(b'jpeg-bytes'), 'camera-test.jpg', 'image/jpeg'),
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()['delivery'], 'photo')
        bot.send_photo.assert_called_once()
        self.assertEqual(bot.send_photo.call_args.args[0], '592831529')
        self.assertIn('Тест камеры Mini App', bot.send_photo.call_args.kwargs['caption'])

    def test_camera_webm_is_sent_as_document(self):
        bot = Mock()
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, 'CAMERA_TEST_RECIPIENT_CHAT_ID', '592831529'),
            patch.object(kpi_web, '_notification_bot', return_value=bot),
            patch.object(kpi_web, '_camera_test_sent_at', {}),
        ):
            response = self.client.post(
                '/api/camera-test',
                headers=self.headers,
                data={
                    'consent': 'yes',
                    'diagnostics': '{}',
                    'media': (BytesIO(b'webm-bytes'), 'camera-test.webm', 'video/webm'),
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()['delivery'], 'document')
        bot.send_document.assert_called_once()

    def test_home_places_compact_shift_module_before_dashboard(self):
        response = self.client.get('/')
        try:
            html = response.get_data(as_text=True)
            self.assertLess(
                html.index('class="modules-section"'),
                html.index('id="personalSection"'),
            )
            self.assertIn('class="module-card shift-module" href="/shift"', html)
            self.assertIn('<span class="module-icon">↗</span>', html)
            self.assertIn('<span class="module-icon">?</span>', html)
            self.assertIn('<span class="module-icon">✓</span>', html)
            self.assertNotIn('id="shiftConfigModule"', html)
            self.assertIn('Сегодня в клубах', html)
        finally:
            response.close()

        script = self.client.get('/static/home.js')
        try:
            self.assertNotIn(b'club.red_zone', script.data)
            self.assertNotIn('Ближайшая'.encode(), script.data)
            self.assertIn('Есть бронь'.encode(), script.data)
            self.assertIn('Нет брони'.encode(), script.data)
            self.assertIn(b'class="club-shift-summary"', script.data)
            self.assertIn(b'class="club-live-statuses"', script.data)
            self.assertIn(b'class="shift-person"', script.data)
            self.assertIn(b'class="club-problems-link"', script.data)
            self.assertIn(b'/problems?club=', script.data)
            self.assertIn(b'data-telegram-username', script.data)
            self.assertIn(b'tg.openTelegramLink(url)', script.data)
            self.assertLess(
                script.data.index(b'class="club-shift-summary"'),
                script.data.index(b'class="club-meta"'),
            )
            self.assertIn(
                b'class="summary-tile clickable-card" href="/kpi"',
                script.data,
            )
            self.assertIn(
                b'class="summary-tile clickable-card" href="/problems"',
                script.data,
            )
            self.assertIn(
                b'class="shift-card clickable-card shift-link" href="/shift"',
                script.data,
            )
        finally:
            script.close()

        problem_script = self.client.get('/static/problems.js')
        try:
            self.assertIn(b'class="media-badge photo"', problem_script.data)
            self.assertIn(b'class="media-badge video"', problem_script.data)
            self.assertIn(b"state.meta.clubs.includes(club)", problem_script.data)
            self.assertIn(b"return 'repair'", problem_script.data)
            self.assertIn(b"return 'bot'", problem_script.data)
            self.assertIn(b"return 'general'", problem_script.data)
            self.assertNotIn('▧'.encode(), problem_script.data)
            self.assertNotIn('▶'.encode(), problem_script.data)
        finally:
            problem_script.close()

    def test_shift_module_is_visible_to_employee_and_unlocks_manager_tools(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'OMG_SHIFT_URL', 'http://shift.test/'),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, '_upcoming_shifts', return_value=[{
                'date': '2026-08-10', 'club': 'Дмитровка', 'duration': 12.5,
                'start': '10:00', 'end': '22:30',
            }]),
            patch.object(kpi_web, '_shift_month_summary', return_value={
                'shifts': 8, 'hours': 86.0,
            }),
        ):
            employee_response = self.client.get(
                '/api/shift', headers=self.headers,
            )

        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'OMG_SHIFT_URL', 'http://shift.test/'),
            patch.object(kpi_web, 'get_user', return_value=user(2)),
            patch.object(kpi_web, '_upcoming_shifts', return_value=[]),
            patch.object(kpi_web, '_shift_month_summary', return_value={
                'shifts': 0, 'hours': 0.0,
            }),
        ):
            manager_response = self.client.get(
                '/api/shift', headers=self.headers,
            )

        self.assertEqual(employee_response.status_code, 200)
        employee_payload = employee_response.get_json()
        self.assertEqual(employee_payload['external_url'], 'http://shift.test/')
        self.assertEqual(employee_payload['user_name'], 'Тестер')
        self.assertEqual(employee_payload['role_name'], 'Сотрудник')
        self.assertFalse(employee_payload['can_manage'])
        self.assertEqual(employee_payload['employee_dashboard']['month_summary'], {
            'shifts': 8, 'hours': 86.0,
        })
        self.assertEqual(
            employee_payload['employee_dashboard']['upcoming_shifts'][0]['club'],
            'Дмитровка',
        )
        self.assertEqual(
            employee_payload['employee_dashboard']['upcoming_shifts'][0]['start'],
            '10:00',
        )
        self.assertEqual(manager_response.status_code, 200)
        manager_payload = manager_response.get_json()
        self.assertTrue(manager_payload['can_manage'])
        self.assertEqual(manager_payload['role_name'], 'Менеджер')

        shift_css = self.client.get('/static/shift.css')
        try:
            self.assertIn(b'.shift-actions[hidden]', shift_css.data)
            self.assertIn(b'.shift-action[hidden]', shift_css.data)
            self.assertIn(b'display:none !important', shift_css.data)
        finally:
            shift_css.close()

    def test_owner_is_excluded_from_active_kpi_employees(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / 'roles.db'
            conn = sqlite3.connect(db_path)
            conn.executescript(
                '''
                CREATE TABLE users (
                    ID INTEGER PRIMARY KEY,
                    login TEXT,
                    status INTEGER
                );
                INSERT INTO users VALUES (1, '@employee', 0);
                INSERT INTO users VALUES (2, '@manager', 2);
                INSERT INTO users VALUES (3, '@owner', 3);
                '''
            )
            conn.commit()
            conn.close()

            with patch.object(kpi_web, 'DB_PATH', str(db_path)):
                logins = kpi_web._active_employee_logins()

        self.assertEqual(logins, ['@employee', '@manager'])

    def test_owner_home_contains_club_dashboard_without_personal_kpi(self):
        team_rows = [{'login': '@employee', 'shifts': 2, 'zone': '🟢'}]
        management = {
            'participants': 1,
            'average_pct': 0.8,
            'red_zone': 0,
            'active_penalties': 0,
        }
        clubs = [{
            'club': 'Марьино',
            'status': 'Открыт',
            'on_shift': ['Сотрудник'],
            'problems': {'work': 1, 'review': 0},
        }]
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(3)),
            patch.object(
                kpi_web,
                '_team_snapshot',
                return_value=(team_rows, management),
            ),
            patch.object(
                kpi_web,
                '_task_counts',
                return_value={'work': 1, 'review': 0},
            ),
            patch.object(kpi_web, '_problem_counts_by_club', return_value={}),
            patch.object(kpi_web, '_club_dashboard', return_value=clubs),
            patch.object(kpi_web, '_upcoming_shifts') as upcoming,
        ):
            response = self.client.get('/api/home', headers=self.headers)

        payload = response.get_json()
        self.assertEqual(response.status_code, 200)
        self.assertIsNone(payload['personal_kpi'])
        self.assertEqual(payload['upcoming_shifts'], [])
        self.assertEqual(payload['clubs'], clubs)
        upcoming.assert_not_called()

    def test_club_dashboard_includes_shift_telegram_contacts(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / 'home.db'
            conn = sqlite3.connect(db_path)
            conn.executescript(
                '''
                CREATE TABLE clubs (club TEXT, status TEXT);
                CREATE TABLE users (
                    ID INTEGER PRIMARY KEY,
                    login TEXT,
                    first_name TEXT,
                    second_name TEXT,
                    nick_name TEXT
                );
                CREATE TABLE shifts (
                    club TEXT,
                    dt_shift TEXT,
                    shift_login TEXT,
                    shift_first_name TEXT,
                    shift_second_name TEXT
                );
                INSERT INTO clubs VALUES ('Марьино', 'Открыт');
                INSERT INTO users VALUES (
                    1, '@employee', 'Иван', 'Иванов', 'Ваня'
                );
                INSERT INTO shifts VALUES (
                    'Марьино', '2026-08-10', '@employee', 'Иван', 'Иванов'
                );
                '''
            )
            conn.commit()
            conn.close()

            with (
                patch.object(kpi_web, 'DB_PATH', str(db_path)),
                patch.object(kpi_web, '_moscow_today', return_value=date(2026, 8, 10)),
                patch.object(
                    kpi_web, 'get_clubs',
                    return_value={'Марьино': {'is_physical': True}},
                ),
            ):
                clubs = kpi_web._club_dashboard([], {})

        self.assertEqual(clubs[0]['on_shift'], ['Ваня'])
        self.assertEqual(
            clubs[0]['on_shift_contacts'],
            [{'name': 'Ваня', 'login': '@employee'}],
        )

    def test_manager_home_also_contains_operational_club_dashboard(self):
        clubs = [{
            'club': 'Ленинский',
            'status': 'Открыт',
            'on_shift': ['Менеджер'],
            'problems': {'work': 0, 'review': 1},
        }]
        management = {
            'participants': 1,
            'average_pct': 0.9,
            'red_zone': 0,
            'active_penalties': 0,
        }
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(2)),
            patch.object(kpi_web, '_upcoming_shifts', return_value=[]),
            patch.object(kpi_web, 'calculate_monthly_kpi', return_value=[]),
            patch.object(
                kpi_web, '_team_snapshot', return_value=([], management),
            ),
            patch.object(
                kpi_web, '_task_counts', return_value={'work': 0, 'review': 1},
            ),
            patch.object(kpi_web, '_problem_counts_by_club', return_value={}),
            patch.object(kpi_web, '_club_dashboard', return_value=clubs),
        ):
            response = self.client.get('/api/home', headers=self.headers)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()['clubs'], clubs)

    def test_club_employee_sees_only_today_shift_club_bookings(self):
        groups = [{
            'club': 'Марьино',
            'count': 1,
            'participants': 5,
            'bookings': [{
                'start': '2026-08-08T12:00:00',
                'end': '2026-08-08T13:00:00',
                'format': 'Классический VR',
                'participants': 5,
            }],
        }]
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, '_today_shift_clubs', return_value=['Марьино']),
            patch.object(kpi_web, '_club_booking_groups', return_value=groups) as grouped,
            patch.object(kpi_web, 'booking_freshness', return_value={
                'last_synced_at': '2026-08-08T12:00:00+03:00',
                'age_minutes': 1,
                'stale': False,
            }),
        ):
            response = self.client.get('/api/bookings/today', headers=self.headers)

        payload = response.get_json()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload['mode'], 'clubs')
        self.assertEqual(payload['groups'], groups)
        self.assertNotIn('number', payload['groups'][0]['bookings'][0])
        grouped.assert_called_once_with(['Марьино'], unittest.mock.ANY)

    def test_callcenter_employee_sees_upcoming_unpaid_orders(self):
        order = {
            'reservation_at': '2026-08-09T12:00:00',
            'reservation_end_at': '2026-08-09T13:00:00',
            'date': date(2026, 8, 9),
            'club': 'Каширка',
            'booking_format': 'Мероприятие',
            'participants': 12,
            'number': '12345',
            'url': 'https://my.bukza.com/order/12345',
        }
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(
                kpi_web,
                '_today_shift_clubs',
                return_value=['Коллцентр'],
            ),
            patch.object(kpi_web, 'upcoming_unpaid_orders', return_value=[order]),
            patch.object(kpi_web, 'booking_freshness', return_value={
                'last_synced_at': '2026-08-08T12:00:00+03:00',
                'age_minutes': 1,
                'stale': False,
            }),
        ):
            response = self.client.get('/api/bookings/today', headers=self.headers)

        payload = response.get_json()
        self.assertEqual(payload['mode'], 'callcenter')
        self.assertEqual(payload['bookings'][0]['number'], '12345')
        self.assertEqual(payload['bookings'][0]['club'], 'Каширка')

    def test_manager_sees_brief_booking_groups_for_all_clubs(self):
        groups = [{'club': 'Ленинский', 'count': 2, 'participants': 7, 'bookings': []}]
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(2)),
            patch.object(kpi_web, '_club_booking_groups', return_value=groups) as grouped,
            patch.object(kpi_web, '_today_shift_clubs') as shifts,
            patch.object(kpi_web, 'booking_freshness', return_value={
                'last_synced_at': '2026-08-08T12:00:00+03:00',
                'age_minutes': 1,
                'stale': False,
            }),
        ):
            response = self.client.get('/api/bookings/today', headers=self.headers)

        payload = response.get_json()
        self.assertEqual(payload['mode'], 'management')
        self.assertEqual(payload['groups'], groups)
        grouped.assert_called_once_with(
            list(kpi_web.BUKZA_CLUB_CODES.values()),
            unittest.mock.ANY,
        )
        shifts.assert_not_called()

    def test_manager_and_owner_can_open_shift_config(self):
        config = {'version': 'abc', 'clubs': []}
        for role in (2, 3):
            with (
                patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
                patch.object(kpi_web, 'get_user', return_value=user(role)),
                patch.object(kpi_web, 'get_editor_config', return_value=config),
            ):
                response = self.client.get('/api/shift-config', headers=self.headers)

            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.get_json(), config)

    def test_employee_cannot_open_shift_config(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, 'get_editor_config') as editor,
        ):
            response = self.client.get('/api/shift-config', headers=self.headers)

        self.assertEqual(response.status_code, 403)
        editor.assert_not_called()

    def test_problem_form_uses_general_request_type(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, 'get_clubs', return_value={'Марьино': {}}),
        ):
            response = self.client.get('/api/problems-meta', headers=self.headers)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()['types'], [
            'Общее обращение',
            'Ремонт',
            'Улучшение бота',
        ])
        self.assertFalse(response.get_json()['can_view_analytics'])
        self.assertFalse(response.get_json()['can_export_analytics'])

    def test_problem_analytics_are_available_from_technician_role(self):
        payload = {
            'period': {'mode': 'month', 'value': '2026-08', 'label': '08.2026'},
            'summary': {'created': 0},
        }
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(1)),
            patch.object(kpi_web, 'build_task_analytics', return_value=payload) as build,
        ):
            response = self.client.get(
                '/api/problems/analytics?mode=month&month=2026-08',
                headers=self.headers,
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), payload)
        build.assert_called_once_with(
            kpi_web.DB_PATH,
            mode='month',
            month='2026-08',
            year=None,
        )

    def test_problem_analytics_are_hidden_from_employee_role(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, 'build_task_analytics') as build,
        ):
            response = self.client.get(
                '/api/problems/analytics', headers=self.headers,
            )

        self.assertEqual(response.status_code, 403)
        build.assert_not_called()

    def test_problem_text_report_is_manager_only_and_uses_selected_period(self):
        report = {'summary': {'open': 2}}
        bot = Mock()
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(2)),
            patch.object(kpi_web, 'build_task_report', return_value=report) as build,
            patch.object(
                kpi_web, 'format_task_report_html',
                return_value=['<b>Готовый отчёт</b>', 'Продолжение'],
            ),
            patch.object(kpi_web, '_notification_bot', return_value=bot),
        ):
            response = self.client.post(
                '/api/problems/export/text?mode=month&month=2026-08',
                headers=self.headers,
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {'sent': True, 'messages': 2})
        self.assertEqual(bot.send_message.call_count, 2)
        self.assertEqual(bot.send_message.call_args_list[0].args, (
            '1001', '<b>Готовый отчёт</b>',
        ))
        self.assertEqual(
            bot.send_message.call_args_list[0].kwargs['parse_mode'], 'HTML',
        )
        build.assert_called_once_with(
            kpi_web.DB_PATH, mode='month', month='2026-08', year=None,
        )

        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(1)),
        ):
            forbidden = self.client.post(
                '/api/problems/export/text', headers=self.headers,
            )
        self.assertEqual(forbidden.status_code, 403)

    def test_problem_excel_report_contains_summary_and_task_sheets(self):
        report = {
            'period': {'mode': 'month', 'value': '2026-08', 'label': 'август 2026'},
            'generated_at': '2026-08-24T12:30:00+03:00',
            'summary': {
                'created': 1, 'completed': 1, 'work': 1, 'review': 0, 'open': 1,
                'average_first_response_seconds': 1800,
                'first_response_precision': 'exact',
                'average_resolution_seconds': 7200,
                'resolution_precision': 'exact',
            },
            'open_clubs': [{'label': 'Каширка', 'open': 1}],
            'closed_clubs': [{
                'label': 'Каширка', 'count': 1,
                'average_first_response_seconds': 1800,
                'first_response_precision': 'exact',
                'average_resolution_seconds': 7200,
                'resolution_precision': 'exact',
            }],
            'backlog': [],
            'rows': [{
                'id': 15, 'date': '2026-08-20', 'closed_at': '2026-08-21',
                'club': 'Каширка', 'type': 'Ремонт', 'title': 'Шлем 2 зона',
                'description': 'Не включается', 'status': 'Выполнено',
                'age_days': 1, 'is_backlog': False,
                'created_in_period': True, 'completed_in_period': True,
                'feedback': '', 'final_solution': 'Заменили кабель',
                'first_response_seconds': 1800,
                'first_response_precision': 'exact',
                'resolution_seconds': 7200,
                'resolution_precision': 'exact',
                'return_count': 0,
            }],
        }
        sent_document = {}
        bot = Mock()

        def capture_document(chat_id, document, **kwargs):
            sent_document['chat_id'] = chat_id
            sent_document['filename'] = document.name
            sent_document['content'] = document.getvalue()
            sent_document['caption'] = kwargs.get('caption')

        bot.send_document.side_effect = capture_document
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(3)),
            patch.object(kpi_web, '_problem_report_from_request', return_value=report),
            patch.object(kpi_web, '_notification_bot', return_value=bot),
        ):
            response = self.client.post(
                '/api/problems/export/excel?mode=month&month=2026-08',
                headers=self.headers,
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {
            'sent': True, 'filename': 'Taskboard_2026-08.xlsx',
        })
        self.assertEqual(sent_document['chat_id'], '1001')
        self.assertEqual(sent_document['filename'], 'Taskboard_2026-08.xlsx')
        self.assertEqual(sent_document['caption'], '🚩 Доска проблем · август 2026')
        workbook = load_workbook(BytesIO(sent_document['content']))
        self.assertEqual(workbook.sheetnames, ['Сводка', 'Закрытые'])
        self.assertEqual(workbook['Закрытые']['A2'].value, 15)
        self.assertEqual(workbook['Закрытые']['F2'].value, 'Шлем 2 зона')
        self.assertEqual(workbook['Закрытые']['K2'].value, 'Заменили кабель')
        self.assertEqual(workbook['Закрытые']['A2'].font.name, 'Comfortaa')
        workbook.close()

    def test_equipment_is_available_only_from_technician_role(self):
        for role in (1, 2, 3):
            with (
                patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
                patch.object(kpi_web, 'get_user', return_value=user(role)),
                patch.object(
                    kpi_web, 'equipment_list_payload',
                    return_value={'units': []},
                ) as equipment,
            ):
                response = self.client.get('/api/equipment', headers=self.headers)

            self.assertEqual(response.status_code, 200)
            equipment.assert_called_once_with(kpi_web.DB_PATH)

        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, 'equipment_list_payload') as equipment,
        ):
            response = self.client.get('/api/equipment', headers=self.headers)

        self.assertEqual(response.status_code, 403)
        equipment.assert_not_called()

    def test_mini_app_creates_anonymous_problem_in_shared_tasks_table(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / 'tasks.db'
            conn = sqlite3.connect(db_path)
            conn.execute(
                '''CREATE TABLE tasks (
                       ID INTEGER PRIMARY KEY AUTOINCREMENT,
                       dtrep TEXT, type TEXT, club TEXT, title TEXT,
                       photo BLOB, desc TEXT, status TEXT, dtfb TEXT,
                       feedback TEXT
                   )'''
            )
            conn.commit()
            conn.close()

            kpi_web.initialize_repair_schema(str(db_path))
            conn = sqlite3.connect(db_path)
            item_id = conn.execute(
                "SELECT id FROM repair_item_types WHERE name='VR-шлем'"
            ).fetchone()[0]
            location_id = conn.execute(
                "SELECT id FROM repair_locations WHERE club='Марьино' AND name='2 зона'"
            ).fetchone()[0]
            conn.close()

            with (
                patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
                patch.object(kpi_web, 'get_user', return_value=user(0)),
                patch.object(kpi_web, 'DB_PATH', str(db_path)),
                patch.object(kpi_web, 'get_clubs', return_value={'Марьино': {}}),
                patch.object(kpi_web, '_send_problem_notification') as notify,
            ):
                response = self.client.post(
                    '/api/problems',
                    headers=self.headers,
                    data={
                        'type': 'Ремонт',
                        'club': 'Марьино',
                        'title': 'Не работает шлем',
                        'description': 'Не включается второй шлем',
                        'repair_item_id': str(item_id),
                        'repair_location_ids': json.dumps([location_id]),
                    },
                )
                detail_response = self.client.get(
                    '/api/problems/1', headers=self.headers,
                )

            conn = sqlite3.connect(db_path)
            row = conn.execute(
                'SELECT type, club, title, desc, status FROM tasks'
            ).fetchone()
            columns = [item[1] for item in conn.execute('PRAGMA table_info(tasks)')]
            repair_link = conn.execute(
                '''SELECT cases.task_id, items.name, locations.name
                   FROM repair_cases cases
                   JOIN repair_item_types items ON items.id=cases.item_type_id
                   JOIN repair_case_locations links ON links.task_id=cases.task_id
                   JOIN repair_locations locations ON locations.id=links.location_id'''
            ).fetchone()
            repair_event = conn.execute(
                'SELECT event_type FROM repair_events WHERE task_id=1'
            ).fetchone()
            task_event = conn.execute(
                'SELECT event_type FROM task_events WHERE task_id=1'
            ).fetchone()
            conn.close()

        self.assertEqual(response.status_code, 201)
        self.assertEqual(
            row,
            (
                'Ремонт', 'Марьино', 'VR-шлем — 2 зона',
                'Не включается второй шлем', 'В работе',
            ),
        )
        self.assertNotIn('author', columns)
        self.assertEqual(repair_link, (1, 'VR-шлем', '2 зона'))
        self.assertEqual(repair_event, ('created',))
        self.assertEqual(task_event, ('created',))
        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual(
            detail_response.get_json()['repair']['history'][0]['task_id'], 1,
        )
        self.assertEqual(
            detail_response.get_json()['activity'][0]['actor'],
            {'name': 'Тестер', 'login': '@tester'},
        )
        notify.assert_called_once()

    def test_equipment_replacement_confirms_multi_location_repairs(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / 'tasks.db'
            conn = sqlite3.connect(db_path)
            conn.execute(
                '''CREATE TABLE tasks (
                       ID INTEGER PRIMARY KEY AUTOINCREMENT,
                       dtrep TEXT, type TEXT, club TEXT, title TEXT,
                       photo BLOB, desc TEXT, status TEXT, dtfb TEXT,
                       feedback TEXT
                   )'''
            )
            conn.commit()
            conn.close()
            kpi_web.initialize_repair_schema(str(db_path))

            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            item_id = conn.execute(
                "SELECT id FROM repair_item_types WHERE name='VR-шлем'"
            ).fetchone()[0]
            location_ids = [
                row[0] for row in conn.execute(
                    '''SELECT id FROM repair_locations
                       WHERE club='Марьино' AND name IN ('1 зона', '2 зона')
                       ORDER BY name'''
                )
            ]
            with conn:
                cursor = conn.execute(
                    '''INSERT INTO tasks(
                           dtrep, type, club, title, desc, status
                       ) VALUES (
                           '2026-08-17', 'Ремонт', 'Марьино',
                           'VR-шлем — 1–2 зоны', 'Не включается', 'В работе'
                       )'''
                )
                kpi_web.create_repair_case(
                    conn, cursor.lastrowid, 'Марьино', item_id, None, location_ids,
                )
            unit_id = conn.execute(
                '''SELECT units.id FROM equipment_units units
                   WHERE units.location_id=? AND units.item_type_id=?''',
                (location_ids[0], item_id),
            ).fetchone()[0]
            conn.close()

            with (
                patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
                patch.object(kpi_web, 'get_user', return_value=user(1)),
                patch.object(kpi_web, 'DB_PATH', str(db_path)),
                patch.object(kpi_web, '_send_problem_notification') as notify,
            ):
                equipment_list = self.client.get(
                    '/api/equipment', headers=self.headers,
                )
                equipment_detail = self.client.get(
                    f'/api/equipment/{unit_id}', headers=self.headers,
                )
                warning = self.client.post(
                    f'/api/equipment/{unit_id}/replace',
                    headers=self.headers,
                    json={'message': 'Выдан новый шлем'},
                )
                replaced = self.client.post(
                    f'/api/equipment/{unit_id}/replace',
                    headers=self.headers,
                    json={
                        'message': 'Выдан новый шлем',
                        'close_multi_location_tasks': True,
                    },
                )

            conn = sqlite3.connect(db_path)
            task = conn.execute(
                'SELECT status, feedback FROM tasks WHERE ID=1'
            ).fetchone()
            generations = conn.execute(
                '''SELECT generation, active FROM equipment_units
                   WHERE location_id=? AND item_type_id=? ORDER BY generation''',
                (location_ids[0], item_id),
            ).fetchall()
            conn.close()

        self.assertEqual(equipment_list.status_code, 200)
        self.assertEqual(len(equipment_list.get_json()['units']), 2)
        self.assertEqual(equipment_detail.status_code, 200)
        self.assertEqual(len(equipment_detail.get_json()['open_tasks']), 1)
        self.assertEqual(warning.status_code, 409)
        self.assertTrue(warning.get_json()['requires_confirmation'])
        self.assertEqual(replaced.status_code, 200)
        self.assertEqual(replaced.get_json()['closed_task_count'], 1)
        self.assertEqual(task[0], 'Выполнено')
        self.assertIn('Выдан новый шлем', task[1])
        self.assertEqual(generations, [(1, 0), (2, 1)])
        notify.assert_called_once()

    def test_problem_video_is_stored_as_telegram_reference(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / 'tasks.db'
            conn = sqlite3.connect(db_path)
            conn.execute(
                '''CREATE TABLE tasks (
                       ID INTEGER PRIMARY KEY AUTOINCREMENT,
                       dtrep TEXT, type TEXT, club TEXT, title TEXT,
                       photo BLOB, desc TEXT, status TEXT, dtfb TEXT,
                       feedback TEXT
                   )'''
            )
            conn.commit()
            conn.close()
            video_reference = {
                'file_id': 'telegram-video-id',
                'file_unique_id': 'telegram-unique-id',
                'mimetype': 'video/mp4',
                'file_size': 11,
            }

            with (
                patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
                patch.object(kpi_web, 'get_user', return_value=user(0)),
                patch.object(kpi_web, 'DB_PATH', str(db_path)),
                patch.object(kpi_web, 'get_clubs', return_value={'Марьино': {}}),
                patch.object(
                    kpi_web,
                    '_send_problem_notification',
                    return_value=video_reference,
                ) as notify,
            ):
                response = self.client.post(
                    '/api/problems',
                    headers=self.headers,
                    data={
                        'type': 'Общее обращение',
                        'club': 'Марьино',
                        'title': 'Проблема со звуком',
                        'description': 'На видео слышен посторонний шум',
                        'video': (BytesIO(b'video-bytes'), 'problem.mp4', 'video/mp4'),
                    },
                )
                detail = self.client.get('/api/problems/1', headers=self.headers)

            conn = sqlite3.connect(db_path)
            stored = conn.execute(
                '''SELECT telegram_file_id, telegram_file_unique_id,
                          mime_type, file_size
                   FROM task_videos WHERE task_id=1'''
            ).fetchone()
            photo = conn.execute('SELECT photo FROM tasks WHERE ID=1').fetchone()[0]
            conn.close()

        self.assertEqual(response.status_code, 201)
        self.assertEqual(
            stored,
            ('telegram-video-id', 'telegram-unique-id', 'video/mp4', 11),
        )
        self.assertIsNone(photo)
        self.assertTrue(detail.get_json()['has_video'])
        self.assertEqual(notify.call_args.kwargs['video']['content'], b'video-bytes')

    def test_mp4_problem_notification_returns_telegram_reference(self):
        bot = Mock()
        bot.send_video.return_value.video = Mock(
            file_id='telegram-video-id',
            file_unique_id='telegram-unique-id',
        )
        task = {
            'type': 'Общее обращение',
            'club': 'Марьино',
            'title': 'Проблема со звуком',
            'description': 'На записи слышен шум',
        }

        with (
            patch.object(kpi_web, '_notification_bot', return_value=bot),
            patch.object(kpi_web, 'get_clubs', return_value={'Марьино': {}}),
            patch.dict(kpi_web.CHATS, {'reports': -100500}, clear=True),
        ):
            reference = kpi_web._send_problem_notification(
                'created',
                task,
                video={
                    'content': b'video',
                    'filename': 'problem.mp4',
                    'mimetype': 'video/mp4',
                },
            )

        self.assertEqual(reference, {
            'file_id': 'telegram-video-id',
            'file_unique_id': 'telegram-unique-id',
            'mimetype': 'video/mp4',
            'file_size': 5,
        })
        self.assertTrue(bot.send_video.call_args.kwargs['supports_streaming'])
        self.assertEqual(bot.send_video.call_args.args[1].name, 'problem.mp4')

    def test_repair_video_is_copied_to_repair_chat_but_not_main_chat(self):
        bot = Mock()
        bot.send_video.return_value.video = Mock(
            file_id='telegram-video-id',
            file_unique_id='telegram-unique-id',
        )
        task = {
            'type': 'Ремонт',
            'club': 'Марьино',
            'title': 'Шлем — 2 зона',
            'description': 'Не включается',
        }
        actor = {'chatid': '1001', 'login': '@tester', 'name': 'Тестер'}

        with (
            patch.object(kpi_web, '_notification_bot', return_value=bot),
            patch.object(
                kpi_web, 'get_clubs',
                return_value={'Марьино': {'tag': '@maryino'}},
            ),
            patch.dict(
                kpi_web.CHATS,
                {'reports': -1, 'main_group': -2, 'repair_extra': -3},
                clear=True,
            ),
        ):
            kpi_web._send_problem_notification(
                'created',
                task,
                video={
                    'content': b'video',
                    'filename': 'problem.mp4',
                    'mimetype': 'video/mp4',
                },
                actor=actor,
            )

        self.assertEqual(bot.send_video.call_count, 2)
        report_caption = bot.send_video.call_args_list[0].kwargs['caption']
        repair_caption = bot.send_video.call_args_list[1].kwargs['caption']
        main_text = bot.send_message.call_args.args[1]
        self.assertIn('#задачи', report_caption)
        self.assertIn('Не включается', report_caption)
        self.assertNotIn('#задачи', repair_caption)
        self.assertNotIn('@OMGVR_Admin_Bot', repair_caption)
        self.assertIn('Не включается', repair_caption)
        self.assertNotIn('Не включается', main_text)
        self.assertIn('Создал:</b> Тестер (@tester)', main_text)

    def test_problem_confirmation_sends_completed_notification(self):
        task = {
            'ID': 1,
            'type': 'Ремонт',
            'club': 'Марьино',
            'title': 'Не работает шлем',
        }
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, '_change_problem_status', return_value=task),
            patch.object(kpi_web, '_send_problem_notification') as notify,
        ):
            response = self.client.post(
                '/api/problems/1/confirm', headers=self.headers,
            )

        self.assertEqual(response.status_code, 200)
        notify.assert_called_once_with(
            'completed', task,
            actor={'chatid': '1001', 'login': '@tester', 'name': 'Тестер'},
        )

    def test_problem_video_is_streamed_from_telegram(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / 'tasks.db'
            kpi_web._initialize_problem_video_schema(str(db_path))
            conn = sqlite3.connect(db_path)
            conn.execute(
                '''INSERT INTO task_videos (
                       task_id, telegram_file_id, telegram_file_unique_id,
                       mime_type, file_size
                   ) VALUES (1, 'telegram-video-id', 'unique-id', 'video/mp4', 5)'''
            )
            conn.commit()
            conn.close()
            bot = Mock()
            bot.get_file.return_value.file_path = 'videos/problem.mp4'
            bot.download_file.return_value = b'video'

            with (
                patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
                patch.object(kpi_web, 'get_user', return_value=user(0)),
                patch.object(kpi_web, 'DB_PATH', str(db_path)),
                patch.object(kpi_web, '_notification_bot', return_value=bot),
            ):
                response = self.client.get(
                    '/api/problems/1/video', headers=self.headers,
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, b'video')
        self.assertEqual(response.mimetype, 'video/mp4')
        bot.get_file.assert_called_once_with('telegram-video-id')
        bot.download_file.assert_called_once_with('videos/problem.mp4')

    def test_problem_is_removed_when_telegram_rejects_video(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / 'tasks.db'
            conn = sqlite3.connect(db_path)
            conn.execute(
                '''CREATE TABLE tasks (
                       ID INTEGER PRIMARY KEY AUTOINCREMENT,
                       dtrep TEXT, type TEXT, club TEXT, title TEXT,
                       photo BLOB, desc TEXT, status TEXT, dtfb TEXT,
                       feedback TEXT
                   )'''
            )
            conn.commit()
            conn.close()

            with (
                patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
                patch.object(kpi_web, 'get_user', return_value=user(0)),
                patch.object(kpi_web, 'DB_PATH', str(db_path)),
                patch.object(kpi_web, 'get_clubs', return_value={'Марьино': {}}),
                patch.object(
                    kpi_web, '_send_problem_notification', return_value=None,
                ),
            ):
                response = self.client.post(
                    '/api/problems',
                    headers=self.headers,
                    data={
                        'type': 'Общее обращение',
                        'club': 'Марьино',
                        'title': 'Проблема со звуком',
                        'description': 'Telegram временно недоступен',
                        'video': (BytesIO(b'video'), 'problem.mp4', 'video/mp4'),
                    },
                )

            conn = sqlite3.connect(db_path)
            task_count = conn.execute('SELECT COUNT(*) FROM tasks').fetchone()[0]
            conn.close()

        self.assertEqual(response.status_code, 400)
        self.assertIn('Не удалось сохранить видео', response.get_json()['error'])
        self.assertEqual(task_count, 0)

    def test_problem_video_over_20_mb_is_rejected(self):
        upload = Mock(filename='problem.mp4', mimetype='video/mp4')
        upload.read.return_value = b'x' * (20 * 1024 * 1024 + 1)

        with self.assertRaisesRegex(ValueError, 'не больше 20 МБ'):
            kpi_web._read_problem_video(upload)

        upload.read.assert_called_once_with(20 * 1024 * 1024 + 1)

    def test_problem_follows_existing_solution_and_return_lifecycle(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / 'tasks.db'
            conn = sqlite3.connect(db_path)
            conn.executescript(
                '''CREATE TABLE tasks (
                       ID INTEGER PRIMARY KEY AUTOINCREMENT,
                       dtrep TEXT, type TEXT, club TEXT, title TEXT,
                       photo BLOB, desc TEXT, status TEXT, dtfb TEXT,
                       feedback TEXT
                   );
                   INSERT INTO tasks (
                       dtrep, type, club, title, desc, status
                   ) VALUES (
                       '2026-08-07', 'Ремонт', 'Марьино',
                       'Шлем', 'Не включается', 'В работе'
                   );'''
            )
            conn.commit()
            conn.close()

            with (
                patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
                patch.object(kpi_web, 'get_user', return_value=user(1)),
                patch.object(kpi_web, 'DB_PATH', str(db_path)),
                patch.object(kpi_web, '_send_problem_notification'),
            ):
                solution = self.client.post(
                    '/api/problems/1/solution',
                    headers=self.headers,
                    json={'message': 'Переподключил питание'},
                )

            with (
                patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
                patch.object(kpi_web, 'get_user', return_value=user(0)),
                patch.object(kpi_web, 'DB_PATH', str(db_path)),
                patch.object(kpi_web, '_send_problem_notification'),
            ):
                returned = self.client.post(
                    '/api/problems/1/return',
                    headers=self.headers,
                    json={'message': 'Снова выключился'},
                )

            conn = sqlite3.connect(db_path)
            row = conn.execute(
                'SELECT status, dtfb, feedback FROM tasks WHERE ID=1'
            ).fetchone()
            events = conn.execute(
                'SELECT event_type FROM task_events WHERE task_id=1 ORDER BY id'
            ).fetchall()
            conn.close()

        self.assertEqual(solution.status_code, 200)
        self.assertEqual(returned.status_code, 200)
        self.assertEqual(row[0], 'В работе')
        self.assertIsNone(row[1])
        self.assertIn('Переподключил питание', row[2])
        self.assertIn('Снова выключился', row[2])
        self.assertEqual(events, [('solution',), ('returned',)])

    def test_all_active_users_can_read_every_employee_kpi(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, '_active_employee_logins', return_value=['@one']),
            patch.object(
                kpi_web,
                '_employee_logins_with_month_shifts',
                return_value=['@one'],
            ),
            patch.object(kpi_web, '_employee_metadata', return_value={
                '@one': {
                    'role': 0,
                    'role_name': 'Сотрудник',
                    'clubs': ['Дмитровка'],
                },
            }),
            patch.object(kpi_web, 'calculate_monthly_kpi', return_value=[{
                'login': '@one',
                'nickname': 'Первый',
                'shifts': 2,
                'rank': 1,
                'total_pct': 0.8,
            }]) as calculate,
            patch.object(kpi_web, 'list_penalties', return_value=[]),
            patch.object(kpi_web, 'get_kpi_freshness', return_value={
                'latest_metric_date': '2026-07-14',
                'latest_shift_date': '2026-07-15',
            }),
            patch.object(kpi_web, 'get_month_status', return_value={
                'period_month': '2026-07-01',
                'is_closed': False,
            }),
        ):
            response = self.client.get(
                '/api/kpi?month=2026-07&date=2026-07-15',
                headers=self.headers,
            )
            cached_response = self.client.get(
                '/api/kpi?month=2026-07&date=2026-07-15',
                headers=self.headers,
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(cached_response.status_code, 200)
        self.assertEqual(response.get_json()['employees'][0]['login'], '@one')
        self.assertEqual(response.get_json(), cached_response.get_json())
        self.assertEqual(response.get_json()['date'], '2026-07-15')
        self.assertEqual(calculate.call_count, 3)
        self.assertTrue(all(
            call.kwargs['employee_logins'] == ['@one']
            for call in calculate.call_args_list
        ))

    def test_personal_kpi_contains_explanation_and_shift_pace(self):
        row = {
            'login': '@tester',
            'nickname': 'Тестер',
            'shifts': 2,
            'weighted_shifts': 3,
            'rank': 1,
            'zone': '🟢',
            'average_pct': 0.9,
            'total_pct': 0.8,
            'weighted_pct': 0.6,
            'reviews': 1,
            'reviews_plan_per_shift': 0.25,
            'reviews_target': 0.5,
            'reviews_pct': 2,
            'forms': 1,
            'forms_plan_per_shift': 1,
            'forms_target': 2,
            'forms_pct': 0.5,
            'extensions': 0,
            'extensions_plan_per_shift': 0.25,
            'extensions_target': 0.5,
            'extensions_pct': 0,
            'certificates': 0,
            'certificates_plan_per_shift': 125,
            'certificates_target': 250,
            'certificates_pct': 0,
            'subscriptions': 0,
            'subscriptions_plan_per_shift': 250,
            'subscriptions_target': 500,
            'subscriptions_pct': 0,
            'initiatives': 1,
            'initiatives_pct': 0.1,
            'penalty_impact': 0.1,
            'stream_bonus': 0.05,
        }
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(
                kpi_web,
                '_active_employee_logins',
                return_value=['@tester'],
            ),
            patch.object(
                kpi_web,
                '_employee_logins_with_month_shifts',
                return_value=['@tester'],
            ),
            patch.object(
                kpi_web,
                'calculate_monthly_kpi',
                side_effect=[[row], [], []],
            ),
            patch.object(kpi_web, '_employee_metadata', return_value={
                '@tester': {
                    'role': 0,
                    'role_name': 'Сотрудник',
                    'clubs': ['Дмитровка', 'Марьино'],
                },
            }),
            patch.object(kpi_web, 'list_penalties', return_value=[]),
            patch.object(kpi_web, 'get_month_status', return_value={
                'period_month': '2026-07-01',
                'is_closed': False,
            }),
            patch.object(kpi_web, 'get_kpi_freshness', return_value={
                'latest_metric_date': '2026-07-14',
                'latest_shift_date': '2026-07-15',
            }),
        ):
            response = self.client.get(
                '/api/kpi?month=2026-07&date=2026-07-15',
                headers=self.headers,
            )

        payload = response.get_json()
        explanation = payload['my_kpi']['explanation']
        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload['my_kpi']['login'], '@tester')
        self.assertEqual(explanation['pace']['shifts'], 2)
        self.assertEqual(explanation['pace']['projected_pct'], 0.8)
        self.assertAlmostEqual(explanation['pace']['gap_to_green_pct'], 0.1)
        self.assertEqual(explanation['metrics'][1]['needed'], 1)
        initiative = next(
            item
            for item in explanation['metrics']
            if item['key'] == 'initiatives'
        )
        self.assertEqual(initiative['bonus_per_item'], 0.1)
        self.assertEqual(initiative['contribution_pct'], 0.1)
        self.assertEqual(
            payload['freshness']['latest_metric_date'],
            '2026-07-14',
        )
        self.assertEqual(payload['my_kpi']['clubs'], ['Дмитровка', 'Марьино'])

    def test_attention_reasons_follow_confirmed_manager_rules(self):
        reasons = kpi_web._attention_reasons({
            'zone': '🔴',
            'penalties': 1,
            'kpi_change_7d': -0.10,
            'shifts': 3,
            'reviews': 0,
            'forms': 0,
            'extensions': 0,
            'certificates': 0,
            'subscriptions': 0,
            'initiatives': 0,
        })

        self.assertEqual(
            [reason['key'] for reason in reasons],
            ['red_zone', 'penalty', 'drop', 'no_kpi'],
        )

    def test_kpi_payload_marks_seven_day_drop_for_manager(self):
        current = {
            'login': '@one',
            'nickname': 'Первый',
            'shifts': 2,
            'weighted_shifts': 2,
            'rank': 1,
            'zone': '🟡',
            'average_pct': 0.8,
            'total_pct': 0.7,
            'weighted_pct': 0.7,
            'reviews': 1,
            'reviews_pct': 1,
            'forms': 1,
            'forms_pct': 0.5,
            'extensions': 1,
            'extensions_pct': 1,
            'certificates': 0,
            'certificates_pct': 0,
            'subscriptions': 0,
            'subscriptions_pct': 0,
            'initiatives': 1,
            'initiatives_pct': 0.1,
            'penalties': 0,
            'penalty_impact': 0,
            'stream_bonus': 0,
        }
        previous_day = dict(current, total_pct=0.72)
        previous_week = dict(current, total_pct=0.85)
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(2)),
            patch.object(kpi_web, '_active_employee_logins', return_value=['@one']),
            patch.object(
                kpi_web,
                'calculate_monthly_kpi',
                side_effect=[[current], [previous_day], [previous_week]],
            ),
            patch.object(kpi_web, '_employee_metadata', return_value={
                '@one': {
                    'role': 0,
                    'role_name': 'Сотрудник',
                    'clubs': ['Дмитровка'],
                },
            }),
            patch.object(kpi_web, 'list_penalties', return_value=[]),
            patch.object(kpi_web, 'get_month_status', return_value={
                'period_month': '2026-07-01',
                'is_closed': False,
            }),
            patch.object(kpi_web, 'get_kpi_freshness', return_value={
                'latest_metric_date': '2026-07-15',
                'latest_shift_date': '2026-07-15',
            }),
        ):
            response = self.client.get(
                '/api/kpi?month=2026-07&date=2026-07-15',
                headers=self.headers,
            )

        employee = response.get_json()['employees'][0]
        self.assertAlmostEqual(employee['kpi_change_7d'], -0.15)
        self.assertTrue(employee['needs_attention'])
        self.assertEqual(
            [reason['key'] for reason in employee['attention_reasons']],
            ['drop'],
        )
        self.assertEqual(employee['clubs'], ['Дмитровка'])

    def test_employee_metadata_contains_every_club_in_selected_month(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / 'metadata.db'
            conn = sqlite3.connect(db_path)
            conn.executescript(
                '''
                CREATE TABLE users (
                    login TEXT,
                    first_name TEXT,
                    second_name TEXT,
                    status INTEGER
                );
                CREATE TABLE shifts (
                    shift_login TEXT,
                    shift_first_name TEXT,
                    shift_second_name TEXT,
                    dt_shift TEXT,
                    club TEXT
                );
                INSERT INTO users VALUES ('@one', 'One', 'User', 0);
                INSERT INTO shifts VALUES (
                    '@one', 'One', 'User', '2026-07-05', 'Дмитровка'
                );
                INSERT INTO shifts VALUES (
                    '@one', 'One', 'User', '2026-07-12', 'Марьино'
                );
                INSERT INTO shifts VALUES (
                    '@one', 'One', 'User', '2026-08-01', 'Каширка'
                );
                '''
            )
            conn.commit()
            conn.close()

            with patch.object(kpi_web, 'DB_PATH', str(db_path)):
                metadata = kpi_web._employee_metadata(
                    ['@one'],
                    '2026-07-01',
                )

        self.assertEqual(metadata['@one']['role'], 0)
        self.assertEqual(
            metadata['@one']['clubs'],
            ['Дмитровка', 'Марьино'],
        )

    def test_metric_details_are_available_to_employee(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, '_active_employee_logins', return_value=['@one']),
            patch.object(kpi_web, 'get_metric_entries', return_value=[{
                'id': 1,
                'date': '2026-07-05',
                'value': 1,
                'description': 'Отзыв',
                'club': None,
                'status': None,
            }]) as get_entries,
        ):
            response = self.client.get(
                '/api/kpi/details?month=2026-07&date=2026-07-15'
                '&employee_login=@one&metric=reviews',
                headers=self.headers,
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()['entries'][0]['date'], '2026-07-05')
        get_entries.assert_called_once_with(
            '@one',
            '2026-07-01',
            'reviews',
            period_end='2026-07-15',
        )

    def test_dynamic_hashtag_details_are_available_to_employee(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, '_active_employee_logins', return_value=['@one']),
            patch.object(kpi_web, 'get_hashtag_entries', return_value=[{
                'id': 1,
                'date': '2026-07-05',
                'value': None,
                'value_unit': None,
                'description': 'Автосим',
                'club': 'Марьино',
                'status': None,
            }]) as get_entries,
        ):
            response = self.client.get(
                '/api/kpi/details?month=2026-07&date=2026-07-15'
                '&employee_login=@one&metric=hashtag:%23автосим',
                headers=self.headers,
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()['entries'][0]['date'], '2026-07-05')
        get_entries.assert_called_once_with(
            '@one',
            '2026-07-01',
            '#автосим',
            period_end='2026-07-15',
            db_path=kpi_web.DB_PATH,
        )

    def test_daily_analytics_are_available_to_employee(self):
        kpi_web._analytics_cache.clear()
        analytics_row = {
            'login': '@one',
            'nickname': 'Первый',
            'total_pct': 0.8,
            'weighted_pct': 0.6,
            'rank': 1,
            'zone': '🟢',
            'shifts': 2,
            'weighted_shifts': 3,
            'reviews': 1,
            'forms': 2,
            'extensions': 1,
            'certificates': 0,
            'subscriptions': 0,
            'initiatives': 1,
            'penalties': 0,
            'stream': False,
        }
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
            patch.object(kpi_web, '_active_employee_logins', return_value=['@one']),
            patch.object(
                kpi_web,
                '_employee_logins_with_month_shifts',
                return_value=['@one'],
            ),
            patch.object(kpi_web, 'initialize_kpi_calculation_schema'),
            patch.object(
                kpi_web,
                'calculate_daily_kpi_series',
                return_value=[
                    {'date': '2026-07-01', 'employees': [analytics_row]},
                    {'date': '2026-07-02', 'employees': [analytics_row]},
                ],
            ) as calculate,
        ):
            response = self.client.get(
                '/api/kpi/analytics?mode=daily&month=2026-07&date=2026-07-02',
                headers=self.headers,
            )
            cached_response = self.client.get(
                '/api/kpi/analytics?mode=daily&month=2026-07&date=2026-07-02',
                headers=self.headers,
            )

        payload = response.get_json()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(cached_response.status_code, 200)
        self.assertEqual(len(payload['points']), 2)
        self.assertEqual(payload['points'][-1]['team']['kpi'], 0.8)
        self.assertEqual(payload['employees'][0]['login'], '@one')
        calculate.assert_called_once_with(
            '2026-07-01',
            '2026-07-02',
            employee_logins=['@one'],
            ensure_schema=False,
        )

    def test_employee_cannot_add_penalty(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(0)),
        ):
            response = self.client.post(
                '/api/penalties',
                headers=self.headers,
                json={
                    'month': '2026-07',
                    'employee_login': '@one',
                    'reason': 'Причина',
                },
            )

        self.assertEqual(response.status_code, 403)

    def test_manager_adds_fixed_ten_percent_penalty(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(2)),
            patch.object(kpi_web, '_active_employee_logins', return_value=['@one']),
            patch.object(kpi_web, 'add_penalty', return_value=42) as add_penalty,
        ):
            response = self.client.post(
                '/api/penalties',
                headers=self.headers,
                json={
                    'month': '2026-07',
                    'employee_login': '@one',
                    'reason': 'Опоздание',
                },
            )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.get_json()['impact_pct'], 0.10)
        add_penalty.assert_called_once_with(
            '@one',
            '2026-07-01',
            'Опоздание',
            '@tester',
            source='telegram_mini_app',
        )

    def test_settings_are_forbidden_to_manager(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(2)),
        ):
            response = self.client.get(
                '/api/kpi/settings?month=2026-07',
                headers=self.headers,
            )

        self.assertEqual(response.status_code, 403)

    def test_owner_can_save_kpi_settings(self):
        saved = {
            'period_month': '2026-08-01',
            'metrics': [],
            'clubs': [],
        }
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(3)),
            patch.object(
                kpi_web,
                'save_kpi_settings',
                return_value=saved,
            ) as save_settings,
        ):
            response = self.client.put(
                '/api/kpi/settings',
                headers=self.headers,
                json={
                    'month': '2026-08',
                    'metrics': {
                        'Отзывы': 0.25,
                        'Инициативы': 0.10,
                    },
                    'clubs': {
                        'Дмитровка': {
                            'weekday_weight': 0.75,
                            'weekend_weight': 2,
                        },
                    },
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()['period_month'], '2026-08-01')
        save_settings.assert_called_once_with(
            '2026-08-01',
            {
                'Отзывы': 0.25,
                'Инициативы': 0.10,
            },
            {
                'Дмитровка': (0.75, 2),
            },
            '@tester',
        )

    def test_owner_can_save_custom_goal(self):
        settings = {
            'period_month': '2026-08-01', 'metrics': [], 'clubs': [],
            'custom_goals': [],
        }
        payload = {
            'month': '2026-08', 'name': 'Повторы', 'hashtag': '#повторы',
            'calculation_type': 'per_unit', 'audience': 'physical',
            'unit_label': 'шт.', 'value': 1, 'contribution_pct': 0.1,
            'min_profile_shifts': None,
        }
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(3)),
            patch.object(
                kpi_web, 'save_custom_goal', return_value='goal-1',
            ) as save_goal,
            patch.object(kpi_web, 'get_kpi_settings', return_value=settings),
        ):
            response = self.client.post(
                '/api/kpi/goals', headers=self.headers, json=payload,
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()['saved_goal_key'], 'goal-1')
        expected = dict(payload)
        expected.pop('month')
        save_goal.assert_called_once_with('2026-08-01', expected, '@tester')

    def test_manager_cannot_save_custom_goal(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(2)),
        ):
            response = self.client.post(
                '/api/kpi/goals', headers=self.headers,
                json={'month': '2026-08'},
            )
        self.assertEqual(response.status_code, 403)

    def test_month_close_preview_only_warns_about_detected_problems(self):
        rows = [
            {
                'login': '@red',
                'nickname': 'Red',
                'shifts': 2,
                'total_pct': 0.4,
                'zone': '🔴',
                'penalties': 1,
                'reviews': 0,
                'forms': 0,
                'extensions': 0,
                'certificates': 0,
                'subscriptions': 0,
                'initiatives': 0,
            },
            {
                'login': '@idle',
                'nickname': 'Idle',
                'shifts': 0,
                'total_pct': 0,
                'zone': '⚪',
                'penalties': 0,
            },
        ]
        with (
            patch.object(kpi_web, '_active_employee_logins', return_value=[
                '@red',
                '@idle',
            ]),
            patch.object(
                kpi_web,
                'calculate_monthly_kpi',
                return_value=rows,
            ),
            patch.object(kpi_web, '_employee_metadata', return_value={}),
        ):
            preview = kpi_web._month_close_preview('2026-07-01')

        self.assertEqual(preview['summary']['active_employees'], 2)
        self.assertEqual(preview['summary']['participants'], 1)
        self.assertEqual(
            [warning['key'] for warning in preview['warnings']],
            ['no_shifts', 'no_kpi', 'red_zone', 'penalties'],
        )

    def test_manager_closes_month_with_server_snapshot(self):
        snapshot = {
            'period_month': '2026-07-01',
            'date': '2026-07-31',
            'generated_at': '2026-07-31T20:00:00+00:00',
            'summary': {'participants': 1},
            'warnings': [],
            'employees': [],
        }
        saved_status = {
            'period_month': '2026-07-01',
            'is_closed': True,
            'updated_by_login': '@tester',
            'updated_at': '2026-07-31 23:00:00',
            'snapshot': snapshot,
            'snapshot_created_at': '2026-07-31 23:00:00',
        }
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(2)),
            patch.object(
                kpi_web,
                '_month_close_preview',
                return_value=snapshot,
            ),
            patch.object(
                kpi_web,
                'set_month_status',
                return_value=saved_status,
            ) as set_status,
        ):
            response = self.client.post(
                '/api/month-status',
                headers=self.headers,
                json={'month': '2026-07', 'is_closed': True},
            )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()['is_closed'])
        set_status.assert_called_once_with(
            '2026-07-01',
            True,
            '@tester',
            snapshot=snapshot,
        )

    def test_manager_can_reopen_month_without_rebuilding_snapshot(self):
        with (
            patch.object(kpi_web, 'TELEGRAM_API_KEY', BOT_TOKEN),
            patch.object(kpi_web, 'get_user', return_value=user(2)),
            patch.object(kpi_web, '_month_close_preview') as preview,
            patch.object(
                kpi_web,
                'set_month_status',
                return_value={
                    'period_month': '2026-07-01',
                    'is_closed': False,
                },
            ) as set_status,
        ):
            response = self.client.post(
                '/api/month-status',
                headers=self.headers,
                json={'month': '2026-07', 'is_closed': False},
            )

        self.assertEqual(response.status_code, 200)
        preview.assert_not_called()
        set_status.assert_called_once_with(
            '2026-07-01',
            False,
            '@tester',
            snapshot=None,
        )


if __name__ == '__main__':
    unittest.main()
